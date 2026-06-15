"""
Gradio Template Form Component

Handles dynamic form rendering, area selection, ID lookup, and bulk import.
"""
import gradio as gr
import io
import logging
import polars as pl
from functools import partial
from pathlib import Path
from typing import Any

from app.services.registry import TemplateConfig
from app.services.excel_parser import list_sheet_names, read_template_sheet, resolve_sheet_name
from app.services.export_naming import build_export_filename
from app.services.excel_print import persist_export_file, primary_print_area, show_print_dialog
from app.services.section_detector import (
    DetectedArea,
    SectionConfig,
    calculate_next_area,
    detect_multi_areas,
    parse_area_range,
)
from app.services.paste_parse_config import (
    load_paste_parse_config,
    map_sheet_row_from_paste_config,
    parse_text_with_config,
    default_input_area_for_template,
    read_input_area_headers,
    id_target_field_from_config,
    DEFAULT_FIELDS_PER_ROW,
)
from app.services.google_sheets import (
    fetch_all_rows,
    GoogleSheetsError,
    invalidate_sheet_cache,
    lookup_row_by_id,
)
from app.services.gemma4_field_matcher import create_field_matcher
from app.services.import_history import (
    load_import_history, mark_as_processed, mark_as_trash,
    unmark_ids, get_import_stats, clear_history, save_import_history,
)

logger = logging.getLogger(__name__)

MAX_FORM_FIELDS = 40
MAX_IMPORT_PREVIEW_ROWS = 1000
HIDE_PROGRESS = {"show_progress": "hidden"}
ENTRY_MODE_ID_AUTO = "ID Auto"
ENTRY_MODE_MANUAL = "Manual"


def _get_template_id(template: TemplateConfig | dict[str, Any] | None) -> str | None:
    """Resolve template id from TemplateConfig or Gradio state dict."""
    if template is None:
        return None
    if isinstance(template, dict):
        template_id = template.get("id")
        return str(template_id) if template_id else None
    return getattr(template, "id", None)


def _normalize_preview_rows(preview_data: Any) -> list[list[Any]]:
    """Convert Gradio Dataframe input (pandas, etc.) to list of row lists."""
    if preview_data is None:
        return []

    import pandas as pd

    if isinstance(preview_data, pd.DataFrame):
        return [] if preview_data.empty else preview_data.values.tolist()

    import numpy as np

    if isinstance(preview_data, np.ndarray):
        return [] if preview_data.size == 0 else preview_data.tolist()

    try:
        import polars as pl

        if isinstance(preview_data, pl.DataFrame):
            return [] if preview_data.is_empty() else preview_data.rows()
    except ImportError:
        pass

    if isinstance(preview_data, list):
        if not preview_data:
            return []
        if isinstance(preview_data[0], list):
            return preview_data
        return [preview_data]

    return []


def _is_row_selected(row: list[Any]) -> bool:
    """Check whether the first column (checkbox) is selected."""
    if not row:
        return False
    value = row[0]
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes"}
    return bool(value)


def _empty_import_selection() -> dict[str, Any]:
    return {"ids": [], "index": 0}


def _selected_import_ids(preview_rows: list[list[Any]]) -> list[str]:
    """Preserve preview-table order while collecting checked row IDs."""
    ids: list[str] = []
    seen: set[str] = set()
    for row in preview_rows:
        if not _is_row_selected(row) or len(row) < 2:
            continue
        id_val = str(row[1]).strip()
        if not id_val or id_val in seen:
            continue
        ids.append(id_val)
        seen.add(id_val)
    return ids


def _build_sheet_row_cache(
    df: pl.DataFrame,
    id_column: str,
) -> dict[str, dict[str, str]]:
    if df.height == 0 or id_column not in df.columns:
        return {}
    cache: dict[str, dict[str, str]] = {}
    for record in df.iter_rows(named=True):
        id_val = str(record.get(id_column, "")).strip()
        if not id_val:
            continue
        cache[id_val] = {
            str(key): "" if value is None else str(value)
            for key, value in record.items()
        }
    return cache


def _map_sheet_row_to_form_fields(
    template_id: str,
    headers: list[str],
    sheet_row: dict[str, str],
    base_row: dict[str, str],
    id_value: str | None = None,
) -> dict[str, str]:
    paste_config = load_paste_parse_config(template_id)
    if not paste_config:
        return _row_values_for_headers(headers, base_row)
    matched = map_sheet_row_from_paste_config(sheet_row, paste_config)
    id_field = id_target_field_from_config(paste_config)
    if id_field and id_field in headers and id_value:
        matched[id_field] = id_value
    return _merge_mapped_into_row(headers, base_row, matched)


def _import_preview_sync_outputs(
    template: TemplateConfig,
    form_data: list[dict[str, str]],
    selection_state: dict[str, Any],
    sheet_cache: dict[str, dict[str, str]],
    entry_mode: str,
) -> tuple[Any, dict[str, Any], tuple]:
    headers = get_form_field_headers(template)
    if not headers:
        return gr.update(), selection_state, tuple(_empty_field_updates())
    ids = selection_state.get("ids", [])
    if not ids:
        return gr.update(), _empty_import_selection(), tuple(_empty_field_updates())
    index = selection_state.get("index", 0)
    index = max(0, min(index, len(ids) - 1))
    selection_state = {"ids": ids, "index": index}
    row_choices = _row_choices_for_form_data(len(ids))
    selected_label = row_choices[index]
    id_value = ids[index]
    sheet_row = sheet_cache.get(id_value)
    if not sheet_row:
        mapped_row = {header: "" for header in headers}
    else:
        base_idx = 0
        if form_data:
            base_idx = min(base_idx, len(form_data) - 1)
        base_row = form_data[base_idx] if form_data else {header: "" for header in headers}
        mapped_row = _map_sheet_row_to_form_fields(
            template.id,
            headers,
            sheet_row,
            base_row,
            id_value=id_value,
        )
    field_updates = tuple(_field_updates_for_row(headers, mapped_row))
    row_selector_update = _row_selector_update(
        row_choices,
        selected_label,
        entry_mode=entry_mode,
    )
    return row_selector_update, selection_state, field_updates


def _format_last_import(last_import: Any) -> str:
    """Format last_import timestamp for display, falling back on parse errors."""
    if not last_import:
        return "从未"
    from datetime import datetime

    try:
        dt = datetime.fromisoformat(str(last_import).replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        return str(last_import)


def _field_row_count(fields_per_row: int = DEFAULT_FIELDS_PER_ROW) -> int:
    return (MAX_FORM_FIELDS + fields_per_row - 1) // fields_per_row


# UI row layout is fixed at build time; row updates must use the same grouping.
FORM_LAYOUT_FIELDS_PER_ROW = DEFAULT_FIELDS_PER_ROW
FORM_ROW_COUNT = _field_row_count(FORM_LAYOUT_FIELDS_PER_ROW)
_NO_CHANGE_FIELD_UPDATES = tuple(gr.update() for _ in range(MAX_FORM_FIELDS))


def form_refresh_output_count() -> int:
    """Expected tuple length for refresh_data_entry_form / form_refresh_outputs."""
    return 6 + FORM_ROW_COUNT + MAX_FORM_FIELDS


def _form_field_output_components(
    field_rows: list,
    form_field_boxes: list,
) -> list:
    """Gradio outputs must follow component creation order: row, its fields, next row, ..."""
    outputs = []
    for row_idx, field_row in enumerate(field_rows):
        outputs.append(field_row)
        row_start = row_idx * FORM_LAYOUT_FIELDS_PER_ROW
        row_end = min(row_start + FORM_LAYOUT_FIELDS_PER_ROW, MAX_FORM_FIELDS)
        for index in range(row_start, row_end):
            outputs.append(form_field_boxes[index])
    return outputs


def _interleaved_row_field_updates(
    row_updates: list[gr.update],
    field_updates: list[gr.update],
) -> list[gr.update]:
    """Pair row visibility updates with the field slots created inside each row."""
    updates = []
    for row_idx, row_update in enumerate(row_updates):
        updates.append(row_update)
        row_start = row_idx * FORM_LAYOUT_FIELDS_PER_ROW
        row_end = min(row_start + FORM_LAYOUT_FIELDS_PER_ROW, MAX_FORM_FIELDS)
        for index in range(row_start, row_end):
            updates.append(field_updates[index])
    return updates


def split_interleaved_field_refresh_updates(
    interleaved: list,
) -> tuple[list, list]:
    """Split refresh tail (after fields_container) back into row and field update lists."""
    row_updates = []
    field_updates: list = [gr.update() for _ in range(MAX_FORM_FIELDS)]
    pos = 0
    for row_idx in range(FORM_ROW_COUNT):
        row_updates.append(interleaved[pos])
        pos += 1
        row_start = row_idx * FORM_LAYOUT_FIELDS_PER_ROW
        row_end = min(row_start + FORM_LAYOUT_FIELDS_PER_ROW, MAX_FORM_FIELDS)
        for index in range(row_start, row_end):
            field_updates[index] = interleaved[pos]
            pos += 1
    return row_updates, field_updates


def get_fields_per_row(template_id: str) -> int:
    """Load fields-per-row layout from paste config (default 7)."""
    paste_config = load_paste_parse_config(template_id)
    if paste_config is None:
        return DEFAULT_FIELDS_PER_ROW
    return paste_config.fields_per_row


def resolve_default_sheet_name(
    template: TemplateConfig,
    workbook_sheets: list[str],
) -> str | None:
    """Prefer paste-config worksheet, then template default, then first sheet."""
    paste_config = load_paste_parse_config(template.id)
    preferred: list[str] = []
    if paste_config and paste_config.worksheet:
        preferred.append(paste_config.worksheet)
    if template.sheet_name:
        preferred.append(template.sheet_name)
    for name in preferred:
        target = name.strip().lower()
        for sheet in workbook_sheets:
            if sheet.strip().lower() == target:
                return sheet
    return workbook_sheets[0] if workbook_sheets else None


def _resolve_input_area_range(
    template: TemplateConfig,
    sheet_name: str | None,
) -> str | None:
    """Resolve the template input area from paste sections or template defaults."""
    paste_config = load_paste_parse_config(template.id)
    if paste_config and paste_config.sections:
        area_range = str(paste_config.sections[0].get("input_area", "")).strip()
        if area_range:
            return area_range
    preferred_sheet = sheet_name or template.sheet_name or None
    # Respect an explicit sheet selection; only default via paste config when unset.
    if not sheet_name and paste_config and paste_config.worksheet:
        preferred_sheet = paste_config.worksheet
    return default_input_area_for_template(
        Path(template.file_path),
        preferred_sheet,
        template.data_start_row,
    )


def get_form_field_headers(
    template: TemplateConfig,
    sheet_name: str | None = None,
) -> list[str]:
    """Load form field names from the Excel template columns above the input area."""
    area_range = _resolve_input_area_range(template, sheet_name)
    if not area_range:
        return []
    preferred_sheet = sheet_name or template.sheet_name or None
    paste_config = load_paste_parse_config(template.id)
    # When no sheet is selected yet, fall back to the paste-config worksheet.
    if not sheet_name and paste_config and paste_config.worksheet:
        preferred_sheet = paste_config.worksheet
    try:
        return read_input_area_headers(
            Path(template.file_path),
            preferred_sheet,
            area_range,
            header_row=template.header_row,
        )
    except Exception as exc:
        logger.error(f"Failed to read template area headers: {exc}")
        return []


def read_area_form_values(
    workbook_path: Path,
    sheet_name: str,
    area_range: str,
    headers: list[str],
) -> dict[str, str]:
    """Read template cells in an input area and map values to column headers."""
    from openpyxl import load_workbook

    from app.services.excel_parser import format_cell_display, resolve_sheet_name
    from app.services.section_detector import parse_area_range

    coords = parse_area_range(area_range)
    resolved_sheet = resolve_sheet_name(workbook_path, sheet_name)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[resolved_sheet]
        values: dict[str, str] = {}
        area_rows = coords.end_row - coords.start_row + 1
        area_cols = coords.end_col - coords.start_col + 1

        if area_rows == 1:
            for index, col in enumerate(range(coords.start_col, coords.end_col + 1)):
                if index >= len(headers):
                    break
                cell = ws.cell(coords.start_row, col)
                values[headers[index]] = format_cell_display(cell.value)
            return values

        pos = 0
        for row_idx in range(coords.start_row, coords.end_row + 1):
            for col_idx in range(coords.start_col, coords.end_col + 1):
                if pos >= len(headers):
                    return values
                cell = ws.cell(row_idx, col_idx)
                values[headers[pos]] = format_cell_display(cell.value)
                pos += 1
        return values
    finally:
        wb.close()


def write_area_form_values(
    worksheet,
    area_range: str,
    headers: list[str],
    values: dict[str, str],
) -> None:
    """Write header-mapped values into template cells for an input area."""
    from app.services.section_detector import parse_area_range

    coords = parse_area_range(area_range)
    area_rows = coords.end_row - coords.start_row + 1

    if area_rows == 1:
        for index, col in enumerate(range(coords.start_col, coords.end_col + 1)):
            if index >= len(headers):
                break
            header = headers[index]
            cell = worksheet.cell(coords.start_row, col)
            if cell.data_type == "f":
                continue
            text = values.get(header, "")
            cell.value = text if text else None
        return

    pos = 0
    for row_idx in range(coords.start_row, coords.end_row + 1):
        for col_idx in range(coords.start_col, coords.end_col + 1):
            if pos >= len(headers):
                return
            header = headers[pos]
            cell = worksheet.cell(row_idx, col_idx)
            if cell.data_type != "f":
                text = values.get(header, "")
                cell.value = text if text else None
            pos += 1


def build_export_workbook_bytes(
    template: TemplateConfig,
    sheet_name: str,
    form_data: list[dict[str, str]],
) -> bytes:
    """Copy the template workbook and write all form rows into input areas."""
    from openpyxl import load_workbook

    workbook_path = Path(template.file_path)
    resolved_sheet = resolve_sheet_name(workbook_path, sheet_name)
    headers = get_form_field_headers(template, sheet_name)
    if not headers:
        raise ValueError("未找到表头")
    base_area = _resolve_input_area_range(template, sheet_name)
    if not base_area:
        raise ValueError("未找到输入区域")
    move_to = "down"
    offset = 1
    paste_config = load_paste_parse_config(template.id)
    if paste_config and paste_config.sections:
        section = paste_config.sections[0]
        move_to = str(section.get("move_to", "down"))
        offset = int(section.get("offset", 1))
    rows = form_data if form_data else [{header: "" for header in headers}]
    wb = load_workbook(workbook_path)
    try:
        ws = wb[resolved_sheet]
        for row_idx, row_dict in enumerate(rows):
            area_range = base_area
            if row_idx > 0:
                area_range = calculate_next_area(base_area, move_to, offset * row_idx)
            row_values = _row_values_for_headers(headers, row_dict)
            write_area_form_values(ws, area_range, headers, row_values)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    finally:
        wb.close()


def _merge_field_boxes_into_form_data(
    template: TemplateConfig,
    sheet_name: str | None,
    form_data: list[dict[str, str]],
    row_selector_value: str | None,
    field_values: tuple[str, ...],
) -> list[dict[str, str]]:
    """Merge visible textbox values into form_data for the selected row."""
    headers = get_form_field_headers(template, sheet_name)
    if not headers:
        return list(form_data)
    merged = [dict(row) for row in form_data] if form_data else []
    if not merged:
        merged = [{header: "" for header in headers}]
    target_idx = _resolve_row_index(row_selector_value, len(merged))
    row_dict = dict(merged[target_idx])
    for index, header in enumerate(headers):
        if index < len(field_values):
            row_dict[header] = field_values[index] or ""
    merged[target_idx] = _row_values_for_headers(headers, row_dict)
    return merged


def _hidden_field_slot_update() -> gr.update:
    return gr.update(visible=False, value="", label="")


def _empty_field_updates() -> list[gr.update]:
    return [_hidden_field_slot_update() for _ in range(MAX_FORM_FIELDS)]


def _empty_row_updates() -> list[gr.update]:
    return [gr.update(visible=False) for _ in range(FORM_ROW_COUNT)]


def _build_row_updates(headers: list[str]) -> list[gr.update]:
    """Show gr.Row containers when any child field slot in that row is active."""
    updates: list[gr.update] = []
    for row_idx in range(FORM_ROW_COUNT):
        row_start = row_idx * FORM_LAYOUT_FIELDS_PER_ROW
        row_end = min(row_start + FORM_LAYOUT_FIELDS_PER_ROW, MAX_FORM_FIELDS)
        has_visible = any(index < len(headers) for index in range(row_start, row_end))
        updates.append(gr.update(visible=has_visible))
    return updates


def _first_section_config(template_id: str) -> SectionConfig | None:
    """Build section config from the first paste YAML section entry."""
    paste_config = load_paste_parse_config(template_id)
    if not paste_config or not paste_config.sections:
        return None
    section_dict = paste_config.sections[0]
    input_area = str(section_dict.get("input_area", "")).strip()
    if not input_area:
        return None
    move_to = str(section_dict.get("move_to", "down")).strip().lower()
    if move_to not in {"down", "up", "left", "right"}:
        move_to = "down"
    try:
        offset = int(section_dict.get("offset", 1))
    except (TypeError, ValueError):
        offset = 1
    if offset <= 0:
        offset = 1
    return SectionConfig(input_area=input_area, move_to=move_to, offset=offset)


def _detect_form_areas(
    template: TemplateConfig,
    sheet_name: str,
    base_area: str,
) -> list[DetectedArea]:
    """Scan the worksheet for repeated input areas using paste section rules."""
    workbook_path = Path(template.file_path)
    resolved_sheet = resolve_sheet_name(workbook_path, sheet_name)
    section_config = _first_section_config(template.id)
    if section_config:
        try:
            areas = detect_multi_areas(workbook_path, resolved_sheet, section_config)
            if areas:
                return areas
        except Exception as exc:
            logger.error(f"Area detection failed: {exc}")
    coords = parse_area_range(base_area)
    return [DetectedArea(index=1, area=base_area, coords=coords, has_data=True)]


def _align_form_data_to_areas(
    form_data: list[dict[str, str]],
    headers: list[str],
    area_count: int,
) -> list[dict[str, str]]:
    """Pad or trim in-memory form rows to match the detected area count."""
    aligned = [dict(row) for row in form_data[:area_count]]
    while len(aligned) < area_count:
        aligned.append({header: "" for header in headers})
    return aligned


def _load_form_data_for_areas(
    template: TemplateConfig,
    sheet_name: str,
    headers: list[str],
    areas: list[DetectedArea],
) -> list[dict[str, str]]:
    """Read template cell values for each detected input area."""
    workbook_path = Path(template.file_path)
    rows: list[dict[str, str]] = []
    for area in areas:
        try:
            row_values = read_area_form_values(
                workbook_path,
                sheet_name,
                area.area,
                headers,
            )
        except Exception as exc:
            logger.error(f"Failed to read area {area.area}: {exc}")
            row_values = {header: "" for header in headers}
        rows.append(row_values)
    return rows


def build_row_brief(
    row_values: dict[str, str],
    static_values: dict[str, str],
    headers: list[str],
) -> str:
    """Build dropdown brief: non-empty values that differ from template static cells."""
    parts: list[str] = []
    for header in headers:
        value = str(row_values.get(header, "") or "").strip()
        if not value:
            continue
        static = str(static_values.get(header, "") or "").strip()
        if value == static:
            continue
        parts.append(value)
    return " | ".join(parts)


def format_row_choice_label(
    row_index: int,
    area_range: str | None,
    brief: str,
) -> str:
    """Format a 区域选择 dropdown entry."""
    if area_range and ":" in area_range:
        prefix = area_range
    else:
        prefix = f"Row {row_index}"
    if brief:
        return f"{prefix} — {brief}"
    return prefix


def _row_choices_for_areas(
    areas: list[DetectedArea],
    form_rows: list[dict[str, str]],
    static_values: dict[str, str],
    headers: list[str],
) -> list[str]:
    choices: list[str] = []
    for index, area in enumerate(areas):
        row_values = form_rows[index] if index < len(form_rows) else {}
        brief = build_row_brief(row_values, static_values, headers)
        choices.append(format_row_choice_label(area.index, area.area, brief))
    return choices


def _load_static_template_row_values(
    template: TemplateConfig,
    sheet_name: str,
    headers: list[str],
    area_range: str,
) -> dict[str, str]:
    """Template defaults from the configured input area on disk."""
    try:
        return read_area_form_values(
            Path(template.file_path),
            sheet_name,
            area_range,
            headers,
        )
    except Exception as exc:
        logger.error(f"Failed to read static template row: {exc}")
        return {header: "" for header in headers}


def _row_choices_for_form_data(
    form_data_len: int,
    *,
    areas: list[DetectedArea] | None = None,
    form_rows: list[dict[str, str]] | None = None,
    static_values: dict[str, str] | None = None,
    headers: list[str] | None = None,
) -> list[str]:
    if areas and form_rows is not None and static_values is not None and headers:
        return _row_choices_for_areas(areas, form_rows, static_values, headers)
    return [f"Row {index + 1}" for index in range(form_data_len)]


def _resolve_row_index(row_selector_value: str | None, form_data_len: int) -> int:
    if form_data_len <= 0:
        return 0
    if not row_selector_value:
        return 0
    if row_selector_value.startswith("Row "):
        number_text = row_selector_value[4:].split(" —", 1)[0].strip()
        try:
            return max(0, min(form_data_len - 1, int(number_text) - 1))
        except ValueError:
            pass
    if " — " in row_selector_value:
        prefix = row_selector_value.split(" — ", 1)[0].strip()
        if prefix.startswith("Row "):
            try:
                return max(0, min(form_data_len - 1, int(prefix[4:].strip()) - 1))
            except ValueError:
                pass
    return 0


def _row_values_for_headers(headers: list[str], row_dict: dict[str, str]) -> dict[str, str]:
    return {header: row_dict.get(header, "") for header in headers}


def _merge_mapped_into_row(
    headers: list[str],
    existing_row: dict[str, str],
    mapped: dict[str, str],
) -> dict[str, str]:
    """Overlay mapped data-source values onto an existing form row; keep unmapped template cells."""
    merged = _row_values_for_headers(headers, existing_row)
    for key, value in mapped.items():
        if key not in headers:
            continue
        if value is None:
            continue
        text = str(value).strip()
        if text:
            merged[key] = text
    return merged


def _field_updates_for_row(headers: list[str], row_values: dict[str, str]) -> list[gr.update]:
    updates: list[gr.update] = []
    for index in range(MAX_FORM_FIELDS):
        if index < len(headers):
            header = headers[index]
            updates.append(
                gr.update(
                    label=header,
                    value=row_values.get(header, ""),
                    visible=True,
                    interactive=True,
                )
            )
        else:
            updates.append(_hidden_field_slot_update())
    return updates


def _paste_input_update(
    headers: list[str],
    *,
    entry_mode: str = ENTRY_MODE_ID_AUTO,
    interactive: bool = True,
) -> gr.update:
    count = len(headers)
    manual = entry_mode == ENTRY_MODE_MANUAL
    placeholder = "粘贴 Tab 分隔的一行（可从 Excel 复制）"
    if count:
        if manual:
            placeholder += "；将按 YAML 配置解析并填入表单"
        else:
            placeholder += "；填入「区域选择」当前区域，未选则第 1 个"
    return gr.update(
        value="",
        visible=count > 0 and manual,
        interactive=interactive,
        placeholder=placeholder,
        label=f"粘贴数据（已配置 {count} 个字段）" if count else "粘贴数据",
    )


def _row_selector_update(
    choices: list[str],
    value: str | None,
    *,
    entry_mode: str = ENTRY_MODE_ID_AUTO,
    active: bool = True,
) -> gr.update:
    id_auto = entry_mode == ENTRY_MODE_ID_AUTO
    return gr.update(
        choices=choices,
        value=value,
        visible=id_auto and active,
    )


def on_entry_mode_change(
    entry_mode: str,
    template: TemplateConfig | None,
) -> tuple:
    """Sync entry mode state and toggle row selector vs paste input visibility."""
    headers = get_form_field_headers(template) if template else []
    has_headers = bool(headers)
    id_auto = entry_mode == ENTRY_MODE_ID_AUTO
    return (
        entry_mode,
        gr.update(visible=id_auto and has_headers),
        _paste_input_update(headers, entry_mode=entry_mode),
    )


def _should_parse_paste(paste_text: str) -> bool:
    stripped = (paste_text or "").strip()
    if not stripped:
        return False
    return "\t" in stripped or "\n" in stripped


def _build_field_updates(
    headers: list[str],
    row_values: dict[str, str],
    entry_mode: str = ENTRY_MODE_ID_AUTO,
) -> tuple[list[gr.update], list[gr.update], gr.update]:
    updates = _field_updates_for_row(headers, row_values)
    row_updates = _build_row_updates(headers)
    paste_update = _paste_input_update(headers, entry_mode=entry_mode)
    return updates, row_updates, paste_update


def _build_row_choices_for_template(
    template: TemplateConfig,
    sheet_name: str | None,
    form_data: list[dict[str, str]],
    detected_areas: list[DetectedArea] | None,
) -> list[str]:
    headers = get_form_field_headers(template, sheet_name)
    if not headers:
        return _row_choices_for_form_data(len(form_data))
    area_range = _resolve_input_area_range(template, sheet_name) or ""
    if not area_range and detected_areas:
        area_range = detected_areas[0].area
    static_values = (
        _load_static_template_row_values(template, sheet_name or "", headers, area_range)
        if area_range
        else {header: "" for header in headers}
    )
    return _row_choices_for_form_data(
        len(form_data),
        areas=detected_areas,
        form_rows=form_data,
        static_values=static_values,
        headers=headers,
    )


def _empty_form_session() -> dict[str, Any]:
    return {
        "dirty": False,
        "snapshot": None,
        "pending": None,
        "dialog_step": "",
    }


def _mark_form_dirty(session: dict[str, Any]) -> dict[str, Any]:
    if session.get("dirty"):
        return session
    updated = dict(session)
    updated["dirty"] = True
    return updated


def _capture_form_snapshot(
    template_id: str | None,
    form_data: list[dict[str, str]],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: list[list[Any]] | None,
) -> dict[str, Any]:
    history_payload = None
    if template_id:
        history_payload = load_import_history(template_id).to_dict()
    preview_copy = None
    if import_preview_rows:
        preview_copy = [list(row) for row in import_preview_rows]
    return {
        "form_data": [dict(row) for row in form_data],
        "history": history_payload,
        "import_selection": dict(import_selection or _empty_import_selection()),
        "import_preview_active": bool(import_preview_active),
        "sheet_cache": dict(sheet_cache or {}),
        "import_view": import_view or "unprocessed",
        "import_preview_rows": preview_copy,
    }


def _restore_form_snapshot(template_id: str | None, snapshot: dict[str, Any] | None) -> None:
    if not template_id or not snapshot:
        return
    history_payload = snapshot.get("history")
    if history_payload:
        from app.services.import_history import ImportHistoryConfig

        save_import_history(ImportHistoryConfig.from_dict(history_payload))


def _session_after_clean_refresh(
    session: dict[str, Any],
    snapshot: dict[str, Any],
) -> dict[str, Any]:
    return {
        "dirty": False,
        "snapshot": snapshot,
        "pending": None,
        "dialog_step": "",
    }


def _begin_unsaved_switch_dialog(session: dict[str, Any], pending: dict[str, Any]) -> dict[str, Any]:
    updated = dict(session)
    updated["pending"] = pending
    updated["dialog_step"] = "switch"
    return updated


def _begin_unsaved_save_dialog(session: dict[str, Any]) -> dict[str, Any]:
    updated = dict(session)
    updated["dialog_step"] = "save"
    return updated


def _clear_unsaved_dialog(session: dict[str, Any]) -> dict[str, Any]:
    updated = dict(session)
    updated["pending"] = None
    updated["dialog_step"] = ""
    return updated


def _revert_to_session_snapshot(
    template: TemplateConfig | None,
    session: dict[str, Any],
) -> tuple[list[dict[str, str]], dict[str, Any], dict[str, dict[str, str]], bool, str, list[list[Any]] | None]:
    snapshot = session.get("snapshot") or {}
    template_id = _get_template_id(template)
    _restore_form_snapshot(template_id, snapshot)
    form_data = [dict(row) for row in snapshot.get("form_data", [])]
    import_selection = dict(snapshot.get("import_selection") or _empty_import_selection())
    sheet_cache = dict(snapshot.get("sheet_cache") or {})
    import_preview_active = bool(snapshot.get("import_preview_active"))
    import_view = str(snapshot.get("import_view") or "unprocessed")
    import_preview_rows = snapshot.get("import_preview_rows")
    preview_copy = [list(row) for row in import_preview_rows] if import_preview_rows else None
    return form_data, import_selection, sheet_cache, import_preview_active, import_view, preview_copy


def _unsaved_dialog_updates(step: str) -> tuple[gr.update, gr.update]:
    return (
        gr.update(visible=step == "switch"),
        gr.update(visible=step == "save"),
    )


def _hold_form_refresh_outputs(
    form_data: list[dict[str, str]],
    detected_areas: list,
) -> tuple:
    """Keep the mounted form unchanged while an unsaved-change dialog is open."""
    return (
        gr.update(),
        detected_areas,
        form_data,
        gr.update(),
        gr.update(),
        gr.update(),
        *_interleaved_row_field_updates(_empty_row_updates(), _empty_field_updates()),
    )


def try_sheet_select(
    new_sheet: str | None,
    committed_sheet: str | None,
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    detected_areas: list,
    entry_mode: str,
    session: dict[str, Any],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
) -> tuple:
    """Intercept worksheet picks when the form session has unsaved edits."""
    preview_rows = _normalize_preview_rows(import_preview_rows)
    if (
        session.get("dirty")
        and new_sheet
        and committed_sheet
        and new_sheet != committed_sheet
    ):
        gr.Warning("当前有未保存的更改，切换工作表前请确认")
        pending = {"type": "sheet", "target": new_sheet}
        switch_update, save_update = _unsaved_dialog_updates("switch")
        return (
            committed_sheet,
            _begin_unsaved_switch_dialog(session, pending),
            switch_update,
            save_update,
            *_hold_form_refresh_outputs(form_data, detected_areas),
        )
    refresh = refresh_data_entry_form(
        template, new_sheet, form_data, entry_mode, preserve_form_data=True,
    )
    template_id = _get_template_id(template)
    snapshot = _capture_form_snapshot(
        template_id,
        refresh[2],
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        preview_rows,
    )
    switch_update, save_update = _unsaved_dialog_updates("")
    return (
        new_sheet,
        _session_after_clean_refresh(session, snapshot),
        switch_update,
        save_update,
        *refresh,
    )


def try_template_select(
    template_name: str | None,
    current_template: TemplateConfig | None,
    committed_template_name: str | None,
    form_data: list[dict[str, str]],
    detected_areas: list,
    entry_mode: str,
    session: dict[str, Any],
    committed_sheet: str | None,
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
) -> tuple:
    """Intercept template picks when the form session has unsaved edits."""
    from app.gradio_main import on_template_change

    preview_rows = _normalize_preview_rows(import_preview_rows)
    if (
        session.get("dirty")
        and template_name
        and committed_template_name
        and template_name != committed_template_name
    ):
        gr.Warning("当前有未保存的更改，切换模板前请确认")
        pending = {"type": "template", "target": template_name}
        switch_update, save_update = _unsaved_dialog_updates("switch")
        return (
            committed_template_name,
            committed_template_name,
            current_template,
            gr.update(),
            committed_sheet,
            _begin_unsaved_switch_dialog(session, pending),
            switch_update,
            save_update,
            *_hold_form_refresh_outputs(form_data, detected_areas),
        )
    new_template, sheet_update, default_sheet = on_template_change(template_name, current_template)
    refresh = refresh_data_entry_form(
        new_template, default_sheet, form_data, entry_mode, preserve_form_data=True,
    )
    template_id = _get_template_id(new_template)
    snapshot = _capture_form_snapshot(
        template_id,
        refresh[2],
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        preview_rows,
    )
    switch_update, save_update = _unsaved_dialog_updates("")
    return (
        template_name,
        template_name,
        new_template,
        sheet_update,
        default_sheet,
        _session_after_clean_refresh(session, snapshot),
        switch_update,
        save_update,
        *refresh,
    )


def complete_pending_sheet_navigation(
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    entry_mode: str,
    session: dict[str, Any],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
    *,
    clear_dirty: bool = False,
) -> tuple:
    """Switch to the deferred worksheet while keeping in-memory form data."""
    pending = session.get("pending") or {}
    target_sheet = pending.get("target")
    preview_rows = _normalize_preview_rows(import_preview_rows)
    refresh = refresh_data_entry_form(
        template,
        target_sheet,
        form_data,
        entry_mode,
        preserve_form_data=True,
    )
    template_id = _get_template_id(template)
    snapshot = _capture_form_snapshot(
        template_id,
        refresh[2],
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        preview_rows,
    )
    switch_update, save_update = _unsaved_dialog_updates("")
    if clear_dirty:
        session_update = _session_after_clean_refresh(session, snapshot)
    else:
        session_update = _clear_unsaved_dialog(session)
        session_update["snapshot"] = snapshot
    return (
        target_sheet,
        session_update,
        switch_update,
        save_update,
        *refresh,
    )


def complete_pending_template_navigation(
    form_data: list[dict[str, str]],
    entry_mode: str,
    session: dict[str, Any],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
    *,
    clear_dirty: bool = False,
) -> tuple:
    """Switch to the deferred template while keeping in-memory form data."""
    from app.gradio_main import on_template_change

    pending = session.get("pending") or {}
    template_name = pending.get("target")
    preview_rows = _normalize_preview_rows(import_preview_rows)
    new_template, sheet_update, default_sheet = on_template_change(template_name, None)
    refresh = refresh_data_entry_form(
        new_template,
        default_sheet,
        form_data,
        entry_mode,
        preserve_form_data=True,
    )
    template_id = _get_template_id(new_template)
    snapshot = _capture_form_snapshot(
        template_id,
        refresh[2],
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        preview_rows,
    )
    switch_update, save_update = _unsaved_dialog_updates("")
    if clear_dirty:
        session_update = _session_after_clean_refresh(session, snapshot)
    else:
        session_update = _clear_unsaved_dialog(session)
        session_update["snapshot"] = snapshot
    return (
        template_name,
        template_name,
        new_template,
        sheet_update,
        default_sheet,
        session_update,
        switch_update,
        save_update,
        update_import_stats(new_template),
        *refresh,
    )


def handle_unsaved_switch_no(
    session: dict[str, Any],
    template: TemplateConfig | None,
    committed_sheet: str | None,
    form_data: list[dict[str, str]],
    detected_areas: list,
) -> tuple:
    """Cancel deferred navigation and keep all current form data."""
    switch_update, save_update = _unsaved_dialog_updates("")
    return (
        committed_sheet,
        _clear_unsaved_dialog(session),
        switch_update,
        save_update,
        *_hold_form_refresh_outputs(form_data, detected_areas),
    )


def handle_unsaved_switch_no_with_template(
    session: dict[str, Any],
    template: TemplateConfig | None,
    committed_template_name: str | None,
    committed_sheet: str | None,
    form_data: list[dict[str, str]],
    detected_areas: list,
) -> tuple:
    """Cancel deferred navigation and keep all current form data."""
    switch_update, save_update = _unsaved_dialog_updates("")
    return (
        gr.update(),
        committed_template_name,
        gr.update(),
        gr.update(),
        committed_sheet,
        _clear_unsaved_dialog(session),
        switch_update,
        save_update,
        update_import_stats(template),
        *_hold_form_refresh_outputs(form_data, detected_areas),
    )


def handle_unsaved_save_with_template(
    session: dict[str, Any],
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    committed_sheet: str | None,
    row_selector_value: str | None,
    entry_mode: str,
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
    *field_values: str,
) -> tuple:
    pending = session.get("pending") or {}
    if pending.get("type") == "template":
        return save_export_and_complete_pending_template(
            template,
            form_data,
            committed_sheet,
            row_selector_value,
            entry_mode,
            session,
            import_selection,
            import_preview_active,
            sheet_cache,
            import_view,
            import_preview_rows,
            *field_values,
        )
    sheet_result = save_export_and_complete_pending_sheet(
        template,
        form_data,
        committed_sheet,
        row_selector_value,
        entry_mode,
        session,
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        import_preview_rows,
        *field_values,
    )
    return (
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        *sheet_result[:4],
        update_import_stats(template),
        *sheet_result[4:],
    )


def save_export_and_complete_pending_sheet(
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    committed_sheet: str | None,
    row_selector_value: str | None,
    entry_mode: str,
    session: dict[str, Any],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
    *field_values: str,
) -> tuple:
    """Export current form data (same as 导出 Excel), then switch worksheet."""
    merged = form_data
    if template and committed_sheet:
        merged = _merge_field_boxes_into_form_data(
            template,
            committed_sheet,
            form_data,
            row_selector_value,
            field_values,
        )
        handle_export(
            template,
            merged,
            committed_sheet,
            row_selector_value,
        )
    return complete_pending_sheet_navigation(
        template,
        merged,
        entry_mode,
        session,
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        import_preview_rows,
        clear_dirty=True,
    )


def save_export_and_complete_pending_template(
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    committed_sheet: str | None,
    row_selector_value: str | None,
    entry_mode: str,
    session: dict[str, Any],
    import_selection: dict[str, Any],
    import_preview_active: bool,
    sheet_cache: dict[str, dict[str, str]],
    import_view: str,
    import_preview_rows: Any,
    *field_values: str,
) -> tuple:
    """Export current form data (same as 导出 Excel), then switch template."""
    merged = form_data
    if template and committed_sheet:
        merged = _merge_field_boxes_into_form_data(
            template,
            committed_sheet,
            form_data,
            row_selector_value,
            field_values,
        )
        handle_export(
            template,
            merged,
            committed_sheet,
            row_selector_value,
        )
    return complete_pending_template_navigation(
        merged,
        entry_mode,
        session,
        import_selection,
        import_preview_active,
        sheet_cache,
        import_view,
        import_preview_rows,
        clear_dirty=True,
    )


def stay_on_current_view(session: dict[str, Any]) -> tuple:
    """Close unsaved dialogs and keep the current template/worksheet."""
    switch_update, save_update = _unsaved_dialog_updates("")
    return _clear_unsaved_dialog(session), switch_update, save_update


def confirm_unsaved_switch_yes(session: dict[str, Any]) -> tuple:
    """First-step confirm: offer save or return."""
    switch_update, save_update = _unsaved_dialog_updates("save")
    return _begin_unsaved_save_dialog(session), switch_update, save_update


def _inactive_form_refresh(form_data: list[dict[str, str]]) -> tuple:
    field_updates = _empty_field_updates()
    row_updates = _empty_row_updates()
    return (
        gr.update(visible=True),
        [],
        form_data,
        gr.update(choices=[], value=None, visible=False),
        gr.update(value="", visible=False, interactive=True),
        gr.update(visible=True),
        *_interleaved_row_field_updates(row_updates, field_updates),
    )


def refresh_data_entry_form(
    template: TemplateConfig | None,
    sheet_name: str | None,
    form_data: list[dict[str, str]],
    entry_mode: str = ENTRY_MODE_ID_AUTO,
    *,
    preserve_form_data: bool = False,
) -> tuple:
    """
    Detect areas and populate dynamic form fields from paste/sections config.

    Returns updates for:
    form_container, detected_areas_state, form_data_state,
    row_selector, paste_input, fields_container,
    then each field row and its textboxes in creation order.
    """
    if not template or not sheet_name:
        return _inactive_form_refresh(form_data)

    headers = get_form_field_headers(template, sheet_name)
    if not headers:
        gr.Warning("未找到模板输入区域或表头，请检查 Excel 模板或区域配置")
        return _inactive_form_refresh(form_data)

    active_area_range = _resolve_input_area_range(template, sheet_name)
    detected_areas: list[DetectedArea] = []
    if active_area_range:
        detected_areas = _detect_form_areas(template, sheet_name, active_area_range)
    if detected_areas:
        if preserve_form_data and form_data:
            form_data = _align_form_data_to_areas(form_data, headers, len(detected_areas))
        else:
            form_data = _load_form_data_for_areas(template, sheet_name, headers, detected_areas)
    else:
        form_data = [{header: "" for header in headers}]
    static_area = active_area_range or (detected_areas[0].area if detected_areas else "")
    static_values = (
        _load_static_template_row_values(template, sheet_name, headers, static_area)
        if static_area
        else {header: "" for header in headers}
    )
    row_choices = _row_choices_for_form_data(
        len(form_data),
        areas=detected_areas or None,
        form_rows=form_data,
        static_values=static_values,
        headers=headers,
    )
    row_values = form_data[0] if form_data else {header: "" for header in headers}
    field_updates, row_updates, status_update = _build_field_updates(
        headers, row_values, entry_mode
    )
    return (
        gr.update(visible=True),
        detected_areas,
        form_data,
        _row_selector_update(
            row_choices,
            row_choices[0] if row_choices else None,
            entry_mode=entry_mode,
            active=True,
        ),
        status_update,
        gr.update(visible=True),
        *_interleaved_row_field_updates(row_updates, field_updates),
    )


def _begin_paste_parse() -> gr.update:
    return gr.update(interactive=False)


def apply_pasted_form_data(
    template: TemplateConfig | None,
    paste_text: str,
    form_data: list[dict[str, str]],
    row_selector_value: str | None,
    detected_areas: list | None = None,
    sheet_name: str | None = None,
    session: dict[str, Any] | None = None,
) -> tuple:
    """Parse tab-separated paste text and fill form rows using paste YAML index rules."""
    headers = get_form_field_headers(template) if template else []
    if not template:
        gr.Warning("请先选择模板")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    if not headers:
        gr.Warning("未找到模板输入区域或表头")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    if not _should_parse_paste(paste_text):
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    paste_config = load_paste_parse_config(template.id)
    if not paste_config:
        gr.Warning("模板粘贴配置加载失败")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    try:
        parsed_rows = parse_text_with_config(paste_text.strip(), paste_config)
    except ValueError as exc:
        gr.Warning(f"解析失败：{exc}")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    if not parsed_rows:
        gr.Warning("未能解析粘贴数据")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    has_values = any(
        value
        for row in parsed_rows
        for key, value in row.items()
        if key != "order" and str(value).strip()
    )
    if not has_values:
        gr.Warning("粘贴数据未匹配任何字段，请检查 YAML index 配置")
        return form_data, gr.update(), gr.update(interactive=True), session or _empty_form_session(), *_NO_CHANGE_FIELD_UPDATES
    if not form_data:
        form_data = [{header: "" for header in headers}]
    target_idx = _resolve_row_index(row_selector_value, len(form_data))
    for offset, parsed in enumerate(parsed_rows):
        row_idx = target_idx + offset
        while len(form_data) <= row_idx:
            form_data.append({header: "" for header in headers})
        form_data[row_idx] = _row_values_for_headers(headers, parsed)
    row_choices = _build_row_choices_for_template(
        template,
        sheet_name,
        form_data,
        detected_areas or [],
    )
    selected_row = row_choices[target_idx] if target_idx < len(row_choices) else row_choices[0]
    field_updates = _field_updates_for_row(headers, form_data[target_idx])
    gr.Info(f"已填充 {len(parsed_rows)} 行数据到表单")
    return (
        form_data,
        gr.update(choices=row_choices, value=selected_row),
        gr.update(value="", interactive=True),
        _mark_form_dirty(session or _empty_form_session()),
        *field_updates,
    )


def _begin_advance_next() -> gr.update:
    return gr.update(interactive=False)


def advance_to_next_area(
    template: TemplateConfig | None,
    sheet_name: str | None,
    form_data: list[dict[str, str]],
    row_selector_value: str | None,
    import_selection_state: dict[str, Any],
    sheet_cache: dict[str, dict[str, str]],
    import_preview_active: bool,
    entry_mode: str,
    session: dict[str, Any],
    detected_areas: list,
    *field_values: str,
) -> tuple:
    """Save the current row and move the row selector to the next area."""
    if (
        import_preview_active
        and template
        and import_selection_state.get("ids")
    ):
        ids = import_selection_state["ids"]
        current_idx = max(0, min(import_selection_state.get("index", 0), len(ids) - 1))
        next_idx = current_idx + 1
        if next_idx >= len(ids):
            gr.Info("已是最后一个选中行")
            next_idx = current_idx
        else:
            gr.Info(f"已切换到选中行 {next_idx + 1}/{len(ids)}")
        new_state = {"ids": ids, "index": next_idx}
        row_selector_update, new_state, field_updates = _import_preview_sync_outputs(
            template,
            form_data,
            new_state,
            sheet_cache,
            entry_mode,
        )
        return (
            form_data,
            row_selector_update,
            gr.update(interactive=True),
            new_state,
            session,
            *field_updates,
        )
    if not template or not sheet_name:
        gr.Warning("请先选择模板和工作表")
        return (
            form_data,
            gr.update(),
            gr.update(interactive=True),
            gr.update(),
            session,
            *_NO_CHANGE_FIELD_UPDATES,
        )
    headers = get_form_field_headers(template, sheet_name)
    if not headers:
        gr.Warning("未找到模板输入区域或表头")
        return (
            form_data,
            gr.update(),
            gr.update(interactive=True),
            gr.update(),
            session,
            *_empty_field_updates(),
        )
    merged = _merge_field_boxes_into_form_data(
        template,
        sheet_name,
        form_data,
        row_selector_value,
        field_values,
    )
    current_idx = _resolve_row_index(row_selector_value, len(merged))
    next_idx = current_idx + 1
    if next_idx >= len(merged):
        gr.Info("已是最后一个区域")
        next_idx = current_idx
    else:
        gr.Info(f"已切换到 Row {next_idx + 1}")
    row_choices = _build_row_choices_for_template(
        template,
        sheet_name,
        merged,
        detected_areas or [],
    )
    selected_row = row_choices[next_idx] if next_idx < len(row_choices) else f"Row {next_idx + 1}"
    field_updates = _field_updates_for_row(headers, merged[next_idx])
    return (
        merged,
        gr.update(value=selected_row),
        gr.update(interactive=True),
        gr.update(),
        _mark_form_dirty(session or _empty_form_session()),
        *field_updates,
    )


def sync_form_fields_to_row(
    template: TemplateConfig | None,
    row_selector_value: str | None,
    form_data: list[dict[str, str]],
    import_selection_state: dict[str, Any],
    sheet_cache: dict[str, dict[str, str]],
    import_preview_active: bool,
    entry_mode: str,
) -> tuple:
    """Show the selected form row in field textboxes."""
    if (
        import_preview_active
        and template
        and import_selection_state.get("ids")
    ):
        ids = import_selection_state["ids"]
        index = _resolve_row_index(row_selector_value, len(ids))
        new_state = {"ids": ids, "index": index}
        _, _, field_updates = _import_preview_sync_outputs(
            template,
            form_data,
            new_state,
            sheet_cache,
            entry_mode,
        )
        return *field_updates, new_state
    if not template:
        return *_empty_field_updates(), gr.update()
    headers = get_form_field_headers(template)
    if not headers:
        return *_empty_field_updates(), gr.update()
    if not form_data:
        return tuple(_field_updates_for_row(headers, {header: "" for header in headers})), gr.update()
    target_idx = _resolve_row_index(row_selector_value, len(form_data))
    row_values = _row_values_for_headers(headers, form_data[target_idx])
    return tuple(_field_updates_for_row(headers, row_values)), gr.update()


def handle_import_preview_selection_change(
    preview_data: Any,
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    selection_state: dict[str, Any],
    sheet_cache: dict[str, dict[str, str]],
    preview_active: bool,
    entry_mode: str,
) -> tuple:
    """Lightweight preview of checked import rows using cached sheet data."""
    if not preview_active or not template:
        return gr.update(), gr.update(), *_NO_CHANGE_FIELD_UPDATES
    preview_rows = _normalize_preview_rows(preview_data)
    selected_ids = _selected_import_ids(preview_rows)
    prev_ids = selection_state.get("ids", [])
    if not selected_ids:
        headers = get_form_field_headers(template)
        if not headers:
            return gr.update(), _empty_import_selection(), *_NO_CHANGE_FIELD_UPDATES
        row_count = len(form_data) if form_data else 1
        row_choices = _row_choices_for_form_data(row_count)
        target_idx = min(selection_state.get("index", 0), max(0, row_count - 1))
        row_values = (
            form_data[target_idx]
            if form_data and target_idx < len(form_data)
            else {header: "" for header in headers}
        )
        return (
            _row_selector_update(
                row_choices,
                row_choices[target_idx] if row_choices else None,
                entry_mode=entry_mode,
            ),
            _empty_import_selection(),
            *_field_updates_for_row(headers, row_values),
        )
    if selected_ids == prev_ids:
        return gr.update(), gr.update(), *_NO_CHANGE_FIELD_UPDATES
    new_state = {"ids": selected_ids, "index": 0}
    row_selector_update, new_state, field_updates = _import_preview_sync_outputs(
        template,
        form_data,
        new_state,
        sheet_cache,
        entry_mode,
    )
    return row_selector_update, new_state, *field_updates


def _restore_form_row_selector(
    template: TemplateConfig,
    form_data: list[dict[str, str]],
    entry_mode: str,
) -> gr.update:
    headers = get_form_field_headers(template)
    row_count = len(form_data) if form_data else 1
    row_choices = _row_choices_for_form_data(row_count)
    return _row_selector_update(
        row_choices,
        row_choices[0] if row_choices else None,
        entry_mode=entry_mode,
        active=bool(headers),
    )


def _import_refresh_tail(
    sheet_cache: dict[str, dict[str, str]],
    preview_active: bool,
) -> tuple[dict[str, dict[str, str]], dict[str, Any], bool]:
    return sheet_cache, _empty_import_selection(), preview_active


def update_import_stats(template: TemplateConfig | dict[str, Any] | None) -> str:
    """
    Update import statistics display
    
    Returns:
        Markdown formatted stats
    """
    template_id = _get_template_id(template)
    if not template_id:
        return "📊 导入统计：未选择模板"
    
    try:
        stats = get_import_stats(template_id)
        last_import = _format_last_import(stats.get("last_import"))
        
        return (
            f"📊 **导入统计** | "
            f"已处理: **{stats['processed_count']}** | "
            f"垃圾数据: **{stats['trash_count']}** | "
            f"最后导入: {last_import}"
        )
    except Exception as e:
        logger.error(f"Failed to get import stats: {e}")
        return "📊 导入统计：加载失败"


def _begin_bulk_refresh() -> tuple[Any, str]:
    """Fast UI feedback before a Google Sheet fetch."""
    return (
        gr.update(interactive=False),
        "📊 **导入统计** | 正在从 Google Sheet 加载数据...",
    )


def _begin_import_selected() -> tuple[Any, Any]:
    return gr.update(interactive=False), gr.update(interactive=False)


def _import_failure_returns(
    form_data: list[dict[str, str]],
    template: TemplateConfig | dict[str, Any] | None,
    session: dict[str, Any] | None = None,
) -> tuple:
    headers = get_form_field_headers(template) if template and isinstance(template, TemplateConfig) else []
    field_updates = (
        tuple(_field_updates_for_row(headers, form_data[0]))
        if headers and form_data
        else _NO_CHANGE_FIELD_UPDATES
    )
    return (
        form_data,
        gr.update(),
        gr.update(),
        gr.update(interactive=True),
        gr.update(interactive=True),
        update_import_stats(template),
        gr.update(),
        gr.update(),
        session or _empty_form_session(),
        *field_updates,
    )


def _preview_columns(df: pl.DataFrame, id_column: str, limit: int = 3) -> list[str]:
    return [column for column in df.columns if column != id_column][:limit]


def _build_import_preview_rows(
    df: pl.DataFrame,
    id_column: str,
    include_ids: set[str] | None,
    exclude_ids: set[str] | None,
    status_label: str,
    max_rows: int,
) -> list[list[Any]]:
    if df.height == 0 or id_column not in df.columns:
        return []

    id_expr = pl.col(id_column).cast(pl.Utf8).str.strip_chars()
    filtered = df
    if include_ids is not None:
        filtered = filtered.filter(id_expr.is_in(list(include_ids)))
    if exclude_ids:
        filtered = filtered.filter(~id_expr.is_in(list(exclude_ids)))

    preview_cols = _preview_columns(df, id_column)
    rows: list[list[Any]] = []
    for record in filtered.head(max_rows).iter_rows(named=True):
        id_val = str(record.get(id_column, ""))
        preview_vals = [str(record.get(column, "")) for column in preview_cols]
        rows.append([False, id_val, status_label, " | ".join(preview_vals)])
    return rows


def _load_template_sheet_df(
    credentials: Any,
    data_source: Any,
    *,
    force_refresh: bool = False,
) -> pl.DataFrame:
    return fetch_all_rows(
        credentials,
        data_source.sheet_url,
        data_source.worksheet_name,
        force_refresh=force_refresh,
    )


def _find_id_field_key(template_id: str) -> str | None:
    """
    Find the ID field key from paste config
    
    Args:
        template_id: Template identifier
        
    Returns:
        ID field key if found, None otherwise
    """
    paste_config = load_paste_parse_config(template_id)
    if not paste_config:
        return None
    return id_target_field_from_config(paste_config)


def _id_field_header_index(template_id: str, headers: list[str]) -> int | None:
    """Return the form field index for the paste-config ID field, if present."""
    id_field = _find_id_field_key(template_id)
    if not id_field:
        return None
    try:
        return headers.index(id_field)
    except ValueError:
        return None


def _id_lookup_no_change(
    form_data: list[dict[str, str]],
    template_id: str | None,
    headers: list[str],
    session: dict[str, Any],
) -> tuple:
    """Keep form state unchanged and re-enable the ID field after a lookup attempt."""
    updates = list(_NO_CHANGE_FIELD_UPDATES)
    if template_id:
        id_idx = _id_field_header_index(template_id, headers)
        if id_idx is not None:
            updates[id_idx] = gr.update(interactive=True)
    return (form_data, *updates, session)


def _begin_id_lookup_gate(
    field_index: int,
    field_value: str,
    template: TemplateConfig | None,
    entry_mode: str,
    import_preview_active: bool,
) -> gr.update:
    """Disable the ID field while fetching sheet data; no-op for other fields."""
    if import_preview_active:
        return gr.update()
    if entry_mode != ENTRY_MODE_ID_AUTO or not template:
        return gr.update()
    headers = get_form_field_headers(template)
    id_idx = _id_field_header_index(template.id, headers)
    if id_idx is None or field_index != id_idx:
        return gr.update()
    if not str(field_value or "").strip():
        return gr.update()
    return gr.update(interactive=False)


def handle_id_field_lookup(
    field_index: int,
    field_value: str,
    template: TemplateConfig | None,
    credentials: Any,
    form_data: list[dict[str, str]],
    row_selector_value: str | None,
    entry_mode: str,
    import_preview_active: bool,
    session: dict[str, Any],
) -> tuple:
    """Look up a sheet row by ID and populate the current form row (ID Auto mode)."""
    headers = get_form_field_headers(template) if template else []
    template_id = _get_template_id(template)
    session = session or _empty_form_session()

    def no_change() -> tuple:
        return _id_lookup_no_change(form_data, template_id, headers, session)

    if import_preview_active:
        return no_change()
    if entry_mode != ENTRY_MODE_ID_AUTO or not template or not template_id:
        return no_change()
    id_idx = _id_field_header_index(template_id, headers)
    if id_idx is None or field_index != id_idx:
        return no_change()
    id_value = str(field_value or "").strip()
    if not id_value:
        return no_change()
    if not credentials:
        gr.Warning("请先连接 Google 账号")
        return no_change()
    from app.services.data_source import load_template_data_source

    data_source = load_template_data_source(template_id)
    if not data_source:
        gr.Warning("模板未配置数据源")
        return no_change()
    paste_config = load_paste_parse_config(template_id)
    if not paste_config:
        gr.Warning("模板粘贴配置加载失败")
        return no_change()
    try:
        sheet_df = _load_template_sheet_df(credentials, data_source)
        sheet_row = lookup_row_by_id(sheet_df, data_source.id_column, id_value)
        if not sheet_row:
            gr.Warning(f"未找到 ID: {id_value}")
            return no_change()
        matched = map_sheet_row_from_paste_config(sheet_row, paste_config)
        id_field_key = id_target_field_from_config(paste_config)
        if id_field_key:
            matched[id_field_key] = id_value
        if not matched:
            gr.Warning("未能匹配任何字段")
            return no_change()
        if not form_data:
            form_data = [{header: "" for header in headers}]
        target_idx = _resolve_row_index(row_selector_value, len(form_data))
        while len(form_data) <= target_idx:
            form_data.append({header: "" for header in headers})
        existing_row = form_data[target_idx] if target_idx < len(form_data) else {}
        form_data[target_idx] = _merge_mapped_into_row(headers, existing_row, matched)
        field_updates = _field_updates_for_row(headers, form_data[target_idx])
        gr.Info(f"已自动填充 ID {id_value} 的数据")
        return (form_data, *field_updates, _mark_form_dirty(session))
    except GoogleSheetsError as exc:
        gr.Warning(str(exc))
        return no_change()
    except Exception as exc:
        logger.error(f"ID lookup failed: {exc}")
        gr.Warning(f"查询失败：{exc}")
        return no_change()


def build_form_tab(
    current_template: gr.State,
    credentials_state: gr.State,
    form_data_state: gr.State,
    detected_areas_state: gr.State,
    template_selector: gr.Dropdown | None = None,
) -> dict:
    """
    Build the data entry tab
    
    Returns:
        Dict of component references for event binding
    """
    components = {}
    import_view_state = gr.State("unprocessed")
    components["import_view_state"] = import_view_state
    import_sheet_cache_state = gr.State({})
    components["import_sheet_cache_state"] = import_sheet_cache_state
    import_selection_state = gr.State(_empty_import_selection())
    components["import_selection_state"] = import_selection_state
    import_preview_active_state = gr.State(False)
    components["import_preview_active_state"] = import_preview_active_state
    entry_mode_state = gr.State(ENTRY_MODE_ID_AUTO)
    components["entry_mode_state"] = entry_mode_state
    form_session_state = gr.State(_empty_form_session())
    components["form_session_state"] = form_session_state
    committed_sheet_state = gr.State(None)
    components["committed_sheet_state"] = committed_sheet_state
    committed_template_name_state = gr.State(None)
    components["committed_template_name_state"] = committed_template_name_state
    
    with gr.Column():
        gr.Markdown("## 表单数据")
        entry_mode = gr.Radio(
            choices=[ENTRY_MODE_ID_AUTO, ENTRY_MODE_MANUAL],
            value=ENTRY_MODE_ID_AUTO,
            label="录入方式",
            interactive=True,
        )
        components["entry_mode"] = entry_mode
        with gr.Row():
            sheet_selector = gr.Dropdown(
                label="选择工作表",
                choices=[],
                value=None,
                interactive=True,
            )
            components["sheet_selector"] = sheet_selector
            row_selector = gr.Dropdown(
                label="区域选择",
                choices=[],
                value=None,
                visible=True,
            )
            components["row_selector"] = row_selector
        paste_input = gr.Textbox(
            label="粘贴数据",
            placeholder="粘贴 Tab 分隔的一行（可从 Excel 复制）；将按 YAML 配置解析并填入表单",
            lines=3,
            max_lines=8,
            visible=False,
            interactive=True,
        )
        components["paste_input"] = paste_input
        with gr.Column(visible=False) as unsaved_switch_group:
            gr.Markdown("### 有未保存的更改\n是否继续切换？当前数据将保留。")
            with gr.Row():
                unsaved_switch_yes_btn = gr.Button("是")
                unsaved_switch_no_btn = gr.Button("否", variant="primary")
        with gr.Column(visible=False) as unsaved_save_group:
            gr.Markdown("### 是否先保存？\n保存将导出 Excel；返回将留在当前页面。")
            with gr.Row():
                unsaved_save_btn = gr.Button("保存并切换")
                unsaved_stay_btn = gr.Button("返回当前", variant="primary")
        components["unsaved_switch_group"] = unsaved_switch_group
        components["unsaved_save_group"] = unsaved_save_group
        
        # Field grid (field slots toggle via refresh; container stays mounted)
        with gr.Column(visible=True) as form_container:
            form_field_boxes: list[gr.Textbox] = []
            field_rows: list[gr.Row] = []
            # Row/column shells stay visible; only textbox slots toggle via refresh updates.
            with gr.Column(visible=True) as fields_container:
                for row_start in range(0, MAX_FORM_FIELDS, FORM_LAYOUT_FIELDS_PER_ROW):
                    with gr.Row(visible=True) as field_row:
                        field_rows.append(field_row)
                        row_end = min(row_start + FORM_LAYOUT_FIELDS_PER_ROW, MAX_FORM_FIELDS)
                        for index in range(row_start, row_end):
                            field_box = gr.Textbox(
                                label=f"字段 {index + 1}",
                                visible=False,
                                interactive=True,
                                scale=1,
                                min_width=100,
                            )
                            form_field_boxes.append(field_box)

            components["form_field_boxes"] = form_field_boxes
            components["field_rows"] = field_rows
            components["fields_container"] = fields_container
            with gr.Row(elem_classes=["form-next-row"]):
                next_area_btn = gr.Button("下一个", scale=0, min_width=100)
            components["next_area_btn"] = next_area_btn
        
        components["form_container"] = form_container
        
        # Action buttons
        with gr.Row():
            export_btn = gr.Button("导出 Excel", variant="primary")
            print_btn = gr.Button("打印预览")
        export_download = gr.File(label="导出文件", visible=False, interactive=False)
        
        components["export_btn"] = export_btn
        components["print_btn"] = print_btn
        components["export_download"] = export_download
        
        # Bulk import section (always visible; no gr.Group — avoids Gradio panel background)
        gr.Markdown("## 批量导入")
        with gr.Row():
            import_stats = gr.Markdown("📊 导入统计：加载中...")
        
        with gr.Row():
            refresh_btn = gr.Button("🔄 刷新未处理数据")
            show_processed_btn = gr.Button("📝 查看已处理", variant="secondary")
            show_trash_btn = gr.Button("🗑️ 查看垃圾数据", variant="secondary")
        
        import_preview = gr.Dataframe(
            headers=["选择", "ID", "状态", "数据预览"],
            datatype=["bool", "str", "str", "str"],
            interactive=True,
            wrap=True,
            visible=False
        )
        
        with gr.Row():
            import_btn = gr.Button(
                "✅ 导入选中行",
                variant="primary",
                visible=False
            )
            
            mark_trash_btn = gr.Button(
                "🗑️ 标记为垃圾",
                variant="secondary",
                visible=False
            )
            
            restore_btn = gr.Button(
                "↩️ 恢复为未处理",
                variant="secondary",
                visible=False,
            )
            
            clear_history_btn = gr.Button(
                "🧹 清空历史",
                variant="stop",
                visible=False,
                size="sm"
            )
        
        components["import_stats"] = import_stats
        components["refresh_btn"] = refresh_btn
        components["show_processed_btn"] = show_processed_btn
        components["show_trash_btn"] = show_trash_btn
        components["import_preview"] = import_preview
        components["import_btn"] = import_btn
        components["mark_trash_btn"] = mark_trash_btn
        components["restore_btn"] = restore_btn
        components["clear_history_btn"] = clear_history_btn
    
    # Event bindings
    
    # Load import stats when template changes
    current_template.change(
        fn=update_import_stats,
        inputs=[current_template],
        outputs=[import_stats],
        **HIDE_PROGRESS,
    )
    
    form_refresh_outputs = [
        form_container,
        detected_areas_state,
        form_data_state,
        row_selector,
        paste_input,
        fields_container,
        *_form_field_output_components(field_rows, form_field_boxes),
    ]

    guarded_sheet_outputs = [
        committed_sheet_state,
        form_session_state,
        unsaved_switch_group,
        unsaved_save_group,
        *form_refresh_outputs,
    ]

    # User-initiated sheet picks only: .change also fires on template-driven gr.update
    # and can refresh with a stale current_template, leaving visible slots at label "".
    sheet_selector.select(
        fn=try_sheet_select,
        inputs=[
            sheet_selector,
            committed_sheet_state,
            current_template,
            form_data_state,
            detected_areas_state,
            entry_mode_state,
            form_session_state,
            import_selection_state,
            import_preview_active_state,
            import_sheet_cache_state,
            import_view_state,
            import_preview,
        ],
        outputs=guarded_sheet_outputs,
        **HIDE_PROGRESS,
    )

    entry_mode.change(
        fn=on_entry_mode_change,
        inputs=[entry_mode, current_template],
        outputs=[entry_mode_state, row_selector, paste_input],
        **HIDE_PROGRESS,
    )

    id_lookup_outputs = [form_data_state, *form_field_boxes]
    for index, field_box in enumerate(form_field_boxes):
        field_box.blur(
            fn=partial(_begin_id_lookup_gate, index),
            inputs=[
                field_box,
                current_template,
                entry_mode_state,
                import_preview_active_state,
            ],
            outputs=[field_box],
            **HIDE_PROGRESS,
        ).then(
            fn=partial(handle_id_field_lookup, index),
            inputs=[
                field_box,
                current_template,
                credentials_state,
                form_data_state,
                row_selector,
                entry_mode_state,
                import_preview_active_state,
                form_session_state,
            ],
            outputs=[*id_lookup_outputs, form_session_state],
            **HIDE_PROGRESS,
        )
        field_box.change(
            fn=_mark_form_dirty,
            inputs=[form_session_state],
            outputs=[form_session_state],
            **HIDE_PROGRESS,
        )

    paste_apply_outputs = [form_data_state, row_selector, paste_input, form_session_state, *form_field_boxes]

    row_selector.change(
        fn=sync_form_fields_to_row,
        inputs=[
            current_template,
            row_selector,
            form_data_state,
            import_selection_state,
            import_sheet_cache_state,
            import_preview_active_state,
            entry_mode_state,
        ],
        outputs=[import_selection_state, *form_field_boxes],
        **HIDE_PROGRESS,
    )

    import_preview_selection_outputs = [
        row_selector,
        import_selection_state,
        *form_field_boxes,
    ]
    import_preview.change(
        fn=handle_import_preview_selection_change,
        inputs=[
            import_preview,
            current_template,
            form_data_state,
            import_selection_state,
            import_sheet_cache_state,
            import_preview_active_state,
            entry_mode_state,
        ],
        outputs=import_preview_selection_outputs,
        **HIDE_PROGRESS,
    )

    paste_input.submit(
        fn=_begin_paste_parse,
        outputs=[paste_input],
        **HIDE_PROGRESS,
    ).then(
        fn=apply_pasted_form_data,
        inputs=[
            current_template,
            paste_input,
            form_data_state,
            row_selector,
            detected_areas_state,
            sheet_selector,
            form_session_state,
        ],
        outputs=paste_apply_outputs,
        **HIDE_PROGRESS,
    )

    paste_input.change(
        fn=_begin_paste_parse,
        outputs=[paste_input],
        **HIDE_PROGRESS,
    ).then(
        fn=apply_pasted_form_data,
        inputs=[
            current_template,
            paste_input,
            form_data_state,
            row_selector,
            detected_areas_state,
            sheet_selector,
            form_session_state,
        ],
        outputs=paste_apply_outputs,
        **HIDE_PROGRESS,
    )

    next_apply_outputs = [
        form_data_state,
        row_selector,
        next_area_btn,
        import_selection_state,
        form_session_state,
        *form_field_boxes,
    ]
    next_area_btn.click(
        fn=_begin_advance_next,
        outputs=[next_area_btn],
        **HIDE_PROGRESS,
    ).then(
        fn=advance_to_next_area,
        inputs=[
            current_template,
            sheet_selector,
            form_data_state,
            row_selector,
            import_selection_state,
            import_sheet_cache_state,
            import_preview_active_state,
            entry_mode_state,
            form_session_state,
            detected_areas_state,
            *form_field_boxes,
        ],
        outputs=next_apply_outputs,
        **HIDE_PROGRESS,
    )
    
    import_refresh_tail_outputs = [
        import_sheet_cache_state,
        import_selection_state,
        import_preview_active_state,
    ]
    refresh_btn.click(
        fn=_begin_bulk_refresh,
        outputs=[refresh_btn, import_stats],
        **HIDE_PROGRESS,
    ).then(
        fn=handle_refresh_unrecorded,
        inputs=[current_template, credentials_state, form_data_state],
        outputs=[
            import_preview, import_btn, mark_trash_btn, restore_btn,
            clear_history_btn, refresh_btn, import_stats, import_view_state,
            *import_refresh_tail_outputs,
        ],
        **HIDE_PROGRESS,
    )
    
    # Show processed data
    show_processed_btn.click(
        fn=handle_show_processed,
        inputs=[current_template, credentials_state],
        outputs=[
            import_preview, import_btn, mark_trash_btn, restore_btn,
            clear_history_btn, import_view_state,
            *import_refresh_tail_outputs,
        ],
        **HIDE_PROGRESS,
    )
    
    # Show trash data
    show_trash_btn.click(
        fn=handle_show_trash,
        inputs=[current_template, credentials_state],
        outputs=[
            import_preview, import_btn, mark_trash_btn, restore_btn,
            clear_history_btn, import_view_state,
            *import_refresh_tail_outputs,
        ],
        **HIDE_PROGRESS,
    )
    
    import_apply_outputs = [
        form_data_state,
        row_selector,
        import_preview,
        import_btn,
        mark_trash_btn,
        import_stats,
        import_selection_state,
        import_preview_active_state,
        form_session_state,
        *form_field_boxes,
    ]
    # Import selected rows
    import_btn.click(
        fn=_begin_import_selected,
        outputs=[import_btn, mark_trash_btn],
        **HIDE_PROGRESS,
    ).then(
        fn=handle_import_selected,
        inputs=[
            import_preview,
            current_template,
            form_data_state,
            credentials_state,
            detected_areas_state,
            import_sheet_cache_state,
            entry_mode_state,
            sheet_selector,
            form_session_state,
        ],
        outputs=import_apply_outputs,
        **HIDE_PROGRESS,
    )
    
    # Mark as trash
    mark_trash_btn.click(
        fn=handle_mark_trash,
        inputs=[import_preview, current_template],
        outputs=[import_preview, mark_trash_btn, import_stats],
        **HIDE_PROGRESS,
    ).then(
        fn=_mark_form_dirty,
        inputs=[form_session_state],
        outputs=[form_session_state],
        **HIDE_PROGRESS,
    )
    
    # Restore selected rows to unprocessed
    restore_btn.click(
        fn=handle_restore_selected,
        inputs=[import_preview, current_template],
        outputs=[import_preview, restore_btn, import_stats],
        **HIDE_PROGRESS,
    ).then(
        fn=_mark_form_dirty,
        inputs=[form_session_state],
        outputs=[form_session_state],
        **HIDE_PROGRESS,
    )
    
    # Clear history
    clear_history_btn.click(
        fn=handle_clear_history,
        inputs=[current_template],
        outputs=[import_stats],
    ).then(
        fn=_mark_form_dirty,
        inputs=[form_session_state],
        outputs=[form_session_state],
        **HIDE_PROGRESS,
    )
    
    # Export filled workbook to exports/ and trigger browser download
    export_btn.click(
        fn=handle_export,
        inputs=[
            current_template,
            form_data_state,
            sheet_selector,
            row_selector,
            *form_field_boxes,
        ],
        outputs=[export_download],
        **HIDE_PROGRESS,
    )

    print_btn.click(
        fn=handle_print_preview,
        inputs=[
            current_template,
            form_data_state,
            sheet_selector,
            row_selector,
            *form_field_boxes,
        ],
        **HIDE_PROGRESS,
    )
    
    components["form_refresh_outputs"] = form_refresh_outputs
    components["guarded_sheet_outputs"] = guarded_sheet_outputs
    components["update_on_template_change"] = [sheet_selector]

    unsaved_dialog_outputs = [
        committed_sheet_state,
        form_session_state,
        unsaved_switch_group,
        unsaved_save_group,
        *form_refresh_outputs,
    ]
    unsaved_switch_yes_btn.click(
        fn=confirm_unsaved_switch_yes,
        inputs=[form_session_state],
        outputs=[form_session_state, unsaved_switch_group, unsaved_save_group],
        **HIDE_PROGRESS,
    )
    unsaved_stay_btn.click(
        fn=stay_on_current_view,
        inputs=[form_session_state],
        outputs=[form_session_state, unsaved_switch_group, unsaved_save_group],
        **HIDE_PROGRESS,
    )
    if template_selector is not None:
        guarded_template_outputs = [
            template_selector,
            committed_template_name_state,
            current_template,
            sheet_selector,
            committed_sheet_state,
            form_session_state,
            unsaved_switch_group,
            unsaved_save_group,
            import_stats,
            *form_refresh_outputs,
        ]
        components["guarded_template_outputs"] = guarded_template_outputs
        unsaved_switch_no_btn.click(
            fn=handle_unsaved_switch_no_with_template,
            inputs=[
                form_session_state,
                current_template,
                committed_template_name_state,
                committed_sheet_state,
                form_data_state,
                detected_areas_state,
            ],
            outputs=guarded_template_outputs,
            **HIDE_PROGRESS,
        )
        unsaved_save_btn.click(
            fn=handle_unsaved_save_with_template,
            inputs=[
                form_session_state,
                current_template,
                form_data_state,
                committed_sheet_state,
                row_selector,
                entry_mode_state,
                import_selection_state,
                import_preview_active_state,
                import_sheet_cache_state,
                import_view_state,
                import_preview,
                *form_field_boxes,
            ],
            outputs=guarded_template_outputs,
            **HIDE_PROGRESS,
        )
    else:
        unsaved_switch_no_btn.click(
            fn=handle_unsaved_switch_no,
            inputs=[
                form_session_state,
                current_template,
                committed_sheet_state,
                form_data_state,
                detected_areas_state,
            ],
            outputs=unsaved_dialog_outputs,
            **HIDE_PROGRESS,
        )
        unsaved_save_btn.click(
            fn=handle_unsaved_save_with_template,
            inputs=[
                form_session_state,
                current_template,
                form_data_state,
                committed_sheet_state,
                row_selector,
                entry_mode_state,
                import_selection_state,
                import_preview_active_state,
                import_sheet_cache_state,
                import_view_state,
                import_preview,
                *form_field_boxes,
            ],
            outputs=unsaved_dialog_outputs,
            **HIDE_PROGRESS,
        )

    return components


def handle_refresh_unrecorded(
    template: TemplateConfig | None,
    credentials: Any,
    form_data: list[dict[str, str]],
) -> tuple:
    """
    Refresh unrecorded data from Google Sheet
    Filters out processed IDs and trash IDs
    
    Returns:
        (import_preview, import_btn, mark_trash_btn, restore_btn,
         clear_history_btn, refresh_btn, import_stats, import_view_state)
    """
    if not template or not credentials:
        gr.Warning("请先配置数据源并连接 Google 账号")
        return (
            gr.update(), 
            gr.update(visible=False), 
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(interactive=True),
            update_import_stats(template),
            "unprocessed",
            *_import_refresh_tail({}, False),
        )
    
    try:
        from app.services.data_source import load_template_data_source
        
        data_source = load_template_data_source(template.id)
        if not data_source:
            gr.Warning("模板未配置数据源")
            return (
                gr.update(), 
                gr.update(visible=False), 
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(interactive=True),
                update_import_stats(template),
                "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        history = load_import_history(template.id)
        exclude_ids = history.processed_ids | history.trash_ids

        logger.info(f"Fetching all rows from sheet: {data_source.worksheet_name}")
        invalidate_sheet_cache(data_source.sheet_url, data_source.worksheet_name)
        df = _load_template_sheet_df(credentials, data_source, force_refresh=True)
        
        if df.height == 0:
            gr.Info("Sheet 中没有数据")
            return (
                gr.update(), 
                gr.update(visible=False), 
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(interactive=True),
                update_import_stats(template),
                "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        if df.height > MAX_IMPORT_PREVIEW_ROWS:
            gr.Warning(
                f"数据量过大（{df.height} 行），仅显示前 {MAX_IMPORT_PREVIEW_ROWS} 行未录入数据"
            )

        unrecorded_rows = _build_import_preview_rows(
            df,
            data_source.id_column,
            include_ids=None,
            exclude_ids=exclude_ids,
            status_label="新数据",
            max_rows=MAX_IMPORT_PREVIEW_ROWS,
        )
        
        if not unrecorded_rows:
            gr.Info("所有数据均已处理或标记")
            return (
                gr.update(), 
                gr.update(visible=False), 
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(visible=False),
                gr.update(interactive=True),
                update_import_stats(template),
                "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        sheet_cache = _build_sheet_row_cache(df, data_source.id_column)
        info_msg = f"找到 {len(unrecorded_rows)} 行未处理数据"
        if len(unrecorded_rows) >= MAX_IMPORT_PREVIEW_ROWS:
            info_msg += f"（已达到显示上限 {MAX_IMPORT_PREVIEW_ROWS} 行）"
        gr.Info(info_msg)
        
        return (
            gr.update(value=unrecorded_rows, visible=True),
            gr.update(visible=True),
            gr.update(visible=True),
            gr.update(visible=False),
            gr.update(visible=True),
            gr.update(interactive=True),
            update_import_stats(template),
            "unprocessed",
            *_import_refresh_tail(sheet_cache, True),
        )
        
    except GoogleSheetsError as e:
        gr.Warning(f"获取 Sheet 数据失败：{e}")
        return (
            gr.update(), 
            gr.update(visible=False), 
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(interactive=True),
            update_import_stats(template),
            "unprocessed",
            *_import_refresh_tail({}, False),
        )
    except Exception as e:
        logger.error(f"Refresh failed: {e}")
        gr.Warning(f"刷新失败：{str(e)}")
        return (
            gr.update(), 
            gr.update(visible=False), 
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(interactive=True),
            update_import_stats(template),
            "unprocessed",
            *_import_refresh_tail({}, False),
        )


def handle_import_selected(
    preview_data: list[list[Any]],
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    credentials: Any,
    detected_areas: list,
    sheet_cache: dict[str, dict[str, str]],
    entry_mode: str,
    sheet_name: str | None,
    session: dict[str, Any],
) -> tuple:
    """
    Import selected rows from preview using Gemma 4 field matching
    
    Returns:
        (form_data_state, row_selector, import_preview, import_btn, mark_trash_btn,
         import_stats, import_selection_state, import_preview_active_state, *field_boxes)
    """
    preview_rows = _normalize_preview_rows(preview_data)
    template_id = _get_template_id(template)
    if not preview_rows or not template_id or not template:
        return _import_failure_returns(form_data, template, session)
    
    try:
        selected_rows = [row for row in preview_rows if _is_row_selected(row)]
        
        if not selected_rows:
            gr.Warning("请勾选要导入的行")
            return _import_failure_returns(form_data, template, session)
        
        from app.services.data_source import load_template_data_source
        
        data_source = load_template_data_source(template_id)
        if not data_source or not credentials:
            gr.Warning("请先配置数据源")
            return _import_failure_returns(form_data, template, session)
        
        paste_config = load_paste_parse_config(template_id)
        if not paste_config:
            gr.Warning("模板配置加载失败")
            return _import_failure_returns(form_data, template, session)
        
        field_matcher = create_field_matcher()
        if field_matcher is None:
            gr.Warning("LLM 模型未加载，使用规则匹配")
        
        sheet_df = None
        if not sheet_cache:
            sheet_df = _load_template_sheet_df(credentials, data_source)
        imported_count = 0
        imported_ids = []
        headers = get_form_field_headers(template)
        
        for row in selected_rows:
            id_value = str(row[1])
            sheet_row = sheet_cache.get(id_value)
            if sheet_row is None:
                if sheet_df is None:
                    sheet_df = _load_template_sheet_df(credentials, data_source)
                sheet_row = lookup_row_by_id(
                    sheet_df,
                    data_source.id_column,
                    id_value,
                )
            
            if not sheet_row:
                logger.warning(f"Row not found for ID: {id_value}")
                continue
            
            if field_matcher is not None:
                matched_fields = field_matcher.match_sheet_fields_to_yaml(
                    sheet_row,
                    paste_config.to_dict()
                )
            else:
                matched_fields = map_sheet_row_from_paste_config(sheet_row, paste_config)
            
            if matched_fields:
                form_data.append(matched_fields)
                imported_ids.append(id_value)
                imported_count += 1
        
        if imported_ids:
            mark_as_processed(template_id, imported_ids)
        
        if imported_count > 0:
            gr.Info(f"✅ 成功导入 {imported_count} 行数据")
        else:
            gr.Warning("未能导入任何数据，请检查配置")
        
        row_choices = _build_row_choices_for_template(
            template,
            sheet_name,
            form_data,
            detected_areas or [],
        )
        row_selector_update = _row_selector_update(
            row_choices,
            row_choices[0] if row_choices else None,
            entry_mode=entry_mode,
        )
        field_updates = (
            tuple(_field_updates_for_row(headers, form_data[0]))
            if headers and form_data
            else _NO_CHANGE_FIELD_UPDATES
        )
        
        return (
            form_data,
            row_selector_update,
            gr.update(visible=False),
            gr.update(interactive=True),
            gr.update(interactive=True),
            update_import_stats(template),
            _empty_import_selection(),
            False,
            _mark_form_dirty(session or _empty_form_session()),
            *field_updates,
        )
        
    except Exception as e:
        logger.error(f"Import failed: {e}")
        gr.Warning(f"导入失败：{str(e)}")
        return _import_failure_returns(form_data, template, session)


def handle_show_processed(
    template: TemplateConfig | None,
    credentials: Any
) -> tuple:
    """
    Show processed data
    
    Returns:
        (import_preview, import_btn, mark_trash_btn, restore_btn,
         clear_history_btn, import_view_state)
    """
    if not template or not credentials:
        gr.Warning("请先配置数据源")
        return (
            gr.update(), gr.update(visible=False), gr.update(visible=False),
            gr.update(visible=False), gr.update(visible=False), "unprocessed",
            *_import_refresh_tail({}, False),
        )
    
    try:
        from app.services.data_source import load_template_data_source
        
        data_source = load_template_data_source(template.id)
        if not data_source:
            gr.Warning("模板未配置数据源")
            return (
                gr.update(), gr.update(visible=False), gr.update(visible=False),
                gr.update(visible=False), gr.update(visible=False), "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        history = load_import_history(template.id)
        
        if not history.processed_ids:
            gr.Info("没有已处理的数据")
            return (
                gr.update(), gr.update(visible=False), gr.update(visible=False),
                gr.update(visible=False), gr.update(visible=False), "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        df = _load_template_sheet_df(credentials, data_source)
        processed_rows = _build_import_preview_rows(
            df,
            data_source.id_column,
            include_ids=history.processed_ids,
            exclude_ids=None,
            status_label="已处理",
            max_rows=MAX_IMPORT_PREVIEW_ROWS,
        )
        
        gr.Info(f"找到 {len(processed_rows)} 行已处理数据")
        sheet_cache = _build_sheet_row_cache(df, data_source.id_column)
        
        return (
            gr.update(value=processed_rows, visible=True),
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(visible=True),
            gr.update(visible=True),
            "processed",
            *_import_refresh_tail(sheet_cache, True),
        )
    except Exception as e:
        logger.error(f"Show processed failed: {e}")
        gr.Warning(f"加载失败：{str(e)}")
        return (
            gr.update(), gr.update(visible=False), gr.update(visible=False),
            gr.update(visible=False), gr.update(visible=False), "unprocessed",
            *_import_refresh_tail({}, False),
        )


def handle_show_trash(
    template: TemplateConfig | None,
    credentials: Any
) -> tuple:
    """
    Show trash data
    
    Returns:
        (import_preview, import_btn, mark_trash_btn, restore_btn,
         clear_history_btn, import_view_state)
    """
    if not template or not credentials:
        gr.Warning("请先配置数据源")
        return (
            gr.update(), gr.update(visible=False), gr.update(visible=False),
            gr.update(visible=False), gr.update(visible=False), "unprocessed",
            *_import_refresh_tail({}, False),
        )
    
    try:
        from app.services.data_source import load_template_data_source
        
        data_source = load_template_data_source(template.id)
        if not data_source:
            gr.Warning("模板未配置数据源")
            return (
                gr.update(), gr.update(visible=False), gr.update(visible=False),
                gr.update(visible=False), gr.update(visible=False), "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        history = load_import_history(template.id)
        
        if not history.trash_ids:
            gr.Info("没有垃圾数据")
            return (
                gr.update(), gr.update(visible=False), gr.update(visible=False),
                gr.update(visible=False), gr.update(visible=False), "unprocessed",
                *_import_refresh_tail({}, False),
            )
        
        df = _load_template_sheet_df(credentials, data_source)
        trash_rows = _build_import_preview_rows(
            df,
            data_source.id_column,
            include_ids=history.trash_ids,
            exclude_ids=None,
            status_label="垃圾",
            max_rows=MAX_IMPORT_PREVIEW_ROWS,
        )
        
        gr.Info(f"找到 {len(trash_rows)} 行垃圾数据")
        sheet_cache = _build_sheet_row_cache(df, data_source.id_column)
        
        return (
            gr.update(value=trash_rows, visible=True),
            gr.update(visible=False),
            gr.update(visible=False),
            gr.update(visible=True),
            gr.update(visible=True),
            "trash",
            *_import_refresh_tail(sheet_cache, True),
        )
    except Exception as e:
        logger.error(f"Show trash failed: {e}")
        gr.Warning(f"加载失败：{str(e)}")
        return (
            gr.update(), gr.update(visible=False), gr.update(visible=False),
            gr.update(visible=False), gr.update(visible=False), "unprocessed",
            *_import_refresh_tail({}, False),
        )


def handle_mark_trash(
    preview_data: Any,
    template: TemplateConfig | dict[str, Any] | None
) -> tuple:
    """
    Mark selected rows as trash
    
    Returns:
        (import_preview, mark_trash_btn, import_stats)
    """
    preview_rows = _normalize_preview_rows(preview_data)
    template_id = _get_template_id(template)
    if not preview_rows or not template_id:
        return gr.update(), gr.update(interactive=True), update_import_stats(template)
    
    try:
        selected_rows = [row for row in preview_rows if _is_row_selected(row)]
        
        if not selected_rows:
            gr.Warning("请勾选要标记的行")
            return gr.update(), gr.update(interactive=True), update_import_stats(template)
        
        ids_to_mark = [str(row[1]) for row in selected_rows]
        
        if mark_as_trash(template_id, ids_to_mark):
            remaining_rows = [row for row in preview_rows if str(row[1]) not in ids_to_mark]
            try:
                gr.Info(f"已标记 {len(ids_to_mark)} 行为垃圾数据")
            except Exception as notify_err:
                logger.debug("Notification failed: %s", notify_err)
            return (
                gr.update(value=remaining_rows, visible=len(remaining_rows) > 0),
                gr.update(interactive=True),
                update_import_stats(template)
            )
        
        gr.Warning("标记失败")
        return gr.update(), gr.update(interactive=True), update_import_stats(template)
            
    except Exception as e:
        logger.error(f"Mark trash failed: {e}")
        gr.Warning(f"标记失败：{str(e)}")
        return gr.update(), gr.update(interactive=True), update_import_stats(template)


def handle_restore_selected(
    preview_data: Any,
    template: TemplateConfig | dict[str, Any] | None,
) -> tuple:
    """
    Restore selected rows from processed/trash back to unprocessed state
    
    Returns:
        (import_preview, restore_btn, import_stats)
    """
    preview_rows = _normalize_preview_rows(preview_data)
    template_id = _get_template_id(template)
    if not preview_rows or not template_id:
        return gr.update(), gr.update(interactive=True), update_import_stats(template)
    
    try:
        selected_rows = [row for row in preview_rows if _is_row_selected(row)]
        
        if not selected_rows:
            gr.Warning("请勾选要恢复的行")
            return gr.update(), gr.update(interactive=True), update_import_stats(template)
        
        ids_to_restore = [str(row[1]) for row in selected_rows]
        
        if unmark_ids(template_id, ids_to_restore):
            remaining_rows = [
                row for row in preview_rows if str(row[1]) not in ids_to_restore
            ]
            gr.Info(f"已恢复 {len(ids_to_restore)} 行为未处理状态")
            return (
                gr.update(value=remaining_rows, visible=len(remaining_rows) > 0),
                gr.update(interactive=True),
                update_import_stats(template),
            )
        
        gr.Warning("恢复失败")
        return gr.update(), gr.update(interactive=True), update_import_stats(template)
            
    except Exception as e:
        logger.error(f"Restore failed: {e}")
        gr.Warning(f"恢复失败：{str(e)}")
        return gr.update(), gr.update(interactive=True), update_import_stats(template)


def handle_clear_history(template: TemplateConfig | None) -> str:
    """
    Clear import history
    
    Returns:
        Updated stats markdown
    """
    if not template:
        return "📊 导入统计：未选择模板"
    
    try:
        if clear_history(template.id):
            gr.Info("✅ 已清空导入历史")
        else:
            gr.Warning("清空失败")
        
        return update_import_stats(template)
    except Exception as e:
        logger.error(f"Clear history failed: {e}")
        gr.Warning(f"清空失败：{str(e)}")
        return update_import_stats(template)


def handle_print_preview(
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    sheet_name: str | None,
    row_selector_value: str | None,
    *field_values: str,
) -> None:
    """Build a filled workbook and open the Windows print-preview dialog."""
    if not template:
        gr.Warning("请先选择模板")
        return None
    if not sheet_name:
        gr.Warning("请先选择工作表")
        return None
    try:
        merged_rows = _merge_field_boxes_into_form_data(
            template,
            sheet_name,
            form_data,
            row_selector_value,
            field_values,
        )
        if not merged_rows:
            gr.Warning("没有可打印的数据")
            return None
        xlsx_bytes = build_export_workbook_bytes(template, sheet_name, merged_rows)
        filename = build_export_filename(Path(template.file_path), merged_rows)
        export_path = persist_export_file(template.id, xlsx_bytes, filename)
        area = primary_print_area(export_path)
        if not area:
            gr.Warning("模板未定义打印区域")
            return None
        show_print_dialog(export_path, area.sheet, area.range)
        gr.Info("已打开打印预览")
        return None
    except Exception as e:
        logger.error(f"Print preview failed: {e}")
        gr.Warning(f"打印预览失败：{str(e)}")
        return None


def handle_export(
    template: TemplateConfig | None,
    form_data: list[dict[str, str]],
    sheet_name: str | None,
    row_selector_value: str | None,
    *field_values: str,
) -> str | None:
    """Export form data to a filled Excel copy and offer it for download."""
    if not template:
        gr.Warning("请先选择模板")
        return None
    if not sheet_name:
        gr.Warning("请先选择工作表")
        return None
    try:
        merged_rows = _merge_field_boxes_into_form_data(
            template,
            sheet_name,
            form_data,
            row_selector_value,
            field_values,
        )
        if not merged_rows:
            gr.Warning("没有可导出的数据")
            return None
        xlsx_bytes = build_export_workbook_bytes(template, sheet_name, merged_rows)
        filename = build_export_filename(Path(template.file_path), merged_rows)
        export_path = persist_export_file(template.id, xlsx_bytes, filename)
        gr.Info(f"已导出：{filename}")
        return str(export_path)
    except Exception as e:
        logger.error(f"Export failed: {e}")
        gr.Warning(f"导出失败：{str(e)}")
        return None
