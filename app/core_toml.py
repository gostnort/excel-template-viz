from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import tomlkit
import tomlkit.exceptions
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from tomlkit import aot, document, string, table

from app.core_registry import TEMPLATES_DIR


CORE_CONFIG_SUFFIX = ".toml"
DEFAULT_DETERMINER = "\t"
DEFAULT_SOURCES: list[dict[str, str | None]] = [{"source1": None}]
DEFAULT_VALUE_FROM_LABEL = "down"
DEFAULT_VALUE_OFFSET = 1
DEFAULT_MOVE_TO = "down"
DEFAULT_INPUT_OFFSET = 1
FIELD_LABEL_KEY = "Input_label"
OPTIONAL_FIELD_KEYS = ("field", "source_file", "source_sheet", "regex")
VALID_DIRECTIONS = {"up", "down", "left", "right"}
VERIFY_SCAN_ROWS = 100
VERIFY_SCAN_COLS = 100


@dataclass
class InputSection:
    """One [[input_section]] row."""

    input_area: str
    move_to: str = DEFAULT_MOVE_TO
    offset: int = DEFAULT_INPUT_OFFSET


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: InputSection.to_dict
        作用: 将单条 [[input_section]] 转换为可序列化字典
        输入:
            无
        输出:
            dict[str, Any] - 含 input_area / move_to / offset 的字典
        """
        return {
            "input_area": self.input_area,  # instance 0 填写值区域
            "move_to": self.move_to,        # 第 k≥1 组值格平移方向
            "offset": self.offset,          # 平移步长
        }



@dataclass
class TomlDefault:
    """One [[fields]] row. Index is ZERO base."""

    Input_label: str
    value_from_label: str = DEFAULT_VALUE_FROM_LABEL
    value_offset: int = DEFAULT_VALUE_OFFSET
    field: str | None = None
    source_file: str | None = None
    source_sheet: str | None = None
    index: int = -1
    regex: str | None = None
    id: bool = False


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: TomlDefault.to_dict
        作用: 将一条 [[fields]] 字段规则转换为可序列化字典；未映射可选键取 None
        输入:
            无
        输出:
            dict[str, Any] - 字段规则的字典形式（内存语义，未映射为 None）
        """
        return {
            FIELD_LABEL_KEY: self.Input_label,
            "value_from_label": self.value_from_label,  # 标签到值格的方向
            "value_offset": self.value_offset,          # 标签到值格的步长
            "field": self.field if not _is_unmapped(self.field) else None,
            "source_file": self.source_file if not _is_unmapped(self.source_file) else None,
            "source_sheet": self.source_sheet if not _is_unmapped(self.source_sheet) else None,
            "index": self.index,
            "regex": self.regex if not _is_unmapped(self.regex) else None,
            "id": self.id,
        }



@dataclass
class VerifyTomlReport:
    """Structured report returned by verify_toml()."""

    ok: bool
    missing_labels: list[str]
    duplicate_labels: list[str]
    out_of_area_labels: list[str]
    located: dict[str, dict[str, int]]
    errors: list[str]
    duplicate_id_sheets: list[str]
    db_id_required: bool
    invalid_db_id: str | None
    db_id: str | None
    id_labels: list[str]
    id_lookup_keys: list[str]


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: VerifyTomlReport.to_dict
        作用: 将校验报告转换为给 UI 的字典契约（拷贝内部列表/字典避免外部改动）
        输入:
            无
        输出:
            dict[str, Any] - 含坐标、id 规则与 db_id 相关字段
        """
        return {
            "ok": self.ok,
            "missing_labels": list(self.missing_labels),
            "duplicate_labels": list(self.duplicate_labels),
            "out_of_area_labels": list(self.out_of_area_labels),
            "located": dict(self.located),
            "errors": list(self.errors),
            "duplicate_id_sheets": list(self.duplicate_id_sheets),
            "db_id_required": self.db_id_required,
            "invalid_db_id": self.invalid_db_id,
            "db_id": self.db_id,
            "id_labels": list(self.id_labels),
            "id_lookup_keys": list(self.id_lookup_keys),
        }


def _core_toml_path(template_id: str) -> Path:
    """
    函数名: _core_toml_path
    作用: 根据模板 ID 拼出其 TOML 配置文件的磁盘路径
    输入:
        template_id (str) - 模板唯一标识
    输出:
        Path - templates/{id}/{id}.toml 的 Path
    """
    return TEMPLATES_DIR / template_id / f"{template_id}{CORE_CONFIG_SUFFIX}"


def _is_unmapped(value: Any) -> bool:
    """
    函数名: _is_unmapped
    作用: 判断一个可选键的值是否视为未映射（None 或空串）
    输入:
        value (Any) - 待判定的值
    输出:
        bool - 未映射返回 True
    """
    return value is None or value == ""


def _needs_literal_string(key: str, value: str) -> bool:
    """
    函数名: _needs_literal_string
    作用: 判断该键值是否需要写成 TOML 字面量字符串（单引号，避免反斜杠转义）
    输入:
        key (str) - 字段键名，regex 始终用字面量
        value (str) - 字段字符串值，含反斜杠时用字面量
    输出:
        bool - 需要字面量返回 True
    """
    # 值本身含单引号时无法用字面量，退回普通字符串
    if "'" in value:
        return False
    if key == "regex":
        return True
    if "\\" in value:
        return True
    return False


def _toml_string(key: str, value: Any) -> Any:
    """
    函数名: _toml_string
    作用: 把内存值转换为可写盘的 TOML 字符串；未映射写空串，必要时用字面量
    输入:
        key (str) - 字段键名，决定是否倾向字面量
        value (Any) - 内存值，None/"" 视为未映射
    输出:
        Any - 普通字符串或 tomlkit 字面量字符串
    """
    if _is_unmapped(value):
        return ""
    text = str(value)  # 统一成文本再判断写法
    if _needs_literal_string(key, text):
        return string(text, literal=True)
    return text


def _extract_toml_text(model_output: str) -> str:
    """
    函数名: _extract_toml_text
    作用: 从可能带 Markdown 代码块围栏的文本中提取纯 TOML
    输入:
        model_output (str) - 用户或 LLM 提供的原始文本
    输出:
        str - 去除围栏后的 TOML 文本
    """
    text = model_output.strip()
    if not text:
        return text
    fenced = re.search(
        r"```(?:toml)?\s*\r?\n?(.*?)```", text, flags=re.DOTALL | re.IGNORECASE
    )  # 优先匹配成对围栏
    if fenced:
        return fenced.group(1).strip()
    # 仅有起始围栏时逐段剥除
    if text.startswith("```"):
        text = re.sub(r"^```(?:toml)?\s*\r?\n?", "", text, count=1, flags=re.IGNORECASE)
        text = re.sub(r"```\s*$", "", text.strip())
    return text.strip()


def _parse_bool(value: Any) -> bool:
    """
    函数名: _parse_bool
    作用: 把布尔/字符串/数字宽松解析为 bool
    输入:
        value (Any) - 待解析值
    输出:
        bool - 解析结果
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("true", "1", "yes")
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _parse_int(value: Any, default: int) -> int:
    """
    函数名: _parse_int
    作用: 把值解析为 int，失败时回退默认值
    输入:
        value (Any) - 待解析值
        default (int) - 解析失败时的回退值
    输出:
        int - 解析结果或默认值
    """
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _optional_string(value: Any) -> str | None:
    """
    函数名: _optional_string
    作用: 规范可选字符串键——未映射转为 None，其余去空白
    输入:
        value (Any) - 原始值
    输出:
        str | None - 未映射为 None，否则为去空白文本
    """
    if _is_unmapped(value):
        return None
    return str(value).strip()


def _resolve_work_sheet(raw: dict[str, Any]) -> str | None:
    """
    函数名: _resolve_work_sheet
    作用: 解析顶层 work_sheet；兼容旧键 worksheet
    输入:
        raw (dict[str, Any]) - tomlkit 解析得到的字典
    输出:
        str | None - 数据录入工作表名
    """
    work_sheet = _optional_string(raw.get("work_sheet"))
    if work_sheet:
        return work_sheet
    return _optional_string(raw.get("worksheet"))


def _parse_sources(raw_sources: Any) -> list[dict[str, str | None]]:
    """
    函数名: _parse_sources
    作用: 解析 [[sources]]，把空串别名规范成 None；无有效项时回退默认
    输入:
        raw_sources (Any) - 原始 sources 数据
    输出:
        list[dict[str, str | None]] - 规范化后的数据源别名列表
    """
    if not isinstance(raw_sources, list) or not raw_sources:
        return [dict(item) for item in DEFAULT_SOURCES]
    sources: list[dict[str, str | None]] = []
    for item in raw_sources:
        if not isinstance(item, dict):
            continue
        parsed_item: dict[str, str | None] = {}
        for key, value in item.items():
            parsed_item[str(key)] = _optional_string(value)  # 空串别名转 None
        if parsed_item:
            sources.append(parsed_item)
    if not sources:
        return [dict(item) for item in DEFAULT_SOURCES]
    return sources


def _input_section_from_dict(raw: Any) -> InputSection | None:
    """
    函数名: _input_section_from_dict
    作用: 解析有且仅有一条的 [[input_section]]；缺 input_area 视为无效
    输入:
        raw (Any) - input_section 原始数据（dict 或仅含一项的 list）
    输出:
        InputSection | None - 解析成功的区段，或 None
    """
    # tomlkit 的 [[input_section]] 会读成 list，按设计只允许一条
    if isinstance(raw, list):
        if len(raw) != 1:
            return None
        raw = raw[0]
    if not isinstance(raw, dict):
        return None
    input_area = str(raw.get("input_area", "")).strip()
    if not input_area:
        return None
    move_to = str(raw.get("move_to", DEFAULT_MOVE_TO)).strip().lower()  # 方向归一化
    offset = _parse_int(raw.get("offset", DEFAULT_INPUT_OFFSET), DEFAULT_INPUT_OFFSET)
    return InputSection(input_area=input_area, move_to=move_to, offset=offset)


def _field_from_dict(raw: Any) -> TomlDefault | None:
    """
    函数名: _field_from_dict
    作用: 解析一条 [[fields]]；缺少 Input_label 视为无效
    输入:
        raw (Any) - 单条字段映射原始数据
    输出:
        TomlDefault | None - 解析成功的字段规则，或 None
    """
    if not isinstance(raw, dict):
        return None
    input_label = str(raw.get(FIELD_LABEL_KEY, "")).strip()
    if not input_label:
        return None
    value_from_label = str(
        raw.get("value_from_label", DEFAULT_VALUE_FROM_LABEL)
    ).strip().lower()  # 方向归一化为小写
    value_offset = _parse_int(raw.get("value_offset", DEFAULT_VALUE_OFFSET), DEFAULT_VALUE_OFFSET)
    index_val = _parse_int(raw.get("index", -1), -1)  # -1 表示不参与文本拆分
    return TomlDefault(
        Input_label=input_label,
        value_from_label=value_from_label,
        value_offset=value_offset,
        field=_optional_string(raw.get("field")),
        source_file=_optional_string(raw.get("source_file")),
        source_sheet=_optional_string(raw.get("source_sheet")),
        index=index_val,
        regex=_optional_string(raw.get("regex")),
        id=_parse_bool(raw.get("id", False)),
    )


def _config_from_dict(raw: dict[str, Any]) -> GetTomlValues | None:
    """
    函数名: _config_from_dict
    作用: 由解析后的 TOML 字典构造 GetTomlValues；缺 work_sheet/input_section/fields 则失败
    输入:
        raw (dict[str, Any]) - tomlkit 解析得到的字典
    输出:
        GetTomlValues | None - 构造成功的配置对象，或 None
    """
    field_rules: list[TomlDefault] = []
    fields_raw = raw.get("fields")
    if isinstance(fields_raw, list):
        for item in fields_raw:
            rule = _field_from_dict(item)
            if rule:
                field_rules.append(rule)
    input_section = _input_section_from_dict(raw.get("input_section"))
    work_sheet = _resolve_work_sheet(raw)
    print_sheet = _optional_string(raw.get("print_sheet"))
    # 三项必备：work_sheet、单条 input_section、至少一条 fields
    if not work_sheet or input_section is None or not field_rules:
        return None
    determiner = str(raw.get("determiner", DEFAULT_DETERMINER)) or DEFAULT_DETERMINER
    db_id = _optional_string(raw.get("db_id"))
    use_independent_db = _parse_bool(raw.get("use_independent_db", True))
    return GetTomlValues(
        determiner=determiner,
        sources=_parse_sources(raw.get("sources")),
        field_rules=field_rules,
        work_sheet=work_sheet,
        print_sheet=print_sheet,
        input_section=input_section,
        db_id=db_id,
        use_independent_db=use_independent_db,
    )


def _dict_to_toml(config: dict[str, Any]) -> str:
    """
    函数名: _dict_to_toml
    作用: 把配置字典序列化为严格 TOML 1.0 文本；未映射可选键写空串 ""，不写 null
    输入:
        config (dict[str, Any]) - 含 determiner/work_sheet/print_sheet/sources/input_section/fields
    输出:
        str - 序列化后的 TOML 文本（以换行结尾）
    """
    # ---- 顶层标量：determiner、work_sheet、print_sheet ----
    doc = document()
    doc["determiner"] = str(config.get("determiner", DEFAULT_DETERMINER))
    work_sheet = config.get("work_sheet", config.get("worksheet", ""))
    doc["work_sheet"] = str(work_sheet or "")
    print_sheet = config.get("print_sheet")
    if print_sheet is not None and str(print_sheet).strip():
        doc["print_sheet"] = str(print_sheet).strip()
    db_id = config.get("db_id")
    if db_id is not None and str(db_id).strip():
        doc["db_id"] = str(db_id).strip()
    doc["use_independent_db"] = config.get("use_independent_db", True)
    # ---- [[sources]]：空别名落盘为空串 ----
    sources_aot = aot()
    for item in config.get("sources", DEFAULT_SOURCES):
        if not isinstance(item, dict):
            continue
        src = table()
        for source_key, source_value in item.items():
            src[str(source_key)] = _toml_string(str(source_key), source_value)
        sources_aot.append(src)
    doc["sources"] = sources_aot
    # ---- [[input_section]]：有且仅有一条 ----
    section = _input_section_from_dict(config.get("input_section"))
    input_sections = aot()
    if section:
        section_row = table()
        section_row["input_area"] = section.input_area
        section_row["move_to"] = section.move_to
        section_row["offset"] = section.offset
        input_sections.append(section_row)
    doc["input_section"] = input_sections
    # ---- [[fields]]：逐条写必有键与可选键 ----
    fields_aot = aot()
    for item in config.get("fields", []):
        rule = _field_from_dict(item) if isinstance(item, dict) else item  # 兼容 dict 与 TomlDefault
        if not isinstance(rule, TomlDefault):
            continue
        row = table()
        row[FIELD_LABEL_KEY] = rule.Input_label
        row["value_from_label"] = rule.value_from_label
        row["value_offset"] = rule.value_offset
        rule_dict = rule.to_dict()
        for key in OPTIONAL_FIELD_KEYS:
            row[key] = _toml_string(key, rule_dict.get(key))  # 未映射写 ""
        row["index"] = rule.index
        row["id"] = rule.id
        fields_aot.append(row)
    doc["fields"] = fields_aot
    dumped = tomlkit.dumps(doc)
    if not dumped.endswith("\n"):
        dumped += "\n"
    return dumped


def _parse_area(area: str) -> tuple[int, int, int, int]:
    """
    函数名: _parse_area
    作用: 把 Excel 区域字符串解析为 (min_row, min_col, max_row, max_col)
    输入:
        area (str) - 形如 "A2:G2" 的区域
    输出:
        tuple[int, int, int, int] - 1-based 行列边界
    """
    # range_boundaries 返回 (min_col, min_row, max_col, max_row)，此处换序为行列
    min_col, min_row, max_col, max_row = range_boundaries(area)
    return min_row, min_col, max_row, max_col


def _cell_text(value: Any) -> str:
    """
    函数名: _cell_text
    作用: 把单元格值规范化为用于标签比对的去空白文本
    输入:
        value (Any) - 单元格原始值
    输出:
        str - None 转空串，其余去空白
    """
    if value is None:
        return ""
    return str(value).strip()


def offset_cell(row: int, col: int, direction: str, offset: int) -> tuple[int, int]:
    """
    函数名: offset_cell
    作用: 在 1-based 坐标上按方向与步长平移，越界或非法方向时抛错
    输入:
        row (int) - 起始行（1-based）
        col (int) - 起始列（1-based）
        direction (str) - up/down/left/right
        offset (int) - 正整数步长
    输出:
        tuple[int, int] - 平移后的 (row, col)
    """
    direction = direction.strip().lower()  # 容忍大小写与空白
    if direction not in VALID_DIRECTIONS:
        raise ValueError(f"unsupported direction: {direction}")
    if offset < 1:
        raise ValueError("offset must be a positive integer")
    if direction == "up":
        row -= offset
    elif direction == "down":
        row += offset
    elif direction == "left":
        col -= offset
    elif direction == "right":
        col += offset
    if row < 1 or col < 1:
        raise ValueError("offset produced an out-of-bounds coordinate")
    return row, col


def _scan_worksheet_labels_diagonal(
    ws: Any,
) -> tuple[dict[str, tuple[int, int]], set[str]]:
    """
    函数名: _scan_worksheet_labels_diagonal
    作用: 按左上→右下斜向波面顺序扫描 100×100，建立标签文本到首见坐标的映射，并记录重复文本
    输入:
        ws (Any) - openpyxl 工作表对象
    输出:
        tuple[dict[str, tuple[int, int]], set[str]] - (标签→(row,col), 重复出现的标签文本集合)
    """
    # 流式拉取 100×100 快照，避免逐格随机读
    grid: list[list[Any]] = []
    for row in ws.iter_rows(max_row=VERIFY_SCAN_ROWS, max_col=VERIFY_SCAN_COLS):
        grid.append([cell.value for cell in row])
    label_to_coord: dict[str, tuple[int, int]] = {}
    duplicate_labels: set[str] = set()
    max_r = len(grid)
    max_c = len(grid[0]) if max_r > 0 else 0
    if max_r == 0 or max_c == 0:
        return label_to_coord, duplicate_labels
    # s = r_idx + c_idx；同一条斜线上 r_idx 小者优先（更靠上）
    for s in range(max_r + max_c - 1):
        for r_idx in range(max_r):
            c_idx = s - r_idx
            if c_idx < 0 or c_idx >= max_c:
                continue
            text = _cell_text(grid[r_idx][c_idx])
            if not text:
                continue
            excel_row = r_idx + 1  # 转回 Excel 1-based 坐标
            excel_col = c_idx + 1
            if text in label_to_coord:
                duplicate_labels.add(text)  # 后续斜向位置再次出现
            else:
                label_to_coord[text] = (excel_row, excel_col)  # 首见即最靠近左上角
    return label_to_coord, duplicate_labels


def _cell_in_area(row: int, col: int, area: tuple[int, int, int, int]) -> bool:
    """
    函数名: _cell_in_area
    作用: 判断坐标是否落在区域矩形内
    输入:
        row (int) - 行（1-based）
        col (int) - 列（1-based）
        area (tuple[int, int, int, int]) - (min_row, min_col, max_row, max_col)
    输出:
        bool - 在区域内返回 True
    """
    min_row, min_col, max_row, max_col = area
    return min_row <= row <= max_row and min_col <= col <= max_col


def _headers_from_first_row(ws: Any) -> list[tuple[int, str]]:
    """
    函数名: _headers_from_first_row
    作用: 读取第 1 行非空单元格，作为标准范式默认配置的字段标签
    输入:
        ws (Any) - openpyxl 工作表对象
    输出:
        list[tuple[int, str]] - (列号, 标签文本) 列表
    """
    headers: list[tuple[int, str]] = []
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        text = _cell_text(ws.cell(row=1, column=col).value)  # 第 1 行该列文本
        if text:
            headers.append((col, text))
    return headers


def _source_ref_key(source_file: str, source_sheet: str) -> str:
    """
    函数名: _source_ref_key
    作用: 拼出外部数据源表引用键 source_file/source_sheet
    输入:
        source_file (str) - sources 别名
        source_sheet (str) - 外部工作表名
    输出:
        str - 如 source1/sheet1
    """
    return f"{source_file}/{source_sheet}"


def _id_lookup_key(rule: TomlDefault) -> str:
    """
    函数名: _id_lookup_key
    作用: 取外部表查行用的列名；已映射 field 优先，否则用 Input_label
    输入:
        rule (TomlDefault) - id=true 的字段规则
    输出:
        str - 查行键
    """
    if not _is_unmapped(rule.field):
        return rule.field  # type: ignore[return-value]
    return rule.Input_label


def _validate_id_rules(cfg: GetTomlValues) -> dict[str, Any]:
    """
    函数名: _validate_id_rules
    作用: 纯 TOML 层校验 id 规则：每外部表至多一个 id=true、汇总 id_lookup_keys、解析 db_id
    输入:
        cfg (GetTomlValues) - 已加载配置
    输出:
        dict[str, Any] - duplicate_id_sheets / db_id_required / invalid_db_id / db_id / id_labels / id_lookup_keys
    """
    duplicate_id_sheets: list[str] = []
    sheet_id_count: dict[str, int] = {}
    id_labels: list[str] = []
    id_lookup_keys: list[str] = []
    seen_lookup: set[str] = set()
    for rule in (cfg.field_rules or []):
        if not rule.id:
            continue
        id_labels.append(rule.Input_label)
        lookup = _id_lookup_key(rule)
        if lookup not in seen_lookup:
            seen_lookup.add(lookup)
            id_lookup_keys.append(lookup)
        if not _is_unmapped(rule.source_file) and not _is_unmapped(rule.source_sheet):
            ref = _source_ref_key(rule.source_file, rule.source_sheet)  # type: ignore[arg-type]
            sheet_id_count[ref] = sheet_id_count.get(ref, 0) + 1
    for ref, count in sheet_id_count.items():
        if count > 1:
            duplicate_id_sheets.append(ref)
    db_id_required = False
    invalid_db_id: str | None = None
    resolved_db_id: str | None = None
    if len(id_labels) == 1:
        resolved_db_id = id_labels[0]
        if cfg.db_id and cfg.db_id != resolved_db_id:
            invalid_db_id = cfg.db_id
    elif len(id_labels) > 1:
        if _is_unmapped(cfg.db_id):
            db_id_required = True
        elif cfg.db_id not in id_labels:
            invalid_db_id = cfg.db_id
        else:
            resolved_db_id = cfg.db_id
    return {
        "duplicate_id_sheets": duplicate_id_sheets,
        "db_id_required": db_id_required,
        "invalid_db_id": invalid_db_id,
        "db_id": resolved_db_id,
        "id_labels": id_labels,
        "id_lookup_keys": id_lookup_keys,
    }


def resolve_db_id(cfg: GetTomlValues) -> str | None:
    """
    函数名: resolve_db_id
    作用: 解析本地 records 主键所对应的 Input_label；无 id 字段时返回 None
    输入:
        cfg (GetTomlValues) - 已加载配置
    输出:
        str | None - 生效的 db_id（Input_label），无则 None
    """
    id_info = _validate_id_rules(cfg)
    return id_info["db_id"]


def _id_ok(id_info: dict[str, Any]) -> bool:
    """id 规则是否通过（无重复表 id、db_id 合法）。"""
    if id_info["duplicate_id_sheets"]:
        return False
    if id_info["db_id_required"]:
        return False
    if id_info["invalid_db_id"]:
        return False
    return True


def _make_report(
    ok: bool,
    missing_labels: list[str],
    duplicate_labels: list[str],
    out_of_area_labels: list[str],
    located: dict[str, dict[str, int]],
    errors: list[str],
    id_info: dict[str, Any],
) -> dict[str, Any]:
    """
    函数名: _make_report
    作用: 合并模板坐标结果与 id 规则结果，构造 verify_toml 报告字典
    输入:
        ok (bool) - 模板坐标部分是否通过
        missing_labels / duplicate_labels / out_of_area_labels / located / errors - 坐标校验
        id_info (dict) - _validate_id_rules 返回值
    输出:
        dict[str, Any] - verify_toml 完整报告
    """
    full_ok = ok and _id_ok(id_info)
    return VerifyTomlReport(
        full_ok,
        missing_labels,
        duplicate_labels,
        out_of_area_labels,
        located,
        errors,
        id_info["duplicate_id_sheets"],
        id_info["db_id_required"],
        id_info["invalid_db_id"],
        id_info["db_id"],
        id_info["id_labels"],
        id_info["id_lookup_keys"],
    ).to_dict()


def _validate_field_regexes(cfg: GetTomlValues) -> list[str]:
    """
    函数名: _validate_field_regexes
    作用: 校验各 [[fields]].regex 是否为合法 Python 正则；失败时带上 Input_label
    输入:
        cfg (GetTomlValues) - 已加载配置
    输出:
        list[str] - 错误信息列表，每项含 Input_label
    """
    errors: list[str] = []
    for rule in cfg.field_rules:
        pattern = rule.regex
        if _is_unmapped(pattern):
            continue
        try:
            re.compile(str(pattern))
        except re.error as exc:
            errors.append(f"{rule.Input_label}: 正则表达式无效 ({pattern!r}): {exc}")
    return errors


def load_toml(template_id: str) -> GetTomlValues | None:
    """
    函数名: load_toml
    作用: 按模板 ID 读取并解析 TOML（只解析文本，不打开 xlsx 校验）
    输入:
        template_id (str) - 模板唯一标识
    输出:
        GetTomlValues | None - 解析成功的配置，或 None（不存在/解析失败/结构不全）
    """
    path = _core_toml_path(template_id)
    if not path.exists():
        return None
    try:
        raw = tomlkit.loads(path.read_text(encoding="utf-8"))
    except tomlkit.exceptions.ParseError:
        return None
    if not isinstance(raw, dict):
        return None
    return _config_from_dict(raw)


def verify_toml(template_path: Path, cfg: GetTomlValues) -> dict[str, Any]:
    """
    函数名: verify_toml
    作用: UI 唯一校验入口；模板坐标印证 + TOML 层 id/db_id 规则校验
    输入:
        template_path (Path) - 当前模板 xlsx 路径
        cfg (GetTomlValues) - 已加载的 TOML 配置
    输出:
        dict[str, Any] - 完整报告（坐标、duplicate_id_sheets、db_id 等）
    """
    missing_labels: list[str] = []
    duplicate_labels: list[str] = []
    out_of_area_labels: list[str] = []
    located: dict[str, dict[str, int]] = {}
    errors: list[str] = []
    id_info = _validate_id_rules(cfg)  # TOML 层 id 规则，不依赖 xlsx
    regex_errors = _validate_field_regexes(cfg)
    errors.extend(regex_errors)
    if not cfg.work_sheet:
        errors.append("work_sheet is required")
        return _make_report(False, [], [], [], {}, errors, id_info)
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        if cfg.work_sheet not in wb.sheetnames:
            errors.append(f"work_sheet not found: {cfg.work_sheet}")
            return _make_report(False, [], [], [], {}, errors, id_info)
        ws = wb[cfg.work_sheet]
        try:
            input_area = _parse_area(cfg.input_section.input_area)
        except ValueError as exc:
            errors.append(f"invalid input_area: {exc}")
            return _make_report(False, [], [], [], {}, errors, id_info)
        label_map, duplicate_texts = _scan_worksheet_labels_diagonal(ws)
        for rule in cfg.field_rules:
            if rule.Input_label not in label_map:
                missing_labels.append(rule.Input_label)
                continue
            if rule.Input_label in duplicate_texts:
                duplicate_labels.append(rule.Input_label)
                continue
            label_row, label_col = label_map[rule.Input_label]
            try:
                value_row, value_col = offset_cell(
                    label_row, label_col, rule.value_from_label, rule.value_offset
                )
            except ValueError as exc:
                errors.append(f"{rule.Input_label}: {exc}")
                continue
            if not _cell_in_area(value_row, value_col, input_area):
                out_of_area_labels.append(rule.Input_label)
                continue
            located[rule.Input_label] = {
                "label_row": label_row,
                "label_col": label_col,
                "value_row": value_row,
                "value_col": value_col,
            }
    finally:
        wb.close()
    layout_ok = not missing_labels and not duplicate_labels and not out_of_area_labels and not errors
    return _make_report(
        layout_ok, missing_labels, duplicate_labels, out_of_area_labels, located, errors, id_info
    )


def ensure_exists(
    template_id: str,
    template_path: Path,
    worksheet_name: str | None = None,
) -> bool:
    """
    函数名: ensure_exists
    作用: TOML 不存在或无法解析为合法配置时，按标准范式生成默认配置
    输入:
        template_id (str) - 模板唯一标识
        template_path (Path) - 模板 xlsx 路径
        worksheet_name (str | None) - 目标工作表名，None 时取 active sheet
    输出:
        bool - 已有合法 TOML 或生成成功返回 True
    """
    path = _core_toml_path(template_id)
    if path.exists() and load_toml(template_id) is not None:
        return True
    if not TomlGenerator().Reset(template_id, template_path, worksheet_name):
        return False
    return load_toml(template_id) is not None



class TomlGenerator:
    """First-time creation of default TOML config and serialization."""

    def _worksheet(self, template_path: Path, worksheet_name: str | None) -> tuple[str, Any, Any]:
        """
        函数名: TomlGenerator._worksheet
        作用: 打开工作簿并定位目标工作表；未指定名时取 active sheet
        输入:
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            tuple[str, Any, Any] - (解析出的表名, 工作表对象, 工作簿对象)
        """
        wb = load_workbook(template_path, read_only=True, data_only=True)
        if worksheet_name:
            # 指定名但不存在时关闭工作簿再抛错，避免句柄泄漏
            if worksheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"worksheet not found: {worksheet_name}")
            return worksheet_name, wb[worksheet_name], wb
        ws = wb.active
        return ws.title, ws, wb


    def CreateDefaultFromTemplate(
        self,
        template_path: Path,
        worksheet_name: str | None = None,
    ) -> dict[str, Any]:
        """
        函数名: TomlGenerator.CreateDefaultFromTemplate
        作用: 按标准范式生成默认配置——第 1 行作字段标签，input_area 登记第 2 行值区；不逐格扫描
        输入:
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            dict[str, Any] - 默认配置字典
        """
        resolved_name, ws, wb = self._worksheet(template_path, worksheet_name)
        try:
            headers = _headers_from_first_row(ws)  # 第 1 行非空标签
            if not headers:
                # active/指定表无表头时，改用首个第 1 行有标签的工作表
                for name in wb.sheetnames:
                    if name == resolved_name:
                        continue
                    alt_ws = wb[name]
                    alt_headers = _headers_from_first_row(alt_ws)
                    if alt_headers:
                        resolved_name = name
                        ws = alt_ws
                        headers = alt_headers
                        break
            fields = [
                TomlDefault(Input_label=label, index=-1).to_dict()  # 默认不参与文本拆分，见 toml_config_design
                for _idx, (_col, label) in enumerate(headers)
            ]
            # input_area 取标签覆盖的列范围、固定第 2 行
            if headers:
                start_col = headers[0][0]
                end_col = headers[-1][0]
            else:
                start_col = 1
                end_col = 1
            input_area = f"{get_column_letter(start_col)}2:{get_column_letter(end_col)}2"
            return {
                "determiner": DEFAULT_DETERMINER,
                "work_sheet": resolved_name,
                "print_sheet": "",
                "use_independent_db": True,
                "sources": [dict(item) for item in DEFAULT_SOURCES],
                "input_section": InputSection(input_area=input_area).to_dict(),
                "fields": fields,
            }
        finally:
            wb.close()


    def ConfigToToml(self, config: dict[str, Any]) -> str:
        """
        函数名: TomlGenerator.ConfigToToml
        作用: 把配置字典序列化为严格 TOML 1.0 文本
        输入:
            config (dict[str, Any]) - 配置字典
        输出:
            str - TOML 文本
        """
        return _dict_to_toml(config)


    def Reset(
        self,
        template_id: str,
        template_path: Path,
        worksheet_name: str | None = None,
    ) -> bool:
        """
        函数名: TomlGenerator.Reset
        作用: 用标准范式默认配置覆盖写入模板 TOML 文件
        输入:
            template_id (str) - 模板唯一标识
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            bool - 写入成功返回 True，异常返回 False
        """
        path = _core_toml_path(template_id)
        try:
            default_cfg = self.CreateDefaultFromTemplate(template_path, worksheet_name)
            path.parent.mkdir(parents=True, exist_ok=True)  # 确保目录存在
            path.write_text(self.ConfigToToml(default_cfg), encoding="utf-8")
            return True
        except Exception:
            return False



class GetTomlValues:
    """Loaded TOML config: query, modify in memory, persist via Save / ToDict."""

    def __init__(
        self,
        determiner: str = DEFAULT_DETERMINER,
        sources: list[dict[str, str | None]] | None = None,
        field_rules: list[TomlDefault] | None = None,
        work_sheet: str | None = None,
        print_sheet: str | None = None,
        input_section: InputSection | None = None,
        db_id: str | None = None,
        use_independent_db: bool = True,
    ) -> None:
        """
        函数名: GetTomlValues.__init__
        作用: 构造已加载配置对象；可选参数缺省时给出安全默认
        输入:
            determiner (str) - 文本拆分分隔符
            sources (list[dict[str, str | None]] | None) - 数据源别名列表
            field_rules (list[TomlDefault] | None) - 字段规则列表
            work_sheet (str | None) - 数据录入 / TOML 定位工作表名
            print_sheet (str | None) - UI 打印区选择所用工作表名
            input_section (InputSection | None) - 单条区段
            db_id (str | None) - 本地 records 主键对应的 Input_label；多 id 时必填
            use_independent_db (bool) - 是否使用独立数据库
        输出:
            无
        """
        self.determiner = determiner
        self.sources = sources if sources is not None else [dict(item) for item in DEFAULT_SOURCES]
        self.field_rules = field_rules if field_rules is not None else []
        self.work_sheet = work_sheet
        self.print_sheet = print_sheet
        self.input_section = input_section if input_section is not None else InputSection("A2:A2")
        self.db_id = db_id
        self.use_independent_db = use_independent_db


    def Load(self, template_id: str) -> GetTomlValues | None:
        """
        函数名: GetTomlValues.Load
        作用: 按模板 ID 读取 TOML（仅解析文本，不打开 xlsx 校验）
        输入:
            template_id (str) - 模板唯一标识
        输出:
            GetTomlValues | None - 解析成功的配置，或 None
        """
        return load_toml(template_id)


    def Save(self, template_id: str, toml_text: str | None = None) -> None:
        """
        函数名: GetTomlValues.Save
        作用: 落盘配置；toml_text 为 None 时保存当前实例，否则校验文本结构后再写
        输入:
            template_id (str) - 模板唯一标识
            toml_text (str | None) - 用户/LLM 提供的 TOML 文本，None 表示用当前实例
        输出:
            无
        """
        if toml_text is None:
            cleaned = TomlGenerator().ConfigToToml(self.ToDict())  # 直接序列化当前实例
        else:
            cleaned = _extract_toml_text(toml_text)  # 去掉可能的 Markdown 围栏
            try:
                parsed = tomlkit.loads(cleaned)
            except tomlkit.exceptions.ParseError as exc:
                raise ValueError(f"Invalid TOML: {exc}") from exc
            # 结构须符合当前设计（有 work_sheet/input_section/fields）
            if not isinstance(parsed, dict) or _config_from_dict(parsed) is None:
                raise ValueError("TOML must match the current core_toml design")
            cleaned = TomlGenerator().ConfigToToml(parsed)  # 归一化后再写
        path = _core_toml_path(template_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(cleaned, encoding="utf-8")


    def VerifyToml(self, template_path: Path) -> dict[str, Any]:
        """
        函数名: GetTomlValues.VerifyToml
        作用: 以本实例为配置，对模板 xlsx 调用 verify_toml
        输入:
            template_path (Path) - 模板 xlsx 路径
        输出:
            dict[str, Any] - 校验报告
        """
        return verify_toml(template_path, self)


    def ToDict(self) -> dict[str, Any]:
        """
        函数名: GetTomlValues.ToDict
        作用: 把当前配置转换为可序列化字典
        输入:
            无
        输出:
            dict[str, Any] - 含 determiner/work_sheet/print_sheet/sources/input_section/fields
        """
        return {
            "determiner": self.determiner,
            "work_sheet": self.work_sheet,
            "print_sheet": self.print_sheet,
            "db_id": self.db_id,
            "use_independent_db": self.use_independent_db,
            "sources": [dict(item) for item in self.sources],
            "input_section": self.input_section.to_dict(),
            "fields": [rule.to_dict() for rule in self.field_rules],
        }



def main() -> None:
    """
    函数名: main
    作用: 命令行自测；演示内存配置序列化，并在样例模板存在时跑默认生成与 verify_toml；含修改 input_area 后再次校验
    输入:
        无（样例模板路径写死在函数内）
    输出:
        无（结果打印到标准输出）
    """
    # ---- 1. 纯内存配置 → TOML 文本（不依赖任何 xlsx）----
    print("=== 1. memory serialization ===")
    demo = GetTomlValues(
        work_sheet="Input_sheet",
        print_sheet="Print_sheet",
        field_rules=[
            TomlDefault(Input_label="ID#", index=0, field="ID", source_file="source1",
                        source_sheet="sheet1", id=True),
            TomlDefault(Input_label="Name", index=1, field="name", source_file="source1",
                        source_sheet="sheet1"),
        ],
        input_section=InputSection(input_area="A2:B2", move_to="down", offset=1),
    )
    print(TomlGenerator().ConfigToToml(demo.ToDict()))
    # ---- 2. 样例模板：默认生成 + 文本回环解析 ----
    repo_root = Path(__file__).resolve().parents[1]
    sample_xlsx = repo_root / "docs" / "sample" / "sample_template.xlsx"
    if not sample_xlsx.is_file():
        print(f"[skip] 样例模板不存在: {sample_xlsx}")
        return
    print("=== 2. generate default config from sample template ===")
    default_cfg = TomlGenerator().CreateDefaultFromTemplate(sample_xlsx)  # 取 active sheet
    default_toml = TomlGenerator().ConfigToToml(default_cfg)
    print(default_toml)
    # 回环：序列化文本应能被重新解析为合法配置
    reparsed = _config_from_dict(tomlkit.loads(default_toml))
    print(f"loop parsing: {'success' if reparsed else 'failed'}")
    # ---- 3. verify_toml report ----
    print("=== 3. verify_toml report ===")
    cfg = reparsed if reparsed else demo
    report = verify_toml(sample_xlsx, cfg)  # 在 work_sheet 上搜标签并校验值格
    print(f"ok: {report['ok']}")
    print(f"missing_labels: {report['missing_labels']}")
    print(f"duplicate_labels: {report['duplicate_labels']}")
    print(f"out_of_area_labels: {report['out_of_area_labels']}")
    print(f"located: {report['located']}")
    print(f"errors: {report['errors']}")
    # ---- 4. 内存中收窄 input_area 为 B2:G2，再次 verify_toml ----
    print("=== 4. verify_toml after input_area -> B2:G2 ===")
    cfg.input_section.input_area = "B2:G2"  # 排除 A 列，ID# 值格 A2 应越界
    report2 = verify_toml(sample_xlsx, cfg)
    print(f"ok: {report2['ok']}")
    print(f"missing_labels: {report2['missing_labels']}")
    print(f"duplicate_labels: {report2['duplicate_labels']}")
    print(f"out_of_area_labels: {report2['out_of_area_labels']}")
    print(f"located: {report2['located']}")
    print(f"errors: {report2['errors']}")
    # ---- 5. id 规则：单 id 时自动解析 db_id ----
    print("=== 5. id rules: single id ===")
    base_rules = [
        TomlDefault(
            Input_label="ID#", index=0, field="ID", source_file="source1",
            source_sheet="sheet1", id=True,
        ),
        TomlDefault(
            Input_label="Name", index=1, field="name", source_file="source1",
            source_sheet="sheet1",
        ),
    ]
    cfg3 = GetTomlValues(
        work_sheet=cfg.work_sheet,
        field_rules=base_rules,
        input_section=InputSection(input_area="A2:G2", move_to="down", offset=1),
        sources=[dict(item) for item in cfg.sources],
    )
    report3 = verify_toml(sample_xlsx, cfg3)
    print(f"ok: {report3['ok']}")
    print(f"db_id: {report3['db_id']}")
    print(f"id_labels: {report3['id_labels']}")
    print(f"id_lookup_keys: {report3['id_lookup_keys']}")
    # ---- 6. 多 id 且无 db_id → db_id_required ----
    print("=== 6. id rules: multi id without db_id ===")
    multi_rules = list(base_rules) + [
        TomlDefault(
            Input_label="AltID", index=-1, field="ID", source_file="source1",
            source_sheet="sheet2", id=True,
        ),
    ]
    multi_cfg = GetTomlValues(
        work_sheet=cfg.work_sheet,
        field_rules=multi_rules,
        input_section=InputSection(input_area="A2:G2", move_to="down", offset=1),
        sources=[dict(item) for item in cfg.sources],
    )
    report4 = verify_toml(sample_xlsx, multi_cfg)
    print(f"ok: {report4['ok']}")
    print(f"db_id_required: {report4['db_id_required']}")
    print(f"duplicate_id_sheets: {report4['duplicate_id_sheets']}")
    # ---- 7. 同一 source_sheet 两个 id=true → duplicate_id_sheets ----
    print("=== 7. id rules: duplicate id on same sheet ===")
    dup_rules = list(base_rules) + [
        TomlDefault(
            Input_label="DupID", index=-1, field="Report_date", source_file="source1",
            source_sheet="sheet1", id=True,
        ),
    ]
    dup_cfg = GetTomlValues(
        work_sheet=cfg.work_sheet,
        field_rules=dup_rules,
        input_section=InputSection(input_area="A2:G2", move_to="down", offset=1),
        sources=[dict(item) for item in cfg.sources],
    )
    report5 = verify_toml(sample_xlsx, dup_cfg)
    print(f"ok: {report5['ok']}")
    print(f"duplicate_id_sheets: {report5['duplicate_id_sheets']}")


if __name__ == "__main__":
    main()
