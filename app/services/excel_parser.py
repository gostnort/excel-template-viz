import io
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def resolve_sheet_name(workbook_path: Path, sheet_name: str) -> str:
    # 大小写不敏感匹配工作表名
    xl = pd.ExcelFile(workbook_path)
    target = sheet_name.strip().lower()
    for name in xl.sheet_names:
        if name.strip().lower() == target:
            return name
    raise ValueError(f"工作表 '{sheet_name}' 不存在，可用: {xl.sheet_names}")


def list_sheet_names(workbook_path: Path) -> list[str]:
    # 获取工作簿内的工作表列表
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    names = list(workbook.sheetnames)
    workbook.close()
    return names


def read_template_sheet(
    workbook_path: Path,
    sheet_name: str,
    header_row: int,
    data_start_row: int,
) -> pd.DataFrame:
    # 读取模板工作表，保留原始列名（含尾随空格）
    resolved = resolve_sheet_name(workbook_path, sheet_name)
    preview = pd.read_excel(workbook_path, sheet_name=resolved, header=None)
    if header_row >= len(preview):
        raise ValueError(f"标题行 {header_row} 超出工作表范围")
    headers = [str(v).strip() if pd.notna(v) else f"col_{i}" for i, v in enumerate(preview.iloc[header_row])]
    if data_start_row >= len(preview):
        return pd.DataFrame(columns=headers)
    body = preview.iloc[data_start_row:].copy()
    body.columns = headers
    body = body.reset_index(drop=True)
    return body


def format_cell_display(value: object) -> str:
    # 将单元格值格式化为表单显示字符串
    if value is None:
        return ""
    if isinstance(value, (pd.Series, pd.Index)):
        if value.empty or value.isna().all():
            return ""
        items = [item for item in value.tolist() if pd.notna(item)]
        if not items:
            return ""
        if len(items) == 1:
            return format_cell_display(items[0])
        formatted = [format_cell_display(item) for item in items]
        return ", ".join(item for item in formatted if item)
    if isinstance(value, (list, tuple, set)):
        if not value:
            return ""
        formatted = [format_cell_display(item) for item in value]
        return ", ".join(item for item in formatted if item)
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value)


def build_dataframe_from_form_rows(headers: list[str], rows: list[dict[str, str]]) -> pd.DataFrame:
    # 由带列标题键的表单行构建 DataFrame
    return pd.DataFrame(rows, columns=headers)


def write_template_sheet(
    workbook_path: Path,
    sheet_name: str,
    dataframe: pd.DataFrame,
    header_row: int,
    data_start_row: int,
) -> bytes:
    # 将编辑后的数据写回工作簿并返回 xlsx 字节
    resolved = resolve_sheet_name(workbook_path, sheet_name)
    workbook = load_workbook(workbook_path)
    worksheet = workbook[resolved]
    for col_idx, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=header_row + 1, column=col_idx, value=column_name)
    for row_offset, row in enumerate(dataframe.itertuples(index=False), start=0):
        excel_row = data_start_row + row_offset + 1
        for col_idx, value in enumerate(row, start=1):
            cell_value = None if pd.isna(value) else value
            worksheet.cell(row=excel_row, column=col_idx, value=cell_value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def parse_spreadsheet_id(url_or_id: str) -> str:
    # 从 URL 或纯 ID 提取 Google Spreadsheet ID
    text = url_or_id.strip()
    if not text:
        raise ValueError("Sheet URL 或 ID 不能为空")
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", text)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", text):
        return text
    raise ValueError("无法解析 Spreadsheet ID，请粘贴完整 URL 或有效 ID")
