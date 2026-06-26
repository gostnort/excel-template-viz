"""外部数据源与 Input_sheet 读写（路径 B）。见 docs/data_flow_design.md。"""
from __future__ import annotations
import argparse
import json
import logging
import re
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from .core_toml import GetTomlValues, TomlDefault, offset_cell


logger = logging.getLogger(__name__)


def _cell_empty(value: Any) -> bool:
    """单元格无有效文本内容时视为空。"""
    return value is None or str(value).strip() == ""


def _column_names_for_rule(rule: TomlDefault) -> list[str]:
    """数据源列查找顺序：先 field 再 Input_label。"""
    names: list[str] = []
    if rule.field:
        names.append(rule.field)
    if rule.Input_label not in names:
        names.append(rule.Input_label)
    return names


def _lookup_row_value(row: dict[str, Any], rule: TomlDefault) -> Any:
    """在已匹配的数据源行中按 rule 取列值。"""
    for key in _column_names_for_rule(rule):
        if key in row:
            return row[key]
    return None


def _load_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """
    函数名: _load_sheet_rows
    作用: 将本地 xlsx 工作表读为行字典列表（首行为表头）
    输入:
        workbook_path (Path) - xlsx 路径
        sheet_name (str) - 工作表名
    输出:
        list[dict[str, Any]] - 每行 {列标题: 值}
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return []
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: list[dict[str, Any]] = []
        # 跳过整行空行，按表头键组装 dict
        for values in row_iter:
            if all(_cell_empty(v) for v in values):
                continue
            item: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header or idx >= len(values):
                    continue
                item[header] = values[idx]
            rows.append(item)
        return rows
    finally:
        wb.close()


def _find_row_by_id(rows: list[dict[str, Any]], id_column: str, id_value: Any) -> dict[str, Any] | None:
    """
    函数名: _find_row_by_id
    作用: 在数据源行列表中按 ID 列匹配一行
    输入:
        rows - 表数据
        id_column (str) - ID 列名（通常为 field=ID）
        id_value (Any) - 待查 ID
    输出:
        dict | None - 匹配行或 None
    """
    target = str(id_value).strip()
    for row in rows:
        if id_column not in row:
            continue
        cell = row[id_column]
        if cell is None:
            continue
        if str(cell).strip() == target:
            return row
        # Excel 数字 ID 可能与字符串形式比较
        try:
            if str(int(float(cell))) == target:
                return row
        except (ValueError, TypeError):
            pass
    return None




class Template2DB:
    """按 source_file / source_sheet / Input_label 读外部表并生成标准记录。"""

    def __init__(self, cfg: GetTomlValues) -> None:
        self.cfg = cfg


    def resolve_source_path(self, source_key: str) -> Path | None:
        """
        函数名: resolve_source_path
        作用: 将 [[sources]] 中的别名解析为磁盘路径
        输入:
            source_key (str) - 如 source1
        输出:
            Path | None - 路径未配置或为空时返回 None
        """
        if not source_key:
            return None
        for item in self.cfg.sources:
            if source_key not in item:
                continue
            raw = item[source_key]
            if raw is None or str(raw).strip() == "":
                return None
            return Path(str(raw))
        return None


    def apply_regex(self, value: Any, pattern: str | None) -> Any:
        """
        函数名: apply_regex
        作用: 对单元格值做正则提取；有捕获组取 group(1)
        输入:
            value (Any) - 原始值
            pattern (str | None) - TOML regex
        输出:
            Any - 提取结果或原值
        """
        if pattern is None or str(pattern).strip() == "":
            return value
        if value is None:
            return value
        text = str(value)
        match = re.search(pattern, text)
        if not match:
            return value
        if match.lastindex:
            return match.group(1)
        return match.group(0)


    def _id_column_name(self) -> str | None:
        """优先取 id=true 且已映射 field 的列名作为查行键。"""
        for rule in self.cfg.field_rules:
            if rule.id and rule.field:
                return rule.field
        for rule in self.cfg.field_rules:
            if rule.id:
                return rule.Input_label
        return None


    def fetch_row_by_id(self, id_value: Any) -> dict[str, Any]:
        """
        函数名: fetch_row_by_id
        作用: 路径 B：按 ID 从各 source_sheet 取列产出 incoming（键为 Input_label）
        输入:
            id_value (Any) - 用户指定或 textbox 中的 ID
        输出:
            dict[str, Any] - incoming：键仅为 Input_label（不含 field 名 / 不含主键 id）
        """
        record: dict[str, Any] = {}
        id_column = self._id_column_name()
        sheet_cache: dict[tuple[str, str], list[dict[str, Any]]] = {}
        # 每个 rule 可能指向不同 source_file / source_sheet
        for rule in self.cfg.field_rules:
            if not rule.source_file or not rule.source_sheet:
                continue
            source_path = self.resolve_source_path(rule.source_file)
            if source_path is None or not source_path.is_file():
                logger.warning("source path missing for %s", rule.source_file)
                continue
            cache_key = (str(source_path), rule.source_sheet)
            if cache_key not in sheet_cache:
                sheet_cache[cache_key] = _load_sheet_rows(source_path, rule.source_sheet)
            rows = sheet_cache[cache_key]
            lookup_col = id_column or "ID"
            matched = _find_row_by_id(rows, lookup_col, id_value)
            if matched is None:
                continue
            raw_value = _lookup_row_value(matched, rule)
            # 只写 Input_label 键；落库主键由 core_store 据 id=true 推导或自动生成
            record[rule.Input_label] = self.apply_regex(raw_value, rule.regex)
        return record




class ExcelWriter:
    """按 verify_toml 的 located + input_section k 组平移读写值格；定位用 Input_label 不用 index。"""

    def __init__(self, cfg: GetTomlValues, located: dict[str, dict[str, int]] | None = None) -> None:
        self.cfg = cfg
        # located: {Input_label: {label_row,label_col,value_row,value_col}}，来自 core_toml.verify_toml
        self.located = dict(located) if located else {}


    def _worksheet_name(self, workbook_path: Path) -> str:
        """解析 cfg.worksheet 或回退 active sheet。"""
        wb = load_workbook(workbook_path, read_only=True)
        try:
            if self.cfg.worksheet and self.cfg.worksheet in wb.sheetnames:
                return self.cfg.worksheet
            return wb.active.title
        finally:
            wb.close()


    def _value_cell(self, label: str, instance_k: int) -> tuple[int, int] | None:
        """
        函数名: _value_cell
        作用: 由 located 的 instance 0 值格，按 input_section 平移得第 k 组值格坐标
        输入:
            label (str) - Input_label
            instance_k (int) - 组序（0 即 instance 0，不平移）
        输出:
            tuple[int, int] | None - (row, col)；label 不在 located 时 None
        """
        coord = self.located.get(label)
        if coord is None:
            return None
        value_row = coord["value_row"]
        value_col = coord["value_col"]
        if instance_k <= 0:
            return value_row, value_col
        section = self.cfg.input_section
        # 同方向平移 offset*k；offset_cell 与 core_toml 共用语义
        return offset_cell(value_row, value_col, section.move_to, section.offset * instance_k)


    def _read_instance(self, ws: Any, instance_k: int) -> dict[str, Any]:
        """读取第 k 组全部 located 值格，键为 Input_label。"""
        values: dict[str, Any] = {}
        for label in self.located:
            cell = self._value_cell(label, instance_k)
            if cell is None:
                continue
            values[label] = ws.cell(row=cell[0], column=cell[1]).value
        return values


    def read_values(self, excel_path: Path, instance_k: int = 0) -> dict[str, Any]:
        """
        函数名: read_values
        作用: 读取单个 instance 的填写值格（键为 Input_label）
        输入:
            excel_path (Path) - 模板 xlsx
            instance_k (int) - 组序
        输出:
            dict[str, Any] - {Input_label: 值}
        """
        sheet_name = self._worksheet_name(excel_path)
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            return self._read_instance(wb[sheet_name], instance_k)
        finally:
            wb.close()


    def read_instances(self, excel_path: Path, limit: int = 100) -> list[dict[str, Any]]:
        """
        函数名: read_instances
        作用: 从 instance 0 起逐组读取值格，遇全空组停止
        输入:
            excel_path (Path) - 模板 xlsx
            limit (int) - 最大扫描组数，防止无界循环
        输出:
            list[dict[str, Any]] - 每组一个 {Input_label: 值}
        """
        if not self.located:
            return []
        sheet_name = self._worksheet_name(excel_path)
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            instances: list[dict[str, Any]] = []
            for k in range(limit):
                values = self._read_instance(ws, k)
                # 全空视为区域结束
                if all(_cell_empty(v) for v in values.values()):
                    break
                instances.append(values)
            return instances
        finally:
            wb.close()


    def write_back(
        self,
        excel_path: Path,
        output_path: Path,
        records: list[dict[str, Any]] | dict[str, Any],
        instance_k: int = 0,
    ) -> None:
        """
        函数名: write_back
        作用: 把记录各 Input_label 值写入对应 located 值格（k 组平移）并另存
        输入:
            excel_path (Path) - 源模板
            output_path (Path) - 输出路径
            records - 单条 dict（写第 instance_k 组）或多条 list（从 instance_k 起依次写）
            instance_k (int) - 起始组序
        输出: 无
        """
        if isinstance(records, dict):
            record_list = [records]
        else:
            record_list = list(records)
        if not self.located:
            raise ValueError("located is empty; run verify_toml first")
        sheet_name = self._worksheet_name(excel_path)
        wb = load_workbook(excel_path)
        try:
            ws = wb[sheet_name]
            # 第 i 条记录写入第 instance_k+i 组值格
            for offset_idx, record in enumerate(record_list):
                k = instance_k + offset_idx
                for label in self.located:
                    if label not in record:
                        continue
                    value = record[label]
                    # 空值不覆盖模板既有内容
                    if _cell_empty(value):
                        continue
                    cell = self._value_cell(label, k)
                    if cell is None:
                        continue
                    ws.cell(row=cell[0], column=cell[1]).value = value
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
        finally:
            wb.close()


    def get_print_areas(self, excel_path: Path) -> dict[str, str | None]:
        """
        函数名: get_print_areas
        作用: 读取各 sheet 的 print_area 元数据（不在 TOML 定位模型内，仅读元数据）
        输入:
            excel_path (Path) - xlsx
        输出:
            dict[str, str | None] - 工作表名 → 区域字符串
        """
        # 非 read_only 打开：ReadOnlyWorksheet 无 print_area 属性
        wb = load_workbook(excel_path)
        try:
            result: dict[str, str | None] = {}
            for name in wb.sheetnames:
                raw = wb[name].print_area
                if raw is None:
                    result[name] = None
                else:
                    text = str(raw)
                    # 去掉工作表前缀与 $ 绝对引用
                    if "!" in text:
                        text = text.split("!", 1)[1]
                    result[name] = text.replace("$", "")
            return result
        finally:
            wb.close()




def main() -> None:
    """
    函数名: _demo_main
    作用: 蓝本 Phase 6：用 docs 样本做全量三段验证（路径 A/B + ExcelWriter）
    输入: 无（可选 argparse 覆盖默认样本路径）
    输出: 无（stdout）
    """
    # 模块级不 import core_store，避免与蓝本依赖方向冲突；仅 __main__ 串联
    import tomlkit
    from .core_registry import PROJECT_ROOT
    from .core_store import SecureSQLite, UiProvider, default_db_path
    from .core_toml import _config_from_dict, verify_toml
    docs_dir = PROJECT_ROOT / "docs" / "sample"
    default_toml = docs_dir / "sample_template.toml"
    default_excel = docs_dir / "sample_template.xlsx"
    default_source_xlsx = docs_dir / "执法堂业绩.xlsx"
    default_template_id = "sample_template"
    default_source_id = "8129"
    default_second_id = "250"
    default_textbox = (
        "8129\tClark Kent\t"
        "recreate himself without changing his dob on records on 1978/02/29"
    )
    parser = argparse.ArgumentParser(description="Data Sheet Core sample verification")
    parser.add_argument("--template-id", default=default_template_id, help="DB basename / template id")
    parser.add_argument("--toml", type=Path, default=default_toml, help="TOML config path")
    parser.add_argument("--excel", type=Path, default=default_excel, help="Input_sheet template xlsx")
    parser.add_argument("--textbox", type=str, default=default_textbox, help="path A: tab-separated string")
    parser.add_argument("--source-id", type=str, default=default_source_id, help="path B: external row ID")
    parser.add_argument("--output", type=Path, default=None, help="write-back xlsx; default exports/{template_id}/sample_template_demo_out.xlsx")
    args = parser.parse_args()
    toml_path = Path(args.toml)
    excel_path = Path(args.excel)
    out_path = args.output or (PROJECT_ROOT / "exports" / args.template_id / "sample_template_demo_out.xlsx")
    if not toml_path.is_file():
        raise SystemExit(f"TOML not found: {toml_path}")
    if not excel_path.is_file():
        raise SystemExit(f"template xlsx not found: {excel_path}")
    try:
        raw = tomlkit.loads(toml_path.read_text(encoding="utf-8"))
    except tomlkit.exceptions.ParseError as exc:
        raise SystemExit(f"TOML parse error in {toml_path}: {exc}") from exc
    if not isinstance(raw, dict):
        raise SystemExit(f"TOML root must be a table: {toml_path}")
    cfg = _config_from_dict(raw)
    if cfg is None:
        raise SystemExit(f"TOML has no valid field rules: {toml_path}")
    # TOML 里可能是本机绝对路径；样本库回退到 docs/执法堂业绩.xlsx
    for item in cfg.sources:
        for key, value in item.items():
            if value is None or str(value).strip() == "":
                continue
            candidate = Path(str(value))
            if not candidate.is_file() and default_source_xlsx.is_file():
                item[key] = str(default_source_xlsx)
    db = SecureSQLite(default_db_path(args.template_id))
    ui = UiProvider(cfg, db)
    t2db = Template2DB(cfg)
    # 唯一校验入口：得 located（标签→值格坐标），供 ExcelWriter 按值格读写
    verify_report = verify_toml(excel_path, cfg)
    located = verify_report.get("located", {})
    writer = ExcelWriter(cfg, located=located)
    read_payload: dict[str, Any] = {
        "toml": str(toml_path),
        "excel": str(excel_path),
        "source_xlsx": str(t2db.resolve_source_path("source1") or ""),
        "verify_toml": verify_report,
    }
    try:
        # 路径 A：textbox 按 determiner 拆分为局部 incoming（仅展示，落库见下方覆盖）
        read_payload["textbox_raw"] = args.textbox
        read_payload["textbox_split"] = ui.split_by_determiner(args.textbox)
        read_payload["textbox_incoming"] = ui.record_from_textbox(args.textbox)
        # Excel 读：按 located + input_section 平移逐组读填写值格
        excel_instances = writer.read_instances(excel_path)
        read_payload["excel_instances"] = excel_instances
        try:
            read_payload["print_areas"] = writer.get_print_areas(excel_path)
        except AttributeError as exc:
            read_payload["print_areas_error"] = str(exc)
        # 路径 B + Excel 组：拼成 incoming 后各写一次（store 按 TOML 覆盖，不 merge）
        source_primary = t2db.fetch_row_by_id(args.source_id)
        read_payload["source_incoming_primary"] = source_primary
        excel_row_primary = excel_instances[0] if excel_instances else {}
        incoming_primary = {**source_primary, **excel_row_primary}
        read_payload["persist_incoming_primary"] = incoming_primary
        rid_primary = ui.persist_fields(incoming_primary)
        write_records: list[dict[str, Any]] = []
        row_primary = db.query_by_id(rid_primary)
        if row_primary is not None:
            write_records.append(row_primary)
        # 第二条记录：演示第 k≥1 组值格平移写回
        source_second = t2db.fetch_row_by_id(default_second_id)
        read_payload["source_incoming_second"] = source_second
        excel_row_second = excel_instances[1] if len(excel_instances) > 1 else {}
        incoming_second = {**source_second, **excel_row_second}
        read_payload["persist_incoming_second"] = incoming_second
        rid_second = ui.persist_fields(incoming_second)
        row_second = db.query_by_id(rid_second)
        if row_second is not None:
            write_records.append(row_second)
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # write_records[i] 写入第 i 组值格（instance_k 从 0 起）
        writer.write_back(excel_path, out_path, write_records, instance_k=0)
        read_payload["output_excel"] = str(out_path)
        print("=== 1. 从 Excel / 数据源读取的数据 ===")
        print(json.dumps(read_payload, ensure_ascii=False, indent=2, default=str))
        print("=== 2. 写入 DB 后的数据 ===")
        print(json.dumps(db.query_all(), ensure_ascii=False, indent=2, default=str))
        print("=== 3. Gradio 可获得的数据 ===")
        print("labels:", json.dumps(ui.get_labels(), ensure_ascii=False))
        print("data:", json.dumps(ui.get_data(), ensure_ascii=False, indent=2, default=str))
    finally:
        db.close()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
