from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import tomlkit
import tomlkit.exceptions
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from tomlkit import aot, document, string, table

from app.services.core_registry import TEMPLATES_DIR


CORE_CONFIG_SUFFIX = ".toml"
DEFAULT_DETERMINER = "\t"
DEFAULT_SOURCES: list[dict[str, str | None]] = [{"source1": None}]
DEFAULT_VALUE_FROM_LABEL = "down"
DEFAULT_VALUE_OFFSET = 1
DEFAULT_MOVE_TO = "down"
DEFAULT_INPUT_OFFSET = 1
FIELD_LABEL_KEY = "Input_label"
OPTIONAL_FIELD_KEYS = ("field", "source_file", "source_sheet", "regex")
VALID_DIRECTIONS = {"up", "down", "left", "right"}
VERIFY_SCAN_ROWS = 100
VERIFY_SCAN_COLS = 100


@dataclass
class InputSection:
    """One [[input_section]] row."""

    input_area: str
    move_to: str = DEFAULT_MOVE_TO
    offset: int = DEFAULT_INPUT_OFFSET


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: InputSection.to_dict
        作用: 将单条 [[input_section]] 转换为可序列化字典
        输入:
            无
        输出:
            dict[str, Any] - 含 input_area / move_to / offset 的字典
        """
        return {
            "input_area": self.input_area,  # instance 0 填写值区域
            "move_to": self.move_to,        # 第 k≥1 组值格平移方向
            "offset": self.offset,          # 平移步长
        }



@dataclass
class TomlDefault:
    """One [[fields]] row. Index is ZERO base."""

    Input_label: str
    value_from_label: str = DEFAULT_VALUE_FROM_LABEL
    value_offset: int = DEFAULT_VALUE_OFFSET
    field: str | None = None
    source_file: str | None = None
    source_sheet: str | None = None
    index: int = -1
    regex: str | None = None
    id: bool = False


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: TomlDefault.to_dict
        作用: 将一条 [[fields]] 字段规则转换为可序列化字典；未映射可选键取 None
        输入:
            无
        输出:
            dict[str, Any] - 字段规则的字典形式（内存语义，未映射为 None）
        """
        return {
            FIELD_LABEL_KEY: self.Input_label,
            "value_from_label": self.value_from_label,  # 标签到值格的方向
            "value_offset": self.value_offset,          # 标签到值格的步长
            "field": self.field if not _is_unmapped(self.field) else None,
            "source_file": self.source_file if not _is_unmapped(self.source_file) else None,
            "source_sheet": self.source_sheet if not _is_unmapped(self.source_sheet) else None,
            "index": self.index,
            "regex": self.regex if not _is_unmapped(self.regex) else None,
            "id": self.id,
        }



@dataclass
class VerifyTomlReport:
    """Structured report returned by verify_toml()."""

    ok: bool
    missing_labels: list[str]
    out_of_area_labels: list[str]
    located: dict[str, dict[str, int]]
    errors: list[str]


    def to_dict(self) -> dict[str, Any]:
        """
        函数名: VerifyTomlReport.to_dict
        作用: 将校验报告转换为给 UI 的字典契约（拷贝内部列表/字典避免外部改动）
        输入:
            无
        输出:
            dict[str, Any] - 含 ok / missing_labels / out_of_area_labels / located / errors
        """
        return {
            "ok": self.ok,
            "missing_labels": list(self.missing_labels),        # 找不到的标签
            "out_of_area_labels": list(self.out_of_area_labels),  # 值格越界的标签
            "located": dict(self.located),                       # 通过项坐标
            "errors": list(self.errors),                         # 结构性错误
        }


def _core_toml_path(template_id: str) -> Path:
    """
    函数名: _core_toml_path
    作用: 根据模板 ID 拼出其 TOML 配置文件的磁盘路径
    输入:
        template_id (str) - 模板唯一标识
    输出:
        Path - templates/{id}/{id}.toml 的 Path
    """
    return TEMPLATES_DIR / template_id / f"{template_id}{CORE_CONFIG_SUFFIX}"


def _is_unmapped(value: Any) -> bool:
    """
    函数名: _is_unmapped
    作用: 判断一个可选键的值是否视为未映射（None 或空串）
    输入:
        value (Any) - 待判定的值
    输出:
        bool - 未映射返回 True
    """
    return value is None or value == ""


def _needs_literal_string(key: str, value: str) -> bool:
    """
    函数名: _needs_literal_string
    作用: 判断该键值是否需要写成 TOML 字面量字符串（单引号，避免反斜杠转义）
    输入:
        key (str) - 字段键名，regex 始终用字面量
        value (str) - 字段字符串值，含反斜杠时用字面量
    输出:
        bool - 需要字面量返回 True
    """
    # 值本身含单引号时无法用字面量，退回普通字符串
    if "'" in value:
        return False
    if key == "regex":
        return True
    if "\\" in value:
        return True
    return False


def _toml_string(key: str, value: Any) -> Any:
    """
    函数名: _toml_string
    作用: 把内存值转换为可写盘的 TOML 字符串；未映射写空串，必要时用字面量
    输入:
        key (str) - 字段键名，决定是否倾向字面量
        value (Any) - 内存值，None/"" 视为未映射
    输出:
        Any - 普通字符串或 tomlkit 字面量字符串
    """
    if _is_unmapped(value):
        return ""
    text = str(value)  # 统一成文本再判断写法
    if _needs_literal_string(key, text):
        return string(text, literal=True)
    return text


def _extract_toml_text(model_output: str) -> str:
    """
    函数名: _extract_toml_text
    作用: 从可能带 Markdown 代码块围栏的文本中提取纯 TOML
    输入:
        model_output (str) - 用户或 LLM 提供的原始文本
    输出:
        str - 去除围栏后的 TOML 文本
    """
    text = model_output.strip()
    if not text:
        return text
    fenced = re.search(
        r"```(?:toml)?\s*\r?\n?(.*?)```", text, flags=re.DOTALL | re.IGNORECASE
    )  # 优先匹配成对围栏
    if fenced:
        return fenced.group(1).strip()
    # 仅有起始围栏时逐段剥除
    if text.startswith("```"):
        text = re.sub(r"^```(?:toml)?\s*\r?\n?", "", text, count=1, flags=re.IGNORECASE)
        text = re.sub(r"```\s*$", "", text.strip())
    return text.strip()


def _parse_bool(value: Any) -> bool:
    """
    函数名: _parse_bool
    作用: 把布尔/字符串/数字宽松解析为 bool
    输入:
        value (Any) - 待解析值
    输出:
        bool - 解析结果
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("true", "1", "yes")
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _parse_int(value: Any, default: int) -> int:
    """
    函数名: _parse_int
    作用: 把值解析为 int，失败时回退默认值
    输入:
        value (Any) - 待解析值
        default (int) - 解析失败时的回退值
    输出:
        int - 解析结果或默认值
    """
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _optional_string(value: Any) -> str | None:
    """
    函数名: _optional_string
    作用: 规范可选字符串键——未映射转为 None，其余去空白
    输入:
        value (Any) - 原始值
    输出:
        str | None - 未映射为 None，否则为去空白文本
    """
    if _is_unmapped(value):
        return None
    return str(value).strip()


def _parse_sources(raw_sources: Any) -> list[dict[str, str | None]]:
    """
    函数名: _parse_sources
    作用: 解析 [[sources]]，把空串别名规范成 None；无有效项时回退默认
    输入:
        raw_sources (Any) - 原始 sources 数据
    输出:
        list[dict[str, str | None]] - 规范化后的数据源别名列表
    """
    if not isinstance(raw_sources, list) or not raw_sources:
        return [dict(item) for item in DEFAULT_SOURCES]
    sources: list[dict[str, str | None]] = []
    for item in raw_sources:
        if not isinstance(item, dict):
            continue
        parsed_item: dict[str, str | None] = {}
        for key, value in item.items():
            parsed_item[str(key)] = _optional_string(value)  # 空串别名转 None
        if parsed_item:
            sources.append(parsed_item)
    if not sources:
        return [dict(item) for item in DEFAULT_SOURCES]
    return sources


def _input_section_from_dict(raw: Any) -> InputSection | None:
    """
    函数名: _input_section_from_dict
    作用: 解析有且仅有一条的 [[input_section]]；缺 input_area 视为无效
    输入:
        raw (Any) - input_section 原始数据（dict 或仅含一项的 list）
    输出:
        InputSection | None - 解析成功的区段，或 None
    """
    # tomlkit 的 [[input_section]] 会读成 list，按设计只允许一条
    if isinstance(raw, list):
        if len(raw) != 1:
            return None
        raw = raw[0]
    if not isinstance(raw, dict):
        return None
    input_area = str(raw.get("input_area", "")).strip()
    if not input_area:
        return None
    move_to = str(raw.get("move_to", DEFAULT_MOVE_TO)).strip().lower()  # 方向归一化
    offset = _parse_int(raw.get("offset", DEFAULT_INPUT_OFFSET), DEFAULT_INPUT_OFFSET)
    return InputSection(input_area=input_area, move_to=move_to, offset=offset)


def _field_from_dict(raw: Any) -> TomlDefault | None:
    """
    函数名: _field_from_dict
    作用: 解析一条 [[fields]]；缺少 Input_label 视为无效
    输入:
        raw (Any) - 单条字段映射原始数据
    输出:
        TomlDefault | None - 解析成功的字段规则，或 None
    """
    if not isinstance(raw, dict):
        return None
    input_label = str(raw.get(FIELD_LABEL_KEY, "")).strip()
    if not input_label:
        return None
    value_from_label = str(
        raw.get("value_from_label", DEFAULT_VALUE_FROM_LABEL)
    ).strip().lower()  # 方向归一化为小写
    value_offset = _parse_int(raw.get("value_offset", DEFAULT_VALUE_OFFSET), DEFAULT_VALUE_OFFSET)
    index_val = _parse_int(raw.get("index", -1), -1)  # -1 表示不参与文本拆分
    return TomlDefault(
        Input_label=input_label,
        value_from_label=value_from_label,
        value_offset=value_offset,
        field=_optional_string(raw.get("field")),
        source_file=_optional_string(raw.get("source_file")),
        source_sheet=_optional_string(raw.get("source_sheet")),
        index=index_val,
        regex=_optional_string(raw.get("regex")),
        id=_parse_bool(raw.get("id", False)),
    )


def _config_from_dict(raw: dict[str, Any]) -> GetTomlValues | None:
    """
    函数名: _config_from_dict
    作用: 由解析后的 TOML 字典构造 GetTomlValues；缺 worksheet/input_section/fields 则失败
    输入:
        raw (dict[str, Any]) - tomlkit 解析得到的字典
    输出:
        GetTomlValues | None - 构造成功的配置对象，或 None
    """
    field_rules: list[TomlDefault] = []
    fields_raw = raw.get("fields")
    if isinstance(fields_raw, list):
        for item in fields_raw:
            rule = _field_from_dict(item)
            if rule:
                field_rules.append(rule)
    input_section = _input_section_from_dict(raw.get("input_section"))
    worksheet = _optional_string(raw.get("worksheet"))
    # 三项必备：worksheet、单条 input_section、至少一条 fields
    if not worksheet or input_section is None or not field_rules:
        return None
    determiner = str(raw.get("determiner", DEFAULT_DETERMINER)) or DEFAULT_DETERMINER
    return GetTomlValues(
        determiner=determiner,
        sources=_parse_sources(raw.get("sources")),
        field_rules=field_rules,
        worksheet=worksheet,
        input_section=input_section,
    )


def _dict_to_toml(config: dict[str, Any]) -> str:
    """
    函数名: _dict_to_toml
    作用: 把配置字典序列化为严格 TOML 1.0 文本；未映射可选键写空串 ""，不写 null
    输入:
        config (dict[str, Any]) - 含 determiner/worksheet/sources/input_section/fields
    输出:
        str - 序列化后的 TOML 文本（以换行结尾）
    """
    # ---- 顶层标量：determiner 与 worksheet ----
    doc = document()
    doc["determiner"] = str(config.get("determiner", DEFAULT_DETERMINER))
    doc["worksheet"] = str(config.get("worksheet", ""))
    # ---- [[sources]]：空别名落盘为空串 ----
    sources_aot = aot()
    for item in config.get("sources", DEFAULT_SOURCES):
        if not isinstance(item, dict):
            continue
        src = table()
        for source_key, source_value in item.items():
            src[str(source_key)] = _toml_string(str(source_key), source_value)
        sources_aot.append(src)
    doc["sources"] = sources_aot
    # ---- [[input_section]]：有且仅有一条 ----
    section = _input_section_from_dict(config.get("input_section"))
    input_sections = aot()
    if section:
        section_row = table()
        section_row["input_area"] = section.input_area
        section_row["move_to"] = section.move_to
        section_row["offset"] = section.offset
        input_sections.append(section_row)
    doc["input_section"] = input_sections
    # ---- [[fields]]：逐条写必有键与可选键 ----
    fields_aot = aot()
    for item in config.get("fields", []):
        rule = _field_from_dict(item) if isinstance(item, dict) else item  # 兼容 dict 与 TomlDefault
        if not isinstance(rule, TomlDefault):
            continue
        row = table()
        row[FIELD_LABEL_KEY] = rule.Input_label
        row["value_from_label"] = rule.value_from_label
        row["value_offset"] = rule.value_offset
        rule_dict = rule.to_dict()
        for key in OPTIONAL_FIELD_KEYS:
            row[key] = _toml_string(key, rule_dict.get(key))  # 未映射写 ""
        row["index"] = rule.index
        row["id"] = rule.id
        fields_aot.append(row)
    doc["fields"] = fields_aot
    dumped = tomlkit.dumps(doc)
    if not dumped.endswith("\n"):
        dumped += "\n"
    return dumped


def _parse_area(area: str) -> tuple[int, int, int, int]:
    """
    函数名: _parse_area
    作用: 把 Excel 区域字符串解析为 (min_row, min_col, max_row, max_col)
    输入:
        area (str) - 形如 "A2:G2" 的区域
    输出:
        tuple[int, int, int, int] - 1-based 行列边界
    """
    # range_boundaries 返回 (min_col, min_row, max_col, max_row)，此处换序为行列
    min_col, min_row, max_col, max_row = range_boundaries(area)
    return min_row, min_col, max_row, max_col


def _cell_text(value: Any) -> str:
    """
    函数名: _cell_text
    作用: 把单元格值规范化为用于标签比对的去空白文本
    输入:
        value (Any) - 单元格原始值
    输出:
        str - None 转空串，其余去空白
    """
    if value is None:
        return ""
    return str(value).strip()


def offset_cell(row: int, col: int, direction: str, offset: int) -> tuple[int, int]:
    """
    函数名: offset_cell
    作用: 在 1-based 坐标上按方向与步长平移，越界或非法方向时抛错
    输入:
        row (int) - 起始行（1-based）
        col (int) - 起始列（1-based）
        direction (str) - up/down/left/right
        offset (int) - 正整数步长
    输出:
        tuple[int, int] - 平移后的 (row, col)
    """
    direction = direction.strip().lower()  # 容忍大小写与空白
    if direction not in VALID_DIRECTIONS:
        raise ValueError(f"unsupported direction: {direction}")
    if offset < 1:
        raise ValueError("offset must be a positive integer")
    if direction == "up":
        row -= offset
    elif direction == "down":
        row += offset
    elif direction == "left":
        col -= offset
    elif direction == "right":
        col += offset
    if row < 1 or col < 1:
        raise ValueError("offset produced an out-of-bounds coordinate")
    return row, col


def _find_label_cell(ws: Any, input_label: str) -> tuple[int, int] | None:
    """
    函数名: _find_label_cell
    作用: 在工作表 (1,1) 起、100×100 内从左到右从上到下找首个完全匹配的标签格
    输入:
        ws (Any) - openpyxl 工作表对象
        input_label (str) - 要匹配的标签文本
    输出:
        tuple[int, int] | None - 命中的 (row, col)，未命中为 None
    """
    max_row = min(ws.max_row or VERIFY_SCAN_ROWS, VERIFY_SCAN_ROWS)  # 上限封顶 100 行
    max_col = min(ws.max_column or VERIFY_SCAN_COLS, VERIFY_SCAN_COLS)  # 上限封顶 100 列
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _cell_text(ws.cell(row=row, column=col).value) == input_label:
                return row, col
    return None


def _cell_in_area(row: int, col: int, area: tuple[int, int, int, int]) -> bool:
    """
    函数名: _cell_in_area
    作用: 判断坐标是否落在区域矩形内
    输入:
        row (int) - 行（1-based）
        col (int) - 列（1-based）
        area (tuple[int, int, int, int]) - (min_row, min_col, max_row, max_col)
    输出:
        bool - 在区域内返回 True
    """
    min_row, min_col, max_row, max_col = area
    return min_row <= row <= max_row and min_col <= col <= max_col


def _headers_from_first_row(ws: Any) -> list[tuple[int, str]]:
    """
    函数名: _headers_from_first_row
    作用: 读取第 1 行非空单元格，作为标准范式默认配置的字段标签
    输入:
        ws (Any) - openpyxl 工作表对象
    输出:
        list[tuple[int, str]] - (列号, 标签文本) 列表
    """
    headers: list[tuple[int, str]] = []
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        text = _cell_text(ws.cell(row=1, column=col).value)  # 第 1 行该列文本
        if text:
            headers.append((col, text))
    return headers


def load_toml(template_id: str) -> GetTomlValues | None:
    """
    函数名: load_toml
    作用: 按模板 ID 读取并解析 TOML（只解析文本，不打开 xlsx 校验）
    输入:
        template_id (str) - 模板唯一标识
    输出:
        GetTomlValues | None - 解析成功的配置，或 None（不存在/解析失败/结构不全）
    """
    path = _core_toml_path(template_id)
    if not path.exists():
        return None
    try:
        raw = tomlkit.loads(path.read_text(encoding="utf-8"))
    except tomlkit.exceptions.ParseError:
        return None
    if not isinstance(raw, dict):
        return None
    return _config_from_dict(raw)


def verify_toml(template_path: Path, cfg: GetTomlValues) -> dict[str, Any]:
    """
    函数名: verify_toml
    作用: UI 唯一校验入口；在 worksheet 上搜每个 Input_label 并验证 instance 0 值格是否落在 input_area
    输入:
        template_path (Path) - 当前模板 xlsx 路径
        cfg (GetTomlValues) - 已加载的 TOML 配置
    输出:
        dict[str, Any] - 报告：ok / missing_labels / out_of_area_labels / located / errors
    """
    missing_labels: list[str] = []
    out_of_area_labels: list[str] = []
    located: dict[str, dict[str, int]] = {}
    errors: list[str] = []
    if not cfg.worksheet:
        return VerifyTomlReport(False, [], [], {}, ["worksheet is required"]).to_dict()
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        # worksheet 不存在直接整体失败
        if cfg.worksheet not in wb.sheetnames:
            return VerifyTomlReport(False, [], [], {}, [f"worksheet not found: {cfg.worksheet}"]).to_dict()
        ws = wb[cfg.worksheet]
        try:
            input_area = _parse_area(cfg.input_section.input_area)  # 解析约束矩形
        except ValueError as exc:
            return VerifyTomlReport(False, [], [], {}, [f"invalid input_area: {exc}"]).to_dict()
        for rule in cfg.field_rules:
            label_cell = _find_label_cell(ws, rule.Input_label)  # 每字段从头扫描
            if label_cell is None:
                missing_labels.append(rule.Input_label)  # 标签找不到
                continue
            label_row, label_col = label_cell
            try:
                value_row, value_col = offset_cell(
                    label_row, label_col, rule.value_from_label, rule.value_offset
                )  # 由标签推算 instance 0 值格
            except ValueError as exc:
                errors.append(f"{rule.Input_label}: {exc}")  # 方向/步长非法或越界
                continue
            if not _cell_in_area(value_row, value_col, input_area):
                out_of_area_labels.append(rule.Input_label)  # 值格不在 input_area 内
                continue
            located[rule.Input_label] = {
                "label_row": label_row,
                "label_col": label_col,
                "value_row": value_row,
                "value_col": value_col,
            }
    finally:
        wb.close()
    ok = not missing_labels and not out_of_area_labels and not errors  # 全部无问题才通过
    return VerifyTomlReport(ok, missing_labels, out_of_area_labels, located, errors).to_dict()


def ensure_exists(
    template_id: str,
    template_path: Path,
    worksheet_name: str | None = None,
) -> bool:
    """
    函数名: ensure_exists
    作用: TOML 不存在时按标准范式生成默认配置
    输入:
        template_id (str) - 模板唯一标识
        template_path (Path) - 模板 xlsx 路径
        worksheet_name (str | None) - 目标工作表名，None 时取 active sheet
    输出:
        bool - 已存在或生成成功返回 True
    """
    path = _core_toml_path(template_id)
    if path.exists():
        return True
    return TomlGenerator().Reset(template_id, template_path, worksheet_name)



class TomlGenerator:
    """First-time creation of default TOML config and serialization."""

    def _worksheet(self, template_path: Path, worksheet_name: str | None) -> tuple[str, Any, Any]:
        """
        函数名: TomlGenerator._worksheet
        作用: 打开工作簿并定位目标工作表；未指定名时取 active sheet
        输入:
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            tuple[str, Any, Any] - (解析出的表名, 工作表对象, 工作簿对象)
        """
        wb = load_workbook(template_path, read_only=True, data_only=True)
        if worksheet_name:
            # 指定名但不存在时关闭工作簿再抛错，避免句柄泄漏
            if worksheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"worksheet not found: {worksheet_name}")
            return worksheet_name, wb[worksheet_name], wb
        ws = wb.active
        return ws.title, ws, wb


    def CreateDefaultFromTemplate(
        self,
        template_path: Path,
        worksheet_name: str | None = None,
    ) -> dict[str, Any]:
        """
        函数名: TomlGenerator.CreateDefaultFromTemplate
        作用: 按标准范式生成默认配置——第 1 行作字段标签，input_area 登记第 2 行值区；不逐格扫描
        输入:
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            dict[str, Any] - 默认配置字典
        """
        resolved_name, ws, wb = self._worksheet(template_path, worksheet_name)
        try:
            headers = _headers_from_first_row(ws)  # 第 1 行非空标签
            fields = [
                TomlDefault(Input_label=label, index=idx).to_dict()  # index 取列序 base 0
                for idx, (_col, label) in enumerate(headers)
            ]
            # input_area 取标签覆盖的列范围、固定第 2 行
            if headers:
                start_col = headers[0][0]
                end_col = headers[-1][0]
            else:
                start_col = 1
                end_col = 1
            input_area = f"{get_column_letter(start_col)}2:{get_column_letter(end_col)}2"
            return {
                "determiner": DEFAULT_DETERMINER,
                "worksheet": resolved_name,
                "sources": [dict(item) for item in DEFAULT_SOURCES],
                "input_section": InputSection(input_area=input_area).to_dict(),
                "fields": fields,
            }
        finally:
            wb.close()


    def ConfigToToml(self, config: dict[str, Any]) -> str:
        """
        函数名: TomlGenerator.ConfigToToml
        作用: 把配置字典序列化为严格 TOML 1.0 文本
        输入:
            config (dict[str, Any]) - 配置字典
        输出:
            str - TOML 文本
        """
        return _dict_to_toml(config)


    def Reset(
        self,
        template_id: str,
        template_path: Path,
        worksheet_name: str | None = None,
    ) -> bool:
        """
        函数名: TomlGenerator.Reset
        作用: 用标准范式默认配置覆盖写入模板 TOML 文件
        输入:
            template_id (str) - 模板唯一标识
            template_path (Path) - 模板 xlsx 路径
            worksheet_name (str | None) - 目标工作表名，None 时取 active
        输出:
            bool - 写入成功返回 True，异常返回 False
        """
        path = _core_toml_path(template_id)
        try:
            default_cfg = self.CreateDefaultFromTemplate(template_path, worksheet_name)
            path.parent.mkdir(parents=True, exist_ok=True)  # 确保目录存在
            path.write_text(self.ConfigToToml(default_cfg), encoding="utf-8")
            return True
        except Exception:
            return False



class GetTomlValues:
    """Loaded TOML config: query, modify in memory, persist via Save / ToDict."""

    def __init__(
        self,
        determiner: str = DEFAULT_DETERMINER,
        sources: list[dict[str, str | None]] | None = None,
        field_rules: list[TomlDefault] | None = None,
        worksheet: str | None = None,
        input_section: InputSection | None = None,
    ) -> None:
        """
        函数名: GetTomlValues.__init__
        作用: 构造已加载配置对象；可选参数缺省时给出安全默认
        输入:
            determiner (str) - 文本拆分分隔符
            sources (list[dict[str, str | None]] | None) - 数据源别名列表
            field_rules (list[TomlDefault] | None) - 字段规则列表
            worksheet (str | None) - 目标工作表名
            input_section (InputSection | None) - 单条区段
        输出:
            无
        """
        self.determiner = determiner
        self.sources = sources if sources is not None else [dict(item) for item in DEFAULT_SOURCES]
        self.field_rules = field_rules if field_rules is not None else []
        self.worksheet = worksheet
        self.input_section = input_section if input_section is not None else InputSection("A2:A2")


    def Load(self, template_id: str) -> GetTomlValues | None:
        """
        函数名: GetTomlValues.Load
        作用: 按模板 ID 读取 TOML（仅解析文本，不打开 xlsx 校验）
        输入:
            template_id (str) - 模板唯一标识
        输出:
            GetTomlValues | None - 解析成功的配置，或 None
        """
        return load_toml(template_id)


    def Save(self, template_id: str, toml_text: str | None = None) -> None:
        """
        函数名: GetTomlValues.Save
        作用: 落盘配置；toml_text 为 None 时保存当前实例，否则校验文本结构后再写
        输入:
            template_id (str) - 模板唯一标识
            toml_text (str | None) - 用户/LLM 提供的 TOML 文本，None 表示用当前实例
        输出:
            无
        """
        if toml_text is None:
            cleaned = TomlGenerator().ConfigToToml(self.ToDict())  # 直接序列化当前实例
        else:
            cleaned = _extract_toml_text(toml_text)  # 去掉可能的 Markdown 围栏
            try:
                parsed = tomlkit.loads(cleaned)
            except tomlkit.exceptions.ParseError as exc:
                raise ValueError(f"Invalid TOML: {exc}") from exc
            # 结构须符合当前设计（有 worksheet/input_section/fields）
            if not isinstance(parsed, dict) or _config_from_dict(parsed) is None:
                raise ValueError("TOML must match the current core_toml design")
            cleaned = TomlGenerator().ConfigToToml(parsed)  # 归一化后再写
        path = _core_toml_path(template_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(cleaned, encoding="utf-8")


    def VerifyToml(self, template_path: Path) -> dict[str, Any]:
        """
        函数名: GetTomlValues.VerifyToml
        作用: 以本实例为配置，对模板 xlsx 调用 verify_toml
        输入:
            template_path (Path) - 模板 xlsx 路径
        输出:
            dict[str, Any] - 校验报告
        """
        return verify_toml(template_path, self)


    def ToDict(self) -> dict[str, Any]:
        """
        函数名: GetTomlValues.ToDict
        作用: 把当前配置转换为可序列化字典
        输入:
            无
        输出:
            dict[str, Any] - 含 determiner/worksheet/sources/input_section/fields
        """
        return {
            "determiner": self.determiner,
            "worksheet": self.worksheet,
            "sources": [dict(item) for item in self.sources],
            "input_section": self.input_section.to_dict(),
            "fields": [rule.to_dict() for rule in self.field_rules],
        }



def main() -> None:
    """
    函数名: main
    作用: 命令行自测；演示内存配置序列化，并在样例模板存在时跑默认生成与 verify_toml；含修改 input_area 后再次校验
    输入:
        无（样例模板路径写死在函数内）
    输出:
        无（结果打印到标准输出）
    """
    # ---- 1. 纯内存配置 → TOML 文本（不依赖任何 xlsx）----
    print("=== 1. memory serialization ===")
    demo = GetTomlValues(
        worksheet="Input_sheet",
        field_rules=[
            TomlDefault(Input_label="ID#", index=0, field="ID", source_file="source1",
                        source_sheet="sheet1", id=True),
            TomlDefault(Input_label="Name", index=1, field="name", source_file="source1",
                        source_sheet="sheet1"),
        ],
        input_section=InputSection(input_area="A2:B2", move_to="down", offset=1),
    )
    print(TomlGenerator().ConfigToToml(demo.ToDict()))
    # ---- 2. 样例模板：默认生成 + 文本回环解析 ----
    repo_root = Path(__file__).resolve().parents[2]  # app/services/core_toml.py 上溯到仓库根
    sample_xlsx = repo_root / "docs" / "sample" / "sample_template.xlsx"
    if not sample_xlsx.is_file():
        print(f"[skip] 样例模板不存在: {sample_xlsx}")
        return
    print("=== 2. generate default config from sample template ===")
    default_cfg = TomlGenerator().CreateDefaultFromTemplate(sample_xlsx)  # 取 active sheet
    default_toml = TomlGenerator().ConfigToToml(default_cfg)
    print(default_toml)
    # 回环：序列化文本应能被重新解析为合法配置
    reparsed = _config_from_dict(tomlkit.loads(default_toml))
    print(f"loop parsing: {'success' if reparsed else 'failed'}")
    # ---- 3. verify_toml report ----
    print("=== 3. verify_toml report ===")
    cfg = reparsed if reparsed else demo
    report = verify_toml(sample_xlsx, cfg)  # 在 worksheet 上搜标签并校验值格
    print(f"ok: {report['ok']}")
    print(f"missing_labels: {report['missing_labels']}")
    print(f"out_of_area_labels: {report['out_of_area_labels']}")
    print(f"located: {report['located']}")
    print(f"errors: {report['errors']}")
    # ---- 4. 内存中收窄 input_area 为 B2:G2，再次 verify_toml ----
    print("=== 4. verify_toml after input_area -> B2:G2 ===")
    cfg.input_section.input_area = "B2:G2"  # 排除 A 列，ID# 值格 A2 应越界
    report2 = verify_toml(sample_xlsx, cfg)
    print(f"ok: {report2['ok']}")
    print(f"missing_labels: {report2['missing_labels']}")
    print(f"out_of_area_labels: {report2['out_of_area_labels']}")
    print(f"located: {report2['located']}")
    print(f"errors: {report2['errors']}")


if __name__ == "__main__":
    main()
