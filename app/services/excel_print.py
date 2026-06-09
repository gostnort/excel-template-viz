from __future__ import annotations

import io
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from app.services.registry import PROJECT_ROOT

EXPORTS_DIR = PROJECT_ROOT / "exports"
PRINT_RENDER_DPI = 300
_SCREEN_DPI = 96.0
_PRINT_SCRIPT = Path(__file__).with_name("print_image_dialog.ps1")


@dataclass(frozen=True)
class PrintAreaInfo:
    sheet: str
    range: str

    @property
    def label(self) -> str:
        return f"{self.sheet} · {self.range}"


def _normalize_range(raw: str) -> str:
    text = raw.strip()
    if "!" in text:
        _, text = text.split("!", 1)
    return text.replace("$", "")


def _split_print_area_value(raw: str) -> list[str]:
    return [_normalize_range(part) for part in str(raw).split(",") if part.strip()]


def _is_label_sheet(sheet_name: str) -> bool:
    return sheet_name.strip().lower() == "label"


def _label_sheet_dimension_range(worksheet) -> str | None:
    dimension = _normalize_range(worksheet.calculate_dimension())
    if not dimension or dimension == "A1":
        return None
    return dimension


def list_print_areas(workbook_source: Path | bytes) -> list[PrintAreaInfo]:
    if isinstance(workbook_source, bytes):
        workbook = load_workbook(io.BytesIO(workbook_source), read_only=False, data_only=True)
    else:
        workbook = load_workbook(workbook_source, read_only=False, data_only=True)
    areas: list[PrintAreaInfo] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            raw_area = worksheet.print_area
            if raw_area:
                for area_range in _split_print_area_value(str(raw_area)):
                    areas.append(PrintAreaInfo(sheet=sheet_name, range=area_range))
                continue
            if _is_label_sheet(sheet_name):
                area_range = _label_sheet_dimension_range(worksheet)
                if area_range:
                    areas.append(PrintAreaInfo(sheet=sheet_name, range=area_range))
    finally:
        workbook.close()
    return areas


def primary_print_area(workbook_source: Path | bytes) -> PrintAreaInfo | None:
    areas = list_print_areas(workbook_source)
    for area in areas:
        if _is_label_sheet(area.sheet):
            return area
    return areas[0] if areas else None


def persist_export_file(template_id: str, xlsx_bytes: bytes, filename: str) -> Path:
    dest_dir = EXPORTS_DIR / template_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / filename
    dest_path.write_bytes(xlsx_bytes)
    return dest_path


def read_print_area_values(
    workbook_source: Path | bytes,
    sheet_name: str,
    print_area: str,
) -> list[list[str]]:
    if isinstance(workbook_source, bytes):
        workbook = load_workbook(io.BytesIO(workbook_source), read_only=True, data_only=True)
    else:
        workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        worksheet = workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(print_area)
        rows: list[list[str]] = []
        for row_cells in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            rows.append(["" if cell.value is None else str(cell.value) for cell in row_cells])
        return rows
    finally:
        workbook.close()


def _scale_for_print(value: float) -> int:
    return max(1, int(round(value * PRINT_RENDER_DPI / _SCREEN_DPI)))


def _load_print_font(size: int):
    from PIL import ImageFont

    for name in ("msyh.ttc", "simhei.ttf", "arial.ttf"):
        path = Path("C:/Windows/Fonts") / name
        if path.is_file():
            try:
                return ImageFont.truetype(str(path), size)
            except OSError:
                continue
    return ImageFont.load_default()


def _rows_to_pil_image(rows: list[list[str]]):
    from PIL import Image, ImageDraw

    if not rows:
        raise ValueError("打印区域为空")

    font = _load_print_font(_scale_for_print(14))
    padding_x = _scale_for_print(8)
    padding_y = _scale_for_print(6)
    min_col_width = _scale_for_print(40)
    line_width = max(1, _scale_for_print(1))

    draw_probe = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    col_count = max(len(row) for row in rows)
    col_widths = [0] * col_count
    for row in rows:
        for idx, cell in enumerate(row):
            if idx >= col_count:
                continue
            bbox = draw_probe.textbbox((0, 0), cell, font=font)
            col_widths[idx] = max(col_widths[idx], bbox[2] - bbox[0] + 2 * padding_x)
    col_widths = [max(width, min_col_width) for width in col_widths]
    line_height = draw_probe.textbbox((0, 0), "Ag", font=font)[3] + 2 * padding_y
    image_width = sum(col_widths) + line_width
    image_height = line_height * len(rows) + line_width
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    y = 0
    for row in rows:
        x = 0
        for idx in range(col_count):
            cell = row[idx] if idx < len(row) else ""
            draw.rectangle(
                (x, y, x + col_widths[idx], y + line_height),
                outline="#cccccc",
                width=line_width,
            )
            draw.text((x + padding_x, y + padding_y), cell, fill="black", font=font)
            x += col_widths[idx]
        y += line_height
    return image


def print_image_cache_path(export_path: Path) -> Path:
    return export_path.with_name(f"{export_path.stem}.print.png")


def _trim_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    trimmed = list(rows)
    while trimmed and all(not str(cell).strip() for cell in trimmed[-1]):
        trimmed.pop()
    while trimmed and all(not str(cell).strip() for cell in trimmed[0]):
        trimmed.pop(0)
    return trimmed


def ensure_print_image(export_path: Path, sheet_name: str, print_area: str) -> Path:
    cache_path = print_image_cache_path(export_path)
    if cache_path.is_file() and cache_path.stat().st_mtime >= export_path.stat().st_mtime:
        return cache_path

    rows = _trim_empty_rows(read_print_area_values(export_path, sheet_name, print_area))
    if not rows:
        raise ValueError("打印区域为空")

    image = _rows_to_pil_image(rows)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(
        cache_path,
        format="PNG",
        dpi=(PRINT_RENDER_DPI, PRINT_RENDER_DPI),
        compress_level=1,
    )
    return cache_path


def _print_image_with_dialog(image_path: Path) -> None:
    if not _PRINT_SCRIPT.is_file():
        raise RuntimeError(f"缺少打印脚本: {_PRINT_SCRIPT}")

    completed = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-Sta",
            "-File",
            str(_PRINT_SCRIPT),
            str(image_path),
        ],
        timeout=600,
        check=False,
    )
    if completed.returncode == 2:
        return
    if completed.returncode != 0:
        raise RuntimeError(f"打印失败 (exit {completed.returncode})")


def open_print_preview_dialog(image_path: Path) -> None:
    if sys.platform != "win32":
        raise RuntimeError("打印预览当前仅支持 Windows")
    if not image_path.is_file():
        raise FileNotFoundError(f"打印图片不存在: {image_path}")
    _print_image_with_dialog(image_path)


def show_print_dialog(xlsx_path: Path, sheet_name: str, print_area: str) -> None:
    image_path = ensure_print_image(xlsx_path, sheet_name, print_area)
    open_print_preview_dialog(image_path)
