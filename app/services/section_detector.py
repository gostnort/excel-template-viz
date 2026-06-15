"""
Section Detector - Multi-area detection for Excel templates

Detects repeated sections in Excel templates based on YAML configuration.
"""
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class AreaCoords:
    """Excel area coordinates"""
    start_row: int  # 1-based
    start_col: int  # 1-based
    end_row: int
    end_col: int
    
    def __str__(self) -> str:
        """Convert to Excel range string like 'A1:M2'"""
        return f"{self._col_to_letter(self.start_col)}{self.start_row}:{self._col_to_letter(self.end_col)}{self.end_row}"
    
    @staticmethod
    def _col_to_letter(col: int) -> str:
        """Convert column number to Excel letter"""
        result = ""
        while col > 0:
            col -= 1
            result = chr(65 + col % 26) + result
            col //= 26
        return result


@dataclass
class SectionConfig:
    """Section configuration from YAML"""
    input_area: str  # e.g., "A1:M2"
    move_to: str     # "down" | "up" | "left" | "right"
    offset: int      # number of rows/columns to move


@dataclass
class DetectedArea:
    """A detected area instance"""
    index: int  # 1-based area index
    area: str   # Excel range string
    coords: AreaCoords
    has_data: bool  # Whether area has non-formula content


def parse_area_range(area_str: str) -> AreaCoords:
    """
    Parse Excel area string to coordinates
    
    Args:
        area_str: Excel range like "A1:M2"
    
    Returns:
        AreaCoords with 1-based row/column indices
    
    Examples:
        "A1:M2" -> (1, 1, 2, 13)
        "B5:E10" -> (5, 2, 10, 5)
    """
    # Pattern: "A1:M2"
    pattern = r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$'
    match = re.match(pattern, area_str.strip().upper())
    
    if not match:
        raise ValueError(f"Invalid Excel range format: {area_str}")
    
    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()
    
    start_row = int(start_row_str)
    end_row = int(end_row_str)
    start_col = _letter_to_col(start_col_letter)
    end_col = _letter_to_col(end_col_letter)
    
    if start_row > end_row or start_col > end_col:
        raise ValueError(f"Invalid range: start must be before end: {area_str}")
    
    return AreaCoords(start_row, start_col, end_row, end_col)


def _letter_to_col(letter: str) -> int:
    """Convert Excel column letter to number (1-based)"""
    col = 0
    for char in letter:
        col = col * 26 + (ord(char) - 64)
    return col


def calculate_next_area(
    input_area: str,
    move_to: str,
    offset: int
) -> str:
    """
    Calculate next area coordinates
    
    Args:
        input_area: Current area e.g., "A1:M2"
        move_to: Direction ("down", "up", "left", "right")
        offset: Offset amount (rows or columns)
    
    Returns:
        Next area string
    
    Examples:
        ("A1:M2", "down", 2) -> "A3:M4"
        ("A1:M2", "right", 3) -> "D1:P2"
    """
    coords = parse_area_range(input_area)
    
    if move_to == "down":
        coords.start_row += offset
        coords.end_row += offset
    elif move_to == "up":
        coords.start_row -= offset
        coords.end_row -= offset
    elif move_to == "right":
        coords.start_col += offset
        coords.end_col += offset
    elif move_to == "left":
        coords.start_col -= offset
        coords.end_col -= offset
    else:
        raise ValueError(f"Invalid move_to direction: {move_to}")
    
    # Validate bounds
    if coords.start_row < 1 or coords.start_col < 1:
        raise ValueError(f"Area moved out of bounds: {coords}")
    
    return str(coords)


def is_cell_empty_content(cell: Cell) -> bool:
    """
    Check if cell is empty content
    
    Rules:
    - cell.value is None -> empty
    - cell.value is empty string -> empty
    - cell.data_type == 'f' (formula) -> empty
    - border, fill, font do not count as content
    
    Args:
        cell: openpyxl Cell object
    
    Returns:
        True if empty, False if has content
    """
    if cell is None:
        return True
    
    # Formula cells are considered empty
    if cell.data_type == 'f':
        return True
    
    # Check value
    value = cell.value
    if value is None:
        return True
    
    if isinstance(value, str) and not value.strip():
        return True
    
    # Has actual content
    return False


def _cell_content_key(value: Any) -> str:
    """Normalize a cell value for exact row-content comparison."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _read_area_content(ws: Worksheet, coords: AreaCoords) -> list[list[Any]]:
    """Read cell values from area (excluding formulas)"""
    content = []
    
    for row_idx in range(coords.start_row, coords.end_row + 1):
        row_values = []
        for col_idx in range(coords.start_col, coords.end_col + 1):
            cell = ws.cell(row_idx, col_idx)
            
            # Skip formula cells (count as empty)
            if cell.data_type == 'f':
                row_values.append(None)
            else:
                row_values.append(cell.value)
        
        content.append(row_values)
    
    return content


def _area_content_signature(ws: Worksheet, coords: AreaCoords) -> tuple[str, ...]:
    """Flat tuple of normalized values across the whole input area."""
    signature: list[str] = []
    for row_idx in range(coords.start_row, coords.end_row + 1):
        for col_idx in range(coords.start_col, coords.end_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.data_type == 'f':
                signature.append("")
            else:
                signature.append(_cell_content_key(cell.value))
    return tuple(signature)


def _is_area_completely_empty(ws: Worksheet, coords: AreaCoords) -> bool:
    """
    Check if area is completely empty
    
    Empty means: no non-formula values
    (border, color, formula, text_format don't count as content)
    """
    for row_idx in range(coords.start_row, coords.end_row + 1):
        for col_idx in range(coords.start_col, coords.end_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if not is_cell_empty_content(cell):
                return False
    
    return True


def _areas_have_same_format(
    ws: Worksheet,
    area1_coords: AreaCoords,
    area2_coords: AreaCoords
) -> bool:
    """
    Check if two areas have the same format/structure
    
    Compare non-formula content positions (not values, just where content exists)
    """
    content1 = _read_area_content(ws, area1_coords)
    content2 = _read_area_content(ws, area2_coords)
    
    # Must have same dimensions
    if len(content1) != len(content2):
        return False
    
    # Check if empty/non-empty pattern matches
    for row1, row2 in zip(content1, content2):
        if len(row1) != len(row2):
            return False
        
        for val1, val2 in zip(row1, row2):
            is_empty1 = val1 is None or (isinstance(val1, str) and not val1.strip())
            is_empty2 = val2 is None or (isinstance(val2, str) and not val2.strip())
            
            # Pattern mismatch: one empty, one not
            if is_empty1 != is_empty2:
                return False
    
    return True


def _area_within_sheet(ws: Worksheet, coords: AreaCoords) -> bool:
    if coords.start_row < 1 or coords.start_col < 1:
        return False
    if coords.end_row > ws.max_row or coords.end_col > ws.max_column:
        return False
    return True


def _scan_homogeneous_group(
    ws: Worksheet,
    start_area_str: str,
    section_config: SectionConfig,
) -> tuple[list[str], str | None]:
    """
    Collect consecutive single-step areas whose full-row content matches exactly.

    Returns:
        (area strings in the group, next area to scan after the group or None)
    """
    group: list[str] = []
    current_area_str = start_area_str
    reference_signature: tuple[str, ...] | None = None
    while True:
        try:
            coords = parse_area_range(current_area_str)
        except ValueError:
            break
        if not _area_within_sheet(ws, coords):
            break
        if _is_area_completely_empty(ws, coords):
            break
        signature = _area_content_signature(ws, coords)
        if reference_signature is None:
            reference_signature = signature
            group.append(current_area_str)
        elif signature == reference_signature:
            group.append(current_area_str)
        else:
            return group, current_area_str
        try:
            current_area_str = calculate_next_area(
                current_area_str,
                section_config.move_to,
                section_config.offset,
            )
        except ValueError:
            break
    return group, None


def detect_multi_areas(
    workbook_path: Path,
    sheet_name: str,
    section_config: SectionConfig
) -> list[DetectedArea]:
    """
    Detect the first contiguous block of identical input rows from YAML input_area.

    Each returned area is one dropdown entry. Only consecutive rows whose full
    area content matches exactly are included; a content change, empty row, or
    sheet end stops the block. Later differing groups are ignored.
    """
    logger.info(f"Detecting areas in {sheet_name} with config: {section_config}")
    
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    detected_areas: list[DetectedArea] = []
    scan_area_str = section_config.input_area
    try:
        scan_coords = parse_area_range(scan_area_str)
    except ValueError:
        wb.close()
        return detected_areas
    if not _area_within_sheet(ws, scan_coords):
        wb.close()
        return detected_areas
    if _is_area_completely_empty(ws, scan_coords):
        wb.close()
        return detected_areas
    group_areas, _next_scan = _scan_homogeneous_group(ws, scan_area_str, section_config)
    for area_index, area_str in enumerate(group_areas, start=1):
        coords = parse_area_range(area_str)
        has_data = not _is_area_completely_empty(ws, coords)
        detected_areas.append(DetectedArea(
            index=area_index,
            area=area_str,
            coords=coords,
            has_data=has_data,
        ))
        logger.info(f"Detected area {area_index}: {area_str}, has_data: {has_data}")
    
    wb.close()
    
    logger.info(f"Total detected areas: {len(detected_areas)}")
    return detected_areas


def parse_sections_from_yaml(yaml_dict: dict[str, Any]) -> list[SectionConfig] | None:
    """
    Parse sections configuration from YAML
    
    Args:
        yaml_dict: Loaded YAML dictionary
    
    Returns:
        List of SectionConfig objects, or None if no sections defined
        
    Raises:
        ValueError: If move_to direction is invalid or offset is not positive
    """
    sections_data = yaml_dict.get("sections")
    
    if not sections_data:
        return None
    
    if not isinstance(sections_data, list):
        raise ValueError("sections must be a list")
    
    VALID_DIRECTIONS = {"down", "up", "left", "right"}
    
    sections = []
    for section_dict in sections_data:
        if not isinstance(section_dict, dict):
            continue
        
        input_area = section_dict.get("input_area")
        move_to = section_dict.get("move_to")
        offset = section_dict.get("offset")
        
        if not all([input_area, move_to, offset is not None]):
            logger.warning(f"Incomplete section config: {section_dict}")
            continue
        
        # Validate move_to direction
        move_to_lower = str(move_to).lower()
        if move_to_lower not in VALID_DIRECTIONS:
            raise ValueError(
                f"Invalid move_to direction: '{move_to}'. "
                f"Must be one of {VALID_DIRECTIONS}"
            )
        
        # Validate offset is positive integer
        try:
            offset_int = int(offset)
        except (ValueError, TypeError) as e:
            raise ValueError(f"offset must be an integer, got: {offset}") from e
        
        if offset_int <= 0:
            raise ValueError(f"offset must be a positive integer, got: {offset_int}")
        
        sections.append(SectionConfig(
            input_area=str(input_area),
            move_to=move_to_lower,
            offset=offset_int
        ))
    
    return sections if sections else None
