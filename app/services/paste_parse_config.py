import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml

from app.services.registry import TEMPLATES_DIR

PASTE_CONFIG_SUFFIX = ".paste.yaml"
DEFAULT_FIELDS_PER_ROW = 7
UNMAPPED_FILED = "?"
UNMAPPED_INDEX = -1
STRUCTURAL_ORDER_COLUMN = "order"
RESERVED_TOP_KEYS = frozenset({"determiner", "order", "worksheet", "sections", "fields_per_row"})


@dataclass
class PasteParseRule:
    filed: str
    index: int
    regex: str | None = None
    id_flag: bool = False


@dataclass
class PasteParseConfig:
    determiner: str
    field_rules: dict[str, list[PasteParseRule]]
    order: list[dict[str, Any]] | None = None
    worksheet: str | None = None
    sections: list[dict[str, Any]] | None = None  # Section configurations for multi-area detection
    fields_per_row: int = DEFAULT_FIELDS_PER_ROW
    
    def to_dict(self) -> dict[str, Any]:
        """
        Convert config to dict for field matching and other operations
        
        Returns:
            Dict representation of the config with field rules converted to dict format
        """
        result: dict[str, Any] = {}
        
        # Convert field_rules to dict format
        for field_name, rules in self.field_rules.items():
            result[field_name] = [_rule_to_dict(rule) for rule in rules]
        
        # Add optional fields if present
        if self.order:
            result["order"] = [
                _order_entry_to_dict(item) if isinstance(item, dict) else item
                for item in self.order
            ]
        if self.worksheet:
            result["worksheet"] = self.worksheet
        if self.sections:
            result["sections"] = self.sections
        
        result["determiner"] = self.determiner
        result["fields_per_row"] = self.fields_per_row
        
        return result


def paste_config_path(template_id: str) -> Path:
    """
    Get path to paste config file
    
    Returns: templates/{template_id}/{template_id}.paste.yaml
    """
    return TEMPLATES_DIR / template_id / f"{template_id}{PASTE_CONFIG_SUFFIX}"


def _normalize_determiner(raw: str) -> str:
    value = raw.strip().lower()
    if value in {"tab", "\\t"}:
        return "tab"
    return value


def _split_line(line: str, determiner: str) -> list[str]:
    if determiner == "tab":
        return line.split("\t")
    if determiner == "space":
        return line.split()
    return line.split(determiner)


def _normalize_regex_value(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if text in {"", "None", "null"}:
        return None
    return text


def _is_unmapped(filed: str, index: int) -> bool:
    return (not filed or filed == UNMAPPED_FILED) or index == UNMAPPED_INDEX


def is_structural_order_column(name: str) -> bool:
    """True when a sheet/template header is the reserved paste column ``order``."""
    return name.strip().lower() == STRUCTURAL_ORDER_COLUMN


def structural_order_col_offset(order: list[dict[str, Any]] | None) -> int:
    """Return 1 when paste order reserves column A for structural ``order``."""
    if not order:
        return 0
    for item in order:
        if not isinstance(item, dict):
            continue
        filed = str(item.get("filed", "")).strip()
        index = int(item.get("index", UNMAPPED_INDEX))
        if index == 0 and is_structural_order_column(filed):
            return 1
    return 0


def default_input_area_for_template(
    template_path: Path,
    worksheet_name: str | None,
    data_start_row: int,
) -> str | None:
    """Infer a single-row input area from the template header row and data_start_row."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from app.services.excel_parser import resolve_sheet_name

    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        resolved_sheet = resolve_sheet_name(template_path, worksheet_name) if worksheet_name else None
        if resolved_sheet:
            ws = wb[resolved_sheet]
        elif worksheet_name and worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.active
        if ws is None:
            return None
        last_col = 0
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None and str(cell.value).strip():
                last_col = col_idx
        if last_col < 1:
            return None
        excel_row = data_start_row + 1
        return f"A{excel_row}:{get_column_letter(last_col)}{excel_row}"
    finally:
        wb.close()


def read_input_area_headers(
    template_path: Path,
    worksheet_name: str | None,
    area_range: str,
    *,
    header_row: int | None = None,
) -> list[str]:
    """Read column headers for an input area from the row above the area (or template header row)."""
    from openpyxl import load_workbook

    from app.services.excel_parser import resolve_sheet_name
    from app.services.section_detector import parse_area_range

    coords = parse_area_range(area_range)
    if coords.start_row > 1:
        excel_header_row = coords.start_row - 1
    elif header_row is not None:
        excel_header_row = header_row + 1
    else:
        excel_header_row = 1
    resolved_sheet = resolve_sheet_name(template_path, worksheet_name) if worksheet_name else None
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        if resolved_sheet:
            ws = wb[resolved_sheet]
        elif worksheet_name and worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.active
        if ws is None:
            return []
        headers: list[str] = []
        for col in range(coords.start_col, coords.end_col + 1):
            cell_value = ws.cell(excel_header_row, col).value
            if cell_value is not None:
                text = str(cell_value).strip()
                if text:
                    headers.append(text)
            # Skip blank header cells so wide input areas do not add placeholder columns.
        while headers and headers[-1].startswith("col_"):
            headers.pop()
        seen: set[str] = set()
        unique_headers: list[str] = []
        for header in headers:
            key = header.strip()
            if not key or key in seen:
                continue
            seen.add(key)
            unique_headers.append(header)
        return unique_headers
    finally:
        wb.close()


def read_template_header_at(
    template_path: Path,
    worksheet_name: str | None,
    col_index: int,
) -> str | None:
    """Read one header cell from the template's first row (0-based column index)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if worksheet_name and worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.active
        if ws is None:
            return None
        cell_value = ws.cell(1, col_index + 1).value
        if cell_value is None:
            return None
        text = str(cell_value).strip()
        return text or None
    finally:
        wb.close()


def structural_order_entry_from_template(
    template_path: Path,
    worksheet_name: str | None = None,
) -> dict[str, Any] | None:
    """Return a paste order entry when column A of the template is structural ``order``."""
    header = read_template_header_at(template_path, worksheet_name, 0)
    if header and is_structural_order_column(header):
        return _order_entry_to_dict({"filed": header, "index": 0})
    return None


def resolve_structural_order_entry(
    order: list[dict[str, Any]] | None,
    template_path: Path | None = None,
    worksheet_name: str | None = None,
) -> dict[str, Any] | None:
    """Prefer configured paste order; fall back to the template's column-A header."""
    if order and structural_order_col_offset(order) == 1:
        return _order_entry_to_dict(order[0])
    if template_path is not None:
        return structural_order_entry_from_template(template_path, worksheet_name)
    return None


def structural_order_active(
    order: list[dict[str, Any]] | None,
    template_path: Path | None = None,
    worksheet_name: str | None = None,
) -> bool:
    """True when YAML or the Excel template reserves column A for ``order``."""
    return resolve_structural_order_entry(order, template_path, worksheet_name) is not None


def build_order_entries_from_mappings(
    sheet_columns: list[str],
    mapped_columns: set[str],
    *,
    include_unmapped: bool = True,
    structural_order_seed: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Build paste order list: index-0 ``order`` column plus matched sheet columns by index."""
    new_order: list[dict[str, Any]] = []
    seen: set[str] = set()
    if structural_order_seed:
        seeded = _order_entry_to_dict(structural_order_seed)
        if structural_order_col_offset([seeded]) == 1:
            new_order.append(seeded)
            seen.add(str(seeded.get("filed", "")).strip())
    for idx, col_name in enumerate(sheet_columns):
        col_stripped = col_name.strip()
        if not col_stripped:
            continue
        is_structural = idx == 0 and is_structural_order_column(col_stripped)
        if is_structural or col_stripped in mapped_columns:
            if col_stripped not in seen:
                new_order.append(_order_entry_to_dict({"filed": col_stripped, "index": idx}))
                seen.add(col_stripped)
    if not new_order and include_unmapped:
        new_order = [_default_order_entry()]
    return new_order


def _rule_to_dict(rule: PasteParseRule) -> dict[str, Any]:
    filed = (rule.filed or "").strip() or UNMAPPED_FILED
    if _is_unmapped(filed, rule.index):
        return {
            "ID": rule.id_flag,
            "filed": UNMAPPED_FILED,
            "index": UNMAPPED_INDEX,
            "regex": "None",
        }
    return {
        "ID": rule.id_flag,
        "filed": filed,
        "index": rule.index,
        "regex": "None" if rule.regex is None else rule.regex,
    }


def _order_entry_to_dict(entry: dict[str, Any]) -> dict[str, Any]:
    filed = (str(entry.get("filed", UNMAPPED_FILED)).strip()) or UNMAPPED_FILED
    index = int(entry.get("index", UNMAPPED_INDEX))
    if _is_unmapped(filed, index):
        return _default_order_entry()
    regex = _normalize_regex_value(entry.get("regex"))
    return {
        "ID": bool(entry.get("ID", False)),
        "filed": filed,
        "index": index,
        "regex": "None" if regex is None else regex,
    }


def _default_unmapped_rule(id_flag: bool = False) -> PasteParseRule:
    return PasteParseRule(
        filed=UNMAPPED_FILED,
        index=UNMAPPED_INDEX,
        regex=None,
        id_flag=id_flag,
    )


def _default_order_entry() -> dict[str, Any]:
    return {
        "ID": False,
        "filed": UNMAPPED_FILED,
        "index": UNMAPPED_INDEX,
        "regex": "None",
    }


def _parse_rules(raw_rules: Any) -> list[PasteParseRule]:
    if not isinstance(raw_rules, list):
        return []
    rules: list[PasteParseRule] = []
    for item in raw_rules:
        if not isinstance(item, dict):
            continue
        if "index" not in item:
            continue
        filed_raw = str(item.get("filed", UNMAPPED_FILED)).strip()
        filed = filed_raw if filed_raw else UNMAPPED_FILED
        rules.append(
            PasteParseRule(
                filed=filed,
                index=int(item["index"]),
                regex=_normalize_regex_value(item.get("regex")),
                id_flag=bool(item.get("ID", False)),
            )
        )
    return rules


def config_from_dict(raw: dict[str, Any]) -> PasteParseConfig | None:
    field_rules: dict[str, list[PasteParseRule]] = {}
    for key, value in raw.items():
        if key in RESERVED_TOP_KEYS:
            continue
        rules = _parse_rules(value)
        if rules:
            field_rules[str(key)] = rules
    if not field_rules:
        return None
    order = raw.get("order")
    if order is not None and not isinstance(order, list):
        order = None
    
    worksheet = raw.get("worksheet")
    if worksheet is not None:
        worksheet = str(worksheet).strip()
    
    # Parse sections configuration
    sections = raw.get("sections")
    if sections is not None and not isinstance(sections, list):
        sections = None

    fields_per_row = raw.get("fields_per_row", DEFAULT_FIELDS_PER_ROW)
    if not isinstance(fields_per_row, int) or fields_per_row < 1:
        fields_per_row = DEFAULT_FIELDS_PER_ROW
    
    return PasteParseConfig(
        determiner=_normalize_determiner(str(raw.get("determiner", "tab"))),
        field_rules=field_rules,
        order=order,
        worksheet=worksheet,
        sections=sections,
        fields_per_row=fields_per_row,
    )


def load_paste_parse_config(template_id: str) -> PasteParseConfig | None:
    path = paste_config_path(template_id)
    if not path.exists():
        return None
    try:
        raw = _safe_load_mapping_yaml(path.read_text(encoding="utf-8"))
    except yaml.YAMLError:
        return None
    if not isinstance(raw, dict):
        return None
    try:
        raw = _normalize_field_rule_lists(raw)
    except ValueError:
        return None
    return config_from_dict(raw)


def save_paste_parse_yaml(template_id: str, yaml_text: str, template_headers: list[str] | None = None) -> None:
    cleaned = extract_yaml_text(yaml_text)
    if template_headers is not None:
        config = validate_mapping_yaml(cleaned, template_headers)
        cleaned = config_to_yaml(config)
    else:
        parsed = _safe_load_mapping_yaml(cleaned)
        if not isinstance(parsed, dict) or config_from_dict(parsed) is None:
            raise ValueError("YAML must contain at least one template field mapping")
        cleaned = config_to_yaml(parsed)
    path = paste_config_path(template_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(cleaned, encoding="utf-8")


_RULE_KEY_ORDER = ("ID", "filed", "index", "regex")


def _yaml_scalar(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, str):
        if "regex" in value or "\\" in value or "(" in value:
            escaped = value.replace("'", "''")
            return f"'{escaped}'"
        if value in {"", "?", "tab"} or " " in value or "#" in value or "." in value:
            return f'"{value}"'
        return f'"{value}"'
    return f'"{value}"'


def _normalize_rule_dict(rule: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(rule, dict):
        return _default_order_entry()
    filed = (str(rule.get("filed", UNMAPPED_FILED)).strip()) or UNMAPPED_FILED
    index = int(rule.get("index", UNMAPPED_INDEX))
    if _is_unmapped(filed, index):
        return {
            "ID": bool(rule.get("ID", False)),
            "filed": UNMAPPED_FILED,
            "index": UNMAPPED_INDEX,
            "regex": "None",
        }
    regex = _normalize_regex_value(rule.get("regex"))
    return {
        "ID": bool(rule.get("ID", False)),
        "filed": filed,
        "index": index,
        "regex": "None" if regex is None else regex,
    }


def _format_rule_lines(rule: dict[str, Any], indent: int) -> list[str]:
    rule = _normalize_rule_dict(rule)
    if not isinstance(rule, dict):
        return []
    keys = [key for key in _RULE_KEY_ORDER if key in rule]
    keys.extend(key for key in rule if key not in keys)
    lines: list[str] = []
    for idx, key in enumerate(keys):
        if idx == 0:
            lines.append(f"{' ' * indent}- {key}: {_yaml_scalar(rule[key])}")
        else:
            lines.append(f"{' ' * (indent + 2)}{key}: {_yaml_scalar(rule[key])}")
    return lines


def _hoist_sections_from_order(raw: dict[str, Any]) -> dict[str, Any]:
    """Move ``sections`` out of ``order`` mappings to the document root."""
    normalized = dict(raw)
    order = normalized.get("order")
    if order is None:
        return normalized
    if isinstance(order, dict):
        order_dict = dict(order)
        nested_sections = order_dict.pop("sections", None)
        if nested_sections is not None and normalized.get("sections") is None:
            normalized["sections"] = nested_sections
        normalized["order"] = [order_dict]
        order = normalized["order"]
    if not isinstance(order, list):
        return normalized
    clean_order: list[Any] = []
    for item in order:
        if not isinstance(item, dict):
            clean_order.append(item)
            continue
        item_copy = dict(item)
        nested_sections = item_copy.pop("sections", None)
        if nested_sections is not None and normalized.get("sections") is None:
            normalized["sections"] = nested_sections
        clean_order.append(item_copy)
    normalized["order"] = clean_order
    return normalized


def _normalize_field_rule_lists(raw: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(raw)
    for key, value in raw.items():
        if key in RESERVED_TOP_KEYS:
            if key == "order" and value is not None and not isinstance(value, (list, dict)):
                raise ValueError(f"'{key}' must be a YAML list; each entry must start with '-'")
            continue
        if isinstance(value, dict):
            normalized[key] = [value]
            continue
        if not isinstance(value, list):
            raise ValueError(
                f"Field {key!r} must be a YAML list. Put '-' before each rule, e.g.\n"
                f"{key}:\n  - filed: \"?\"\n    index: -1"
            )
        for item in value:
            if not isinstance(item, dict):
                raise ValueError(f"Field {key!r}: each '-' item must be a mapping with filed/index")
    return _hoist_sections_from_order(normalized)


def _meaningful_order_entries(order: list[Any]) -> list[dict[str, Any]]:
    """Keep order entries that reference a mapped sheet column."""
    meaningful: list[dict[str, Any]] = []
    for item in order:
        if not isinstance(item, dict):
            continue
        filed = (str(item.get("filed", UNMAPPED_FILED)).strip()) or UNMAPPED_FILED
        index = int(item.get("index", UNMAPPED_INDEX))
        if _is_unmapped(filed, index):
            continue
        meaningful.append(item)
    return meaningful


def _format_section_lines(sections: list[dict[str, Any]]) -> list[str]:
    lines: list[str] = ["sections:"]
    for section in sections:
        if not isinstance(section, dict):
            continue
        lines.append("  - input_area: " + _yaml_scalar(str(section.get("input_area", ""))))
        lines.append("    move_to: " + _yaml_scalar(str(section.get("move_to", ""))))
        lines.append("    offset: " + str(section.get("offset", 0)))
    return lines


def config_to_yaml(config: dict[str, Any], *, omit_unmapped_fields: bool = False) -> str:
    config = _normalize_field_rule_lists(config)
    lines: list[str] = [f'determiner: {_yaml_scalar(str(config.get("determiner", "tab")))}']
    fields_per_row = config.get("fields_per_row", DEFAULT_FIELDS_PER_ROW)
    if not isinstance(fields_per_row, int) or fields_per_row < 1:
        fields_per_row = DEFAULT_FIELDS_PER_ROW
    lines.append(f"fields_per_row: {fields_per_row}")
    worksheet = config.get("worksheet")
    if worksheet is not None:
        lines.append(f'worksheet: {_yaml_scalar(str(worksheet))}')
    order = config.get("order")
    if isinstance(order, list) and order:
        meaningful_order = _meaningful_order_entries(order)
        if meaningful_order:
            lines.append("order:")
            for item in meaningful_order:
                lines.extend(_format_rule_lines(item, indent=2))
    sections = config.get("sections")
    if isinstance(sections, list) and sections:
        lines.extend(_format_section_lines(sections))
    for key, value in config.items():
        if key in RESERVED_TOP_KEYS or key == "determiner" or key == "fields_per_row":
            continue
        if not isinstance(value, list):
            continue
        field_lines: list[str] = []
        for rule in value:
            if isinstance(rule, dict):
                if omit_unmapped_fields:
                    filed = (str(rule.get("filed", UNMAPPED_FILED)).strip()) or UNMAPPED_FILED
                    index = int(rule.get("index", UNMAPPED_INDEX))
                    if _is_unmapped(filed, index):
                        continue
                field_lines.extend(_format_rule_lines(rule, indent=2))
        if not field_lines:
            continue
        lines.append(str(key) + ":")
        lines.extend(field_lines)
    return "\n".join(lines) + "\n"


def build_empty_mapping_yaml(template_headers: list[str]) -> str:
    config: dict[str, Any] = {"determiner": "tab"}
    order_entries: list[dict[str, Any]] = [_default_order_entry()]
    for col_idx, header in enumerate(template_headers):
        header_stripped = header.strip()
        if not header_stripped:
            continue
        if is_structural_order_column(header_stripped):
            order_entries = [_order_entry_to_dict({"filed": header_stripped, "index": col_idx})]
            continue
        config[header_stripped] = [_rule_to_dict(_default_unmapped_rule())]
    config["order"] = order_entries
    return config_to_yaml(config)


_REGEX_DOUBLE_QUOTED = re.compile(r'^(\s*regex:\s*)"(.*)"\s*$')


def _normalize_yaml_regex_quotes(yaml_text: str) -> str:
    # YAML double-quoted scalars treat \d as invalid escapes; regex must use single quotes.
    lines: list[str] = []
    for line in yaml_text.splitlines():
        match = _REGEX_DOUBLE_QUOTED.match(line)
        if not match:
            lines.append(line)
            continue
        prefix, inner = match.group(1), match.group(2)
        inner = inner.replace("\\\\", "\\")
        escaped = inner.replace("'", "''")
        lines.append(f"{prefix}'{escaped}'")
    return "\n".join(lines)


def _safe_load_mapping_yaml(yaml_text: str) -> Any:
    normalized = _normalize_yaml_regex_quotes(yaml_text)
    return yaml.safe_load(normalized)


def extract_yaml_text(model_output: str) -> str:
    text = model_output.strip()
    if not text:
        return text

    fenced = re.search(
        r"```(?:yaml|yml)?\s*\r?\n?(.*?)```",
        text,
        flags=re.DOTALL | re.IGNORECASE,
    )
    if fenced:
        return fenced.group(1).strip()

    if text.startswith("```"):
        text = re.sub(r"^```(?:yaml|yml)?\s*\r?\n?", "", text, count=1, flags=re.IGNORECASE)
        text = re.sub(r"```\s*$", "", text.strip())

    return text.strip()


def validate_mapping_yaml(yaml_text: str, template_headers: list[str]) -> dict[str, Any]:
    normalized = extract_yaml_text(yaml_text)
    try:
        parsed = _safe_load_mapping_yaml(normalized)
    except yaml.YAMLError as exc:
        raise ValueError(f"Invalid YAML: {exc}") from exc
    if not isinstance(parsed, dict):
        raise ValueError("Model output is not a valid YAML object")
    parsed = _normalize_field_rule_lists(parsed)
    config = config_from_dict(parsed)
    if config is None:
        raise ValueError("YAML must contain at least one template field mapping")
    allowed = {header.strip() for header in template_headers}
    for field_name in config.field_rules:
        if field_name.strip() not in allowed:
            raise ValueError(
                f"Unknown template field {field_name!r}; check that it exists in the template columns."
            )
    parsed.setdefault("determiner", config.determiner)
    return parsed


def id_target_field_from_config(config: PasteParseConfig | None) -> str | None:
    if config is None:
        return None
    for field_name, rules in config.field_rules.items():
        for rule in rules:
            if rule.id_flag:
                return field_name
    return None


def id_column_from_config(config: PasteParseConfig | None) -> str | None:
    if config is None:
        return None
    for field_name, rules in config.field_rules.items():
        for rule in rules:
            if rule.id_flag and rule.filed and rule.filed != "?":
                return rule.filed
    return None


def resolve_sheet_header(filed: str, sheet_headers: list[str]) -> str | None:
    if not filed or filed == "?":
        return None
    filed_stripped = filed.strip()
    for h in sheet_headers:
        if h == filed_stripped:
            return h
    filed_lower = filed_stripped.lower()
    for h in sheet_headers:
        if h.strip().lower() == filed_lower:
            return h
    return None


def map_sheet_row_from_paste_config(
    row: dict[str, str],
    config: PasteParseConfig,
) -> dict[str, str]:
    parsed: dict[str, str] = {}
    sheet_headers = list(row.keys())
    for field_name, rules in config.field_rules.items():
        for rule in rules:
            if not rule.filed or rule.filed == "?":
                continue
            resolved_header = resolve_sheet_header(rule.filed, sheet_headers)
            if resolved_header is None:
                continue
            raw_val = row.get(resolved_header, "")
            if raw_val is None:
                raw_val = ""
            raw_val = str(raw_val).strip()
            if not raw_val:
                continue
            if rule.regex:
                extracted = _extract_with_regex(raw_val, rule.regex)
                if extracted is None:
                    continue
                val = extracted
            else:
                val = raw_val
            parsed[field_name] = _format_field_value(field_name, val)
            break
    return parsed


def validate_yaml_against_sheet_headers(
    config: PasteParseConfig,
    sheet_headers: list[str],
) -> dict[str, Any]:
    matched: dict[str, str] = {}
    missing: list[str] = []
    id_filed = id_column_from_config(config)
    id_matched = True
    if id_filed:
        resolved_id_header = resolve_sheet_header(id_filed, sheet_headers)
        if not resolved_id_header:
            id_matched = False
    for field_name, rules in config.field_rules.items():
        for rule in rules:
            if not rule.filed or rule.filed == "?":
                continue
            resolved = resolve_sheet_header(rule.filed, sheet_headers)
            if resolved:
                matched[rule.filed] = resolved
            else:
                if rule.filed not in missing:
                    missing.append(rule.filed)
    return {
        "matched": matched,
        "missing": missing,
        "id_matched": id_matched,
    }


def resolve_id_target_field(
    template_id: str,
    data_source_config: Any,
    headers: list[str],
) -> str | None:
    from app.services.data_source import id_target_field

    paste_config = load_paste_parse_config(template_id)
    paste_id = id_target_field_from_config(paste_config)
    header_by_stripped = {header.strip(): header for header in headers}
    if paste_id and paste_id.strip() in header_by_stripped:
        return header_by_stripped[paste_id.strip()]
    return id_target_field(data_source_config, headers)


def _safe_regex_search(pattern: str, raw: str) -> re.Match[str] | None:
    try:
        return re.search(pattern, raw)
    except re.error:
        pass
    if "(?<=" in pattern and r"\d{1,2}" in pattern:
        alt = re.sub(r"\(\?<=[^)]+\)", r"(?:\\d{1,2}/)", pattern, count=1)
        try:
            return re.search(alt, raw)
        except re.error:
            return None
    return None


def _extract_with_regex(raw: str, pattern: str) -> str | None:
    match = _safe_regex_search(pattern, raw)
    if not match:
        return None
    if match.lastindex and match.lastindex >= 1:
        return match.group(1).strip()
    return match.group(0).strip()


def _format_field_value(field_name: str, value: str) -> str:
    stripped_name = field_name.strip()
    if stripped_name in {"MM", "DD"} and value.isdigit():
        return value.zfill(2)
    if stripped_name == "Receiving Date" and "/" in value:
        parts = value.split("/", 1)
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return f"{parts[0].zfill(2)}/{parts[1].zfill(2)}"
    return value


def _apply_field_rule(parts: list[str], rule: PasteParseRule) -> str | None:
    if rule.index < 0 or rule.index >= len(parts):
        return None
    raw = parts[rule.index].strip()
    if not raw:
        return None
    if rule.regex:
        extracted = _extract_with_regex(raw, rule.regex)
        if not extracted:
            return None
        return extracted
    return raw


def parse_line_with_config(
    line: str,
    config: PasteParseConfig,
    order: int,
    reference_year: int | None = None,
) -> dict[str, str]:
    del reference_year  # retained for call-site compatibility
    parts = _split_line(line, config.determiner)
    if not parts:
        raise ValueError("Empty line cannot be parsed")
    parsed: dict[str, str] = {"order": str(order)}
    for field_name, rules in config.field_rules.items():
        for rule in rules:
            value = _apply_field_rule(parts, rule)
            if value is None:
                continue
            parsed[field_name] = _format_field_value(field_name, value)
            break
    return parsed


def parse_text_with_config(
    text: str,
    config: PasteParseConfig,
    reference_year: int | None = None,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    order = 1
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        rows.append(parse_line_with_config(line, config, order, reference_year))
        order += 1
    return rows


def create_default_config_from_template(template_path: Path, worksheet_name: str | None = None) -> PasteParseConfig:
    """
    Create default configuration from Excel template
    
    Reads the first row of the template as field names and creates a basic configuration
    with default field mappings and sections (if applicable).
    
    Args:
        template_path: Path to the Excel template file
        worksheet_name: Optional worksheet name (uses first sheet if None)
        
    Returns:
        PasteParseConfig with default settings
        
    Raises:
        ValueError: If template cannot be read or has no data
    """
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
        
        # Select worksheet
        if worksheet_name and worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.active
        
        if ws is None:
            raise ValueError("No worksheet found in template")
        
        # Read first row as field names (0-based column index for paste YAML)
        header_cells: list[tuple[int, str]] = []
        last_col_with_data = 0
        for col_idx, cell in enumerate(ws[1]):
            value = cell.value
            if value is not None and str(value).strip():
                header_cells.append((col_idx, str(value).strip()))
                last_col_with_data = col_idx + 1
        if not header_cells:
            raise ValueError("Template first row has no field names")
        # Create field rules with unmapped defaults (filed="?", index=-1)
        field_rules: dict[str, list[PasteParseRule]] = {}
        order_entries: list[dict[str, Any]] = [_default_order_entry()]
        for col_idx, field_name in header_cells:
            if is_structural_order_column(field_name):
                order_entries = [_order_entry_to_dict({"filed": field_name, "index": col_idx})]
                continue
            field_rules[field_name] = [_default_unmapped_rule()]
        
        # Create sections configuration if there are multiple columns
        sections = None
        if len(header_cells) > 1:
            # Detect data area from row 2, column 1 to last column with data
            from openpyxl.utils import get_column_letter
            
            start_col = get_column_letter(1)
            end_col = get_column_letter(last_col_with_data)
            
            sections = [{
                "input_area": f"{start_col}2:{end_col}2",  # Second row as input area
                "move_to": "down",  # Move down by default
                "offset": 1  # Offset by 1 row
            }]
        
        wb.close()
        
        return PasteParseConfig(
            determiner="tab",
            field_rules=field_rules,
            order=order_entries,
            worksheet=ws.title if ws.title else None,
            sections=sections,
            fields_per_row=DEFAULT_FIELDS_PER_ROW,
        )
        
    except Exception as e:
        raise ValueError(f"Failed to create default config from template: {e}") from e


def ensure_config_exists(template_id: str, template_path: Path) -> bool:
    """
    Ensure a configuration file exists for the template
    
    Creates a default configuration if none exists.
    Migrates legacy flat-path configs when present.
    
    Args:
        template_id: Template identifier
        template_path: Path to the Excel template file
        
    Returns:
        True if config exists or was created successfully
    """
    config_path = paste_config_path(template_id)
    
    if config_path.exists():
        return True

    legacy_path = TEMPLATES_DIR / f"{template_id}{PASTE_CONFIG_SUFFIX}"
    if legacy_path.exists():
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(legacy_path.read_text(encoding="utf-8"), encoding="utf-8")
        return True
    
    try:
        default_config = create_default_config_from_template(template_path)
        config_dict = default_config.to_dict()
        yaml_text = config_to_yaml(config_dict)
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(yaml_text, encoding='utf-8')
        return True
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Failed to create default config for {template_id}: {e}")
        return False

