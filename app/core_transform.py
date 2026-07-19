"""外部数据源与 Input_sheet 读写（路径 B）。见 docs/data_flow_design.md。"""
from __future__ import annotations
import argparse
import json
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from PIL import Image, ImageDraw, ImageFont
from .core_toml import GetTomlValues, TomlDefault, _validate_id_rules, offset_cell


logger = logging.getLogger(__name__)


def _cell_empty(value: Any) -> bool:
    """单元格无有效文本内容时视为空。"""
    return value is None or str(value).strip() == ""


def _column_names_for_rule(rule: TomlDefault) -> list[str]:
    """数据源列查找顺序：先 field 再 Input_label。"""
    names: list[str] = []
    if rule.field:
        names.append(rule.field)
    if rule.Input_label not in names:
        names.append(rule.Input_label)
    return names


def _lookup_row_value(row: dict[str, Any], rule: TomlDefault) -> Any:
    """在已匹配的数据源行中按 rule 取列值。"""
    for key in _column_names_for_rule(rule):
        if key in row:
            return row[key]
    return None


def _load_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """
    函数名: _load_sheet_rows
    作用: 将本地 xlsx 工作表读为行字典列表（首行为表头）
    输入:
        workbook_path (Path) - xlsx 路径
        sheet_name (str) - 工作表名
    输出:
        list[dict[str, Any]] - 每行 {列标题: 值}
    """
    wb = load_workbook(workbook_path, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return []
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: list[dict[str, Any]] = []
        # 跳过整行空行，按表头键组装 dict
        for values in row_iter:
            if all(_cell_empty(v) for v in values):
                continue
            item: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header or idx >= len(values):
                    continue
                item[header] = values[idx]
            rows.append(item)
        return rows
    finally:
        wb.close()


def _find_row_by_id(rows: list[dict[str, Any]], id_column: str, id_value: Any) -> dict[str, Any] | None:
    """
    函数名: _find_row_by_id
    作用: 在数据源行列表中按 ID 列匹配一行
    输入:
        rows - 表数据
        id_column (str) - ID 列名（通常为 field=ID）
        id_value (Any) - 待查 ID
    输出:
        dict | None - 匹配行或 None
    """
    target = str(id_value).strip()
    for row in rows:
        if id_column not in row:
            continue
        cell = row[id_column]
        if cell is None:
            continue
        if str(cell).strip() == target:
            return row
        # Excel 数字 ID 可能与字符串形式比较
        try:
            if str(int(float(cell))) == target:
                return row
        except (ValueError, TypeError):
            pass
    return None


def _find_row_by_lookup_keys(
    rows: list[dict[str, Any]], lookup_keys: list[str], id_value: Any
) -> dict[str, Any] | None:
    """在数据源行列表中按 id_lookup_keys 全局 OR 匹配一行。"""
    for key in lookup_keys:
        matched = _find_row_by_id(rows, key, id_value)
        if matched is not None:
            return matched
    return None


def _cell_display_text(value: Any) -> str:
    """单元格显示文本（用于打印区指纹与渲染）。"""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _split_print_area_raw(raw: str, default_sheet: str) -> list[tuple[str, str]]:
    """将 ws.print_area 拆成 (sheet_name, A1_range) 列表。"""
    parts = [part.strip() for part in str(raw).split(",") if part.strip()]
    parsed: list[tuple[str, str]] = []
    for part in parts:
        if "!" in part:
            sheet_part, area_part = part.rsplit("!", 1)
            sheet_name = sheet_part.strip().strip("'\"")
            area_text = area_part.replace("$", "")
        else:
            sheet_name = default_sheet
            area_text = part.replace("$", "")
        parsed.append((sheet_name, area_text))
    return parsed


def _read_area_cell_grid(ws: Any, area: str) -> list[list[str]]:
    """读取单个 A1 区域内的单元格显示文本网格。"""
    min_col, min_row, max_col, max_row = range_boundaries(area)
    grid: list[list[str]] = []
    for row_idx in range(min_row, max_row + 1):
        row_cells: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            row_cells.append(_cell_display_text(ws.cell(row=row_idx, column=col_idx).value))
        grid.append(row_cells)
    return grid


def _area_content_key(cells: list[list[str]]) -> tuple[str, ...]:
    """可比较的打印区内容指纹。"""
    return tuple(cell for row in cells for cell in row)


def _area_preview_text(cells: list[list[str]], limit: int = 30) -> str:
    """从非空单元格拼短预览，供下拉标签使用。"""
    parts: list[str] = []
    for row in cells:
        for cell in row:
            if not cell:
                continue
            parts.append(cell)
            joined = " ".join(parts)
            if len(joined) >= limit:
                return joined[:limit]
    return " ".join(parts)[:limit]


def _area_display_label(sheet_name: str, area: str, cells: list[list[str]]) -> str:
    """人类可读打印区标签。"""
    preview = _area_preview_text(cells)
    base = f"{sheet_name}: {area}"
    if preview:
        return f"{base} ({preview})"
    return base


def _column_pixel_width(ws: Any, col_idx: int, scale: float) -> int:
    """openpyxl 列宽 → 像素（近似 Excel 默认）。"""
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    width = dim.width if dim is not None and dim.width is not None else 8.43
    return max(24, int(width * 7 * scale))


def _row_pixel_height(ws: Any, row_idx: int, scale: float) -> int:
    """openpyxl 行高 → 像素。"""
    dim = ws.row_dimensions.get(row_idx)
    height = dim.height if dim is not None and dim.height is not None else 15.0
    return max(16, int(height * scale * 1.33))


class Template2DB:
    """按 source_file / source_sheet / Input_label 读外部表并生成标准记录。"""

    def __init__(self, cfg: GetTomlValues) -> None:
        self.cfg = cfg


    def resolve_source_path(self, source_key: str) -> Path | None:
        """
        函数名: resolve_source_path
        作用: 将 [[sources]] 中的别名解析为磁盘路径
        输入:
            source_key (str) - 如 source1
        输出:
            Path | None - 路径未配置或为空时返回 None
        """
        if not source_key:
            return None
        for item in self.cfg.sources:
            if source_key not in item:
                continue
            raw = item[source_key]
            if raw is None or str(raw).strip() == "":
                return None
            return Path(str(raw))
        return None


    def apply_regex(self, value: Any, pattern: str | None) -> Any:
        """
        函数名: apply_regex
        作用: 对单元格值做正则提取；有捕获组取 group(1)
        输入:
            value (Any) - 原始值
            pattern (str | None) - TOML regex
        输出:
            Any - 提取结果或原值
        """
        if pattern is None or str(pattern).strip() == "":
            return value
        if value is None:
            return value
        text = str(value)
        match = re.search(pattern, text)
        if not match:
            return value
        if match.lastindex:
            return match.group(1)
        return match.group(0)


    def fetch_row_by_id(self, id_value: Any) -> dict[str, Any]:
        """
        函数名: fetch_row_by_id
        作用: 路径 B：按 ID 从各 source_sheet 取列产出 incoming（键为 Input_label）
        输入:
            id_value (Any) - 用户指定或 textbox 中的 ID
        输出:
            dict[str, Any] - incoming：键仅为 Input_label（不含 field 名 / 不含主键 id）
        """
        record: dict[str, Any] = {}
        id_info = _validate_id_rules(self.cfg)
        lookup_keys = id_info["id_lookup_keys"] or ["ID"]
        sheet_cache: dict[tuple[str, str], list[dict[str, Any]]] = {}
        # 每个 rule 可能指向不同 source_file / source_sheet
        for rule in self.cfg.field_rules:
            if not rule.source_file or not rule.source_sheet:
                continue
            source_path = self.resolve_source_path(rule.source_file)
            if source_path is None or not source_path.is_file():
                logger.warning("source path missing for %s", rule.source_file)
                continue
            cache_key = (str(source_path), rule.source_sheet)
            if cache_key not in sheet_cache:
                sheet_cache[cache_key] = _load_sheet_rows(source_path, rule.source_sheet)
            rows = sheet_cache[cache_key]
            matched = _find_row_by_lookup_keys(rows, lookup_keys, id_value)
            if matched is None:
                continue
            raw_value = _lookup_row_value(matched, rule)
            # 只写 Input_label 键；落库主键由 core_store 据 db_id 推导或自动生成
            record[rule.Input_label] = self.apply_regex(raw_value, rule.regex)
        return record




class ExcelWriter:
    """按 verify_toml 的 located + input_section k 组平移读写值格；定位用 Input_label 不用 index。"""

    def __init__(self, cfg: GetTomlValues, located: dict[str, dict[str, int]] | None = None) -> None:
        self.cfg = cfg
        # located: {Input_label: {label_row,label_col,value_row,value_col}}，来自 core_toml.verify_toml
        self.located = dict(located) if located else {}


    def _worksheet_name(self, workbook_path: Path) -> str:
        """解析 cfg.work_sheet 或回退 active sheet。"""
        wb = load_workbook(workbook_path)
        try:
            if self.cfg.work_sheet and self.cfg.work_sheet in wb.sheetnames:
                return self.cfg.work_sheet
            return wb.active.title
        finally:
            wb.close()


    def _value_cell(self, label: str, instance_k: int) -> tuple[int, int] | None:
        """
        函数名: _value_cell
        作用: 由 located 的 instance 0 值格，按 input_section 平移得第 k 组值格坐标
        输入:
            label (str) - Input_label
            instance_k (int) - 组序（0 即 instance 0，不平移）
        输出:
            tuple[int, int] | None - (row, col)；label 不在 located 时 None
        """
        coord = self.located.get(label)
        if coord is None:
            return None
        value_row = coord["value_row"]
        value_col = coord["value_col"]
        if instance_k <= 0:
            return value_row, value_col
        section = self.cfg.input_section
        # 同方向平移 offset*k；offset_cell 与 core_toml 共用语义
        return offset_cell(value_row, value_col, section.move_to, section.offset * instance_k)


    def _read_instance(self, ws_data: Any, ws_form: Any | None, instance_k: int) -> tuple[dict[str, Any], dict[str, bool]]:
        """读取第 k 组全部 located 值格，键为 Input_label。返回 (值字典, 公式掩码字典)"""
        values: dict[str, Any] = {}
        masks: dict[str, bool] = {}
        try:
            for label in self.located:
                cell = self._value_cell(label, instance_k)
                if cell is None:
                    continue
                # If out of bounds, openpyxl cell() might still work but it's safe to check.
                if cell[0] > 1048576 or cell[1] > 16384:
                    raise ValueError("Cell out of bounds")
                
                val = ws_data.cell(row=cell[0], column=cell[1]).value
                values[label] = val
                
                if ws_form:
                    form_val = ws_form.cell(row=cell[0], column=cell[1]).value
                    masks[label] = isinstance(form_val, str) and form_val.startswith("=")
                else:
                    masks[label] = False
        except ValueError:
            # Propagate up to stop reading when we hit sheet bounds via offset_cell
            raise
            
        return values, masks


    def read_values(self, excel_path: Path, instance_k: int = 0) -> tuple[dict[str, Any], dict[str, bool]]:
        """
        函数名: read_values
        作用: 读取单个 instance 的填写值格（键为 Input_label）及公式掩码
        输入:
            excel_path (Path) - 模板 xlsx
            instance_k (int) - 组序
        输出:
            tuple - (值字典, 公式掩码字典)
        """
        sheet_name = self._worksheet_name(excel_path)
        wb_data = load_workbook(excel_path, data_only=True)
        wb_form = load_workbook(excel_path, data_only=False)
        try:
            return self._read_instance(wb_data[sheet_name], wb_form[sheet_name], instance_k)
        finally:
            wb_data.close()
            wb_form.close()


    def get_total_instance_count(self, excel_path: Path) -> int:
        """
        函数名: get_total_instance_count
        作用: 使用二分查找快速找到文件中非空 instance 的总数 (O(log N))
        输入:
            excel_path (Path) - 模板 xlsx
        输出:
            int - 非空 instance 数量
        """
        if not self.located:
            return 0
        sheet_name = self._worksheet_name(excel_path)
        wb_data = load_workbook(excel_path, data_only=True)
        wb_form = load_workbook(excel_path, data_only=False)
        try:
            ws_data = wb_data[sheet_name]
            ws_form = wb_form[sheet_name]
            
            def is_empty_instance(k: int) -> bool:
                try:
                    values, mask = self._read_instance(ws_data, ws_form, k)
                    for label, v in values.items():
                        if not _cell_empty(v) or mask.get(label):
                            return False
                    return True
                except ValueError:
                    return True
            
            # Find upper bound using max_row/max_column
            max_r = ws_data.max_row
            max_c = ws_data.max_column
            offset_val = self.cfg.input_section.offset or 1
            if self.cfg.input_section.move_to in ["down", "up"]:
                high = max_r // offset_val + 2
            else:
                high = max_c // offset_val + 2
                
            low = 0
            ans = 0
            while low <= high:
                mid = (low + high) // 2
                if not is_empty_instance(mid):
                    ans = mid + 1
                    low = mid + 1
                else:
                    high = mid - 1
            return ans
        finally:
            wb_data.close()
            wb_form.close()

    def read_instances(
        self, excel_path: Path, limit: int | None = None, offset_k: int | None = None, reverse: bool = True
    ) -> tuple[list[dict[str, Any]], list[dict[str, bool]]]:
        """
        函数名: read_instances
        作用: 读取指定范围内的 instances。若不传 limit/offset，则从底部读回所有非空行。
        输入:
            excel_path (Path) - 模板 xlsx
            limit (int | None) - 读取条数
            offset_k (int | None) - 起始 instance_k（reverse=True时表示向下的上限，为None则从最新开始）
            reverse (bool) - 是否从大 k 倒序读取（默认倒序）
        输出:
            tuple - (值字典列表, 公式掩码字典列表)，值字典带有 "instance_k" 键
        """
        if not self.located:
            return [], []
            
        total_count = self.get_total_instance_count(excel_path)
        if total_count == 0:
            return [], []
            
        sheet_name = self._worksheet_name(excel_path)
        wb_data = load_workbook(excel_path, data_only=True)
        wb_form = load_workbook(excel_path, data_only=False)
        try:
            ws_data = wb_data[sheet_name]
            ws_form = wb_form[sheet_name]
            instances: list[dict[str, Any]] = []
            masks: list[dict[str, bool]] = []
            
            start_k = offset_k if offset_k is not None else total_count - 1
            if start_k >= total_count:
                start_k = total_count - 1
                
            if reverse:
                k_range = range(start_k, max(-1, start_k - (limit if limit else total_count)), -1)
            else:
                k_range = range(start_k, min(total_count, start_k + (limit if limit else total_count)))
                
            for k in k_range:
                try:
                    values, mask = self._read_instance(ws_data, ws_form, k)
                    values["instance_k"] = k
                    instances.append(values)
                    masks.append(mask)
                except ValueError:
                    break
            return instances, masks
        finally:
            wb_data.close()
            wb_form.close()


    def max_instance_count(self, excel_path: Path) -> int:
        """
        函数名: max_instance_count
        作用: 以 input_area 为第一块，按 move_to/offset 平移，统计与第一块 cell.value 完全一致的最大块数（含 instance 0）
        输入:
            excel_path (Path) - 模板 xlsx
        输出:
            int - 允许的最大 instance 数（含 instance 0），上界 16384
        """
        from openpyxl.utils import range_boundaries  # 仅此方法用到，局部导入避免改动顶层 import
        BOUND = 16384  # 行列统一边界
        section = self.cfg.input_section
        min_col, min_row, max_col, max_row = range_boundaries(section.input_area)
        sheet_name = self._worksheet_name(excel_path)
        # data_only=False 让公式格呈现 "=..." 文本，便于排除
        wb = load_workbook(excel_path, data_only=False)
        try:
            ws = wb[sheet_name]
            # 读第一块各相对坐标的值；"=" 开头的公式格不参与比较
            base: dict[tuple[int, int], Any] = {}
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    value = ws.cell(row=r, column=c).value
                    if isinstance(value, str) and value.startswith("="):
                        continue
                    base[(r - min_row, c - min_col)] = value
            count = 1  # instance 0 自身计入
            for k in range(1, BOUND):
                shift = section.offset * k
                try:
                    # 整块四角同向平移；越过第 1 行/列时 offset_cell 抛错
                    k_min_row, k_min_col = offset_cell(min_row, min_col, section.move_to, shift)
                    k_max_row, k_max_col = offset_cell(max_row, max_col, section.move_to, shift)
                except ValueError:
                    break
                if k_max_row > BOUND or k_max_col > BOUND:
                    break  # 越过 16384 边界
                matched = True
                for rel, base_value in base.items():
                    cell_value = ws.cell(row=k_min_row + rel[0], column=k_min_col + rel[1]).value
                    if cell_value != base_value:
                        matched = False
                        break
                if not matched:
                    break  # 与第一块不一致，到此为止
                count += 1
            return count
        finally:
            wb.close()


    def write_back(
        self,
        excel_path: Path,
        output_path: Path,
        records: list[dict[str, Any]] | dict[str, Any],
        instance_k: int = 0,
    ) -> None:
        """
        函数名: write_back
        作用: 把记录各 Input_label 值写入对应 located 值格（k 组平移）并另存
        输入:
            excel_path (Path) - 源模板
            output_path (Path) - 输出路径
            records - 单条 dict（写第 instance_k 组）或多条 list（从 instance_k 起依次写）
            instance_k (int) - 起始组序
        输出: 无
        """
        if isinstance(records, dict):
            record_list = [records]
        else:
            record_list = list(records)
        if not self.located:
            raise ValueError("located is empty; run verify_toml first")
        sheet_name = self._worksheet_name(excel_path)
        wb = load_workbook(excel_path)  # data_only=False 默认保留公式
        try:
            ws = wb[sheet_name]
            # 第 i 条记录写入第 instance_k+i 组值格
            for offset_idx, record in enumerate(record_list):
                k = record.get("instance_k", instance_k + offset_idx)
                for label in self.located:
                    if label not in record:
                        continue
                    value = record[label]
                    # 空值不覆盖模板既有内容
                    if _cell_empty(value):
                        continue
                    cell = self._value_cell(label, k)
                    if cell is None:
                        continue
                    
                    # 公式格保护：即使有非空输入也不覆盖公式
                    existing = ws.cell(row=cell[0], column=cell[1]).value
                    if isinstance(existing, str) and existing.startswith("="):
                        continue
                        
                    ws.cell(row=cell[0], column=cell[1]).value = value
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
        finally:
            wb.close()


    def get_print_areas(self, excel_path: Path) -> list[dict[str, Any]]:
        """
        函数名: get_print_areas
        作用: 读取 print_sheet 上全部 print_area，按单元格内容去重
        输入:
            excel_path (Path) - xlsx
        输出:
            list[dict] - 唯一区域；键 label / area / sheet / content_key / cells
        """
        # 非 read_only：ReadOnlyWorksheet 无 print_area
        wb = load_workbook(excel_path, data_only=True)
        try:
            sheet_name: str | None = None
            if self.cfg.print_sheet and self.cfg.print_sheet in wb.sheetnames:
                sheet_name = self.cfg.print_sheet
            elif self.cfg.work_sheet and self.cfg.work_sheet in wb.sheetnames:
                sheet_name = self.cfg.work_sheet
            elif wb.active:
                sheet_name = wb.active.title
            if not sheet_name or sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            raw = ws.print_area
            if raw is None:
                return []
            ranges = _split_print_area_raw(str(raw), sheet_name)
            seen_keys: set[tuple[str, ...]] = set()
            unique: list[dict[str, Any]] = []
            for range_sheet, area in ranges:
                if range_sheet not in wb.sheetnames:
                    continue
                range_ws = wb[range_sheet]
                cells = _read_area_cell_grid(range_ws, area)
                content_key = _area_content_key(cells)
                if content_key in seen_keys:
                    continue
                seen_keys.add(content_key)
                unique.append(
                    {
                        "label": _area_display_label(range_sheet, area, cells),
                        "area": area,
                        "sheet": range_sheet,
                        "content_key": content_key,
                        "cells": cells,
                    }
                )
            return unique
        finally:
            wb.close()


    def render_print_area_image(
        self,
        excel_path: Path,
        sheet_name: str,
        area: str,
        scale: float = 2.5,
    ) -> Image.Image:
        """
        函数名: render_print_area_image
        作用: 将单个打印区渲染为内存 PIL 图像（不写磁盘）
        输入:
            excel_path (Path) - xlsx
            sheet_name (str) - 工作表名
            area (str) - A1 区域
            scale (float) - 像素缩放（越大越清晰）
        输出:
            Image.Image - RGB 图像
        """
        wb = load_workbook(excel_path, data_only=True)
        try:
            ws = wb[sheet_name]
            min_col, min_row, max_col, max_row = range_boundaries(area)
            col_widths = [_column_pixel_width(ws, c, scale) for c in range(min_col, max_col + 1)]
            row_heights = [_row_pixel_height(ws, r, scale) for r in range(min_row, max_row + 1)]
            img_w = sum(col_widths) + 2
            img_h = sum(row_heights) + 2
            image = Image.new("RGB", (img_w, img_h), "white")
            draw = ImageDraw.Draw(image)
            try:
                font = ImageFont.truetype("arial.ttf", max(10, int(11 * scale)))
            except OSError:
                font = ImageFont.load_default()
            y_cursor = 1
            for row_offset, row_idx in enumerate(range(min_row, max_row + 1)):
                x_cursor = 1
                row_h = row_heights[row_offset]
                for col_offset, col_idx in enumerate(range(min_col, max_col + 1)):
                    col_w = col_widths[col_offset]
                    x0, y0 = x_cursor, y_cursor
                    x1, y1 = x_cursor + col_w, y_cursor + row_h
                    draw.rectangle((x0, y0, x1, y1), outline="#cccccc", fill="white")
                    text = _cell_display_text(ws.cell(row=row_idx, column=col_idx).value)
                    if text:
                        draw.text((x0 + 4, y0 + 2), text, fill="black", font=font)
                    x_cursor += col_w
                y_cursor += row_h
            return image
        finally:
            wb.close()


    def render_print_area_png_bytes(
        self,
        excel_path: Path,
        sheet_name: str,
        area: str,
        scale: float = 2.5,
    ) -> bytes:
        """render_print_area_image 的 PNG 字节流（内存，不落盘）。"""
        image = self.render_print_area_image(excel_path, sheet_name, area, scale=scale)
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()




def main() -> None:
    """
    函数名: _demo_main
    作用: 蓝本 Phase 6：用 docs 样本做全量三段验证（路径 A/B + ExcelWriter）
    输入: 无（可选 argparse 覆盖默认样本路径）
    输出: 无（stdout）
    """
    # 模块级不 import core_store，避免与蓝本依赖方向冲突；仅 __main__ 串联
    import tomlkit
    from .core_registry import PROJECT_ROOT
    from .core_store import SecureSQLite, UiProvider, default_db_path
    from .core_toml import _config_from_dict, verify_toml
    docs_dir = PROJECT_ROOT / "docs" / "sample"
    default_toml = docs_dir / "sample_template.toml"
    default_excel = docs_dir / "sample_template.xlsx"
    default_source_xlsx = docs_dir / "执法堂业绩.xlsx"
    default_template_id = "sample_template"
    default_source_id = "8129"
    default_second_id = "250"
    default_textbox = (
        "8129\tClark Kent\t"
        "recreate himself without changing his dob on records on 1978/02/29"
    )
    parser = argparse.ArgumentParser(description="Data Sheet Core sample verification")
    parser.add_argument("--template-id", default=default_template_id, help="DB basename / template id")
    parser.add_argument("--toml", type=Path, default=default_toml, help="TOML config path")
    parser.add_argument("--excel", type=Path, default=default_excel, help="Input_sheet template xlsx")
    parser.add_argument("--textbox", type=str, default=default_textbox, help="path A: tab-separated string")
    parser.add_argument("--source-id", type=str, default=default_source_id, help="path B: external row ID")
    parser.add_argument("--output", type=Path, default=None, help="write-back xlsx; default exports/{template_id}/sample_template_demo_out.xlsx")
    args = parser.parse_args()
    toml_path = Path(args.toml)
    excel_path = Path(args.excel)
    out_path = args.output or (PROJECT_ROOT / "exports" / args.template_id / "sample_template_demo_out.xlsx")
    if not toml_path.is_file():
        raise SystemExit(f"TOML not found: {toml_path}")
    if not excel_path.is_file():
        raise SystemExit(f"template xlsx not found: {excel_path}")
    try:
        raw = tomlkit.loads(toml_path.read_text(encoding="utf-8"))
    except tomlkit.exceptions.ParseError as exc:
        raise SystemExit(f"TOML parse error in {toml_path}: {exc}") from exc
    if not isinstance(raw, dict):
        raise SystemExit(f"TOML root must be a table: {toml_path}")
    cfg = _config_from_dict(raw)
    if cfg is None:
        raise SystemExit(f"TOML has no valid field rules: {toml_path}")
    # TOML 里可能是本机绝对路径；样本库回退到 docs/执法堂业绩.xlsx
    for item in cfg.sources:
        for key, value in item.items():
            if value is None or str(value).strip() == "":
                continue
            candidate = Path(str(value))
            if not candidate.is_file() and default_source_xlsx.is_file():
                item[key] = str(default_source_xlsx)
    db = SecureSQLite(default_db_path(args.template_id))
    ui = UiProvider(cfg, db)
    t2db = Template2DB(cfg)
    # 唯一校验入口：得 located（标签→值格坐标），供 ExcelWriter 按值格读写
    verify_report = verify_toml(excel_path, cfg)
    located = verify_report.get("located", {})
    writer = ExcelWriter(cfg, located=located)
    read_payload: dict[str, Any] = {
        "toml": str(toml_path),
        "excel": str(excel_path),
        "source_xlsx": str(t2db.resolve_source_path("source1") or ""),
        "verify_toml": verify_report,
    }
    try:
        # 路径 A：textbox 按 determiner 拆分为局部 incoming（仅展示，落库见下方覆盖）
        read_payload["textbox_raw"] = args.textbox
        read_payload["textbox_split"] = ui.split_by_determiner(args.textbox)
        read_payload["textbox_incoming"] = ui.record_from_textbox(args.textbox)
        # Excel 读：按 located + input_section 平移逐组读填写值格
        excel_instances, excel_masks = writer.read_instances(excel_path)
        read_payload["excel_instances"] = excel_instances
        try:
            read_payload["print_areas"] = writer.get_print_areas(excel_path)
        except AttributeError as exc:
            read_payload["print_areas_error"] = str(exc)
        # 路径 B + Excel 组：拼成 incoming 后各写一次（store 按 TOML 覆盖，不 merge）
        source_primary = t2db.fetch_row_by_id(args.source_id)
        read_payload["source_incoming_primary"] = source_primary
        excel_row_primary = excel_instances[0] if excel_instances else {}
        incoming_primary = {**source_primary, **excel_row_primary}
        read_payload["persist_incoming_primary"] = incoming_primary
        rid_primary = ui.persist_fields(incoming_primary)
        write_records: list[dict[str, Any]] = []
        row_primary = db.query_by_id(rid_primary)
        if row_primary is not None:
            write_records.append(row_primary)
        # 第二条记录：演示第 k≥1 组值格平移写回
        source_second = t2db.fetch_row_by_id(default_second_id)
        read_payload["source_incoming_second"] = source_second
        excel_row_second = excel_instances[1] if len(excel_instances) > 1 else {}
        incoming_second = {**source_second, **excel_row_second}
        read_payload["persist_incoming_second"] = incoming_second
        rid_second = ui.persist_fields(incoming_second)
        row_second = db.query_by_id(rid_second)
        if row_second is not None:
            write_records.append(row_second)
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # write_records[i] 写入第 i 组值格（instance_k 从 0 起）
        writer.write_back(excel_path, out_path, write_records, instance_k=0)
        read_payload["output_excel"] = str(out_path)
        print("=== 1. 从 Excel / 数据源读取的数据 ===")
        print(json.dumps(read_payload, ensure_ascii=False, indent=2, default=str))
        print("=== 2. 写入 DB 后的数据 ===")
        print(json.dumps(db.query_all(), ensure_ascii=False, indent=2, default=str))
        print("=== 3. Gradio 可获得的数据 ===")
        print("labels:", json.dumps(ui.get_labels(), ensure_ascii=False))
        print("data:", json.dumps(ui.get_data(), ensure_ascii=False, indent=2, default=str))
    finally:
        db.close()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
